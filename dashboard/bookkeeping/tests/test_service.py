from datetime import date
from decimal import Decimal
from pathlib import Path
import sys
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import dashboard.bookkeeping.service as bookkeeping_service
import dashboard.bookkeeping.routes as bookkeeping_routes
from dashboard.bookkeeping.service import (
    build_bookkeeping_workbook,
    build_receipt_organization,
    build_sheet_views,
    build_workspace_revision_snapshot,
    build_revenue_item_payloads,
    build_workspace_summary,
    detect_source_from_headers,
    summarize_revenue_rows,
    update_receipt_organization,
)


class DummyExpenseItem(SimpleNamespace):
    def effective_total(self):
        return float(self.total or self.amount or 0)


def test_receipt_organization_uses_service_date_for_operations():
    period = SimpleNamespace(period_start=date(2026, 8, 1))
    structured = {
        "document_type": "payment_app_screenshot",
        "overall_confidence": 0.96,
        "shared_fields": {"payment_date": "2026-07-23"},
        "entries": [
            {
                "category": "cleaning",
                "item_name": "Cleaning",
                "service_date": "2026-07-21",
                "payment_date": "2026-07-23",
            }
        ],
    }

    organization = build_receipt_organization(structured, "IMG_4421.jpeg", period)

    assert organization["status"] == "suggested"
    assert organization["date_source"] == "service_date"
    assert organization["suggested_filename"] == "July 21, 2026 - Cleaning.jpeg"
    assert organization["relative_path"] == (
        "2026/7. July 2026/Cleanings, Maintenance & Misc. Receipts/"
        "July 21, 2026 - Cleaning.jpeg"
    )


def test_receipt_organization_follows_ticket_purchase_naming_order():
    period = SimpleNamespace(period_start=date(2026, 8, 1))
    structured = {
        "document_type": "retail_receipt",
        "overall_confidence": 0.93,
        "shared_fields": {"store_name": "Amazon", "service_date": "2026-08-09"},
        "entries": [
            {
                "category": "supplies",
                "item_name": "Replacement linens",
                "service_date": "2026-08-09",
                "store_name": "Amazon",
            }
        ],
    }

    organization = build_receipt_organization(structured, "Screenshot.png", period)

    assert organization["document_group"] == "purchase"
    assert organization["suggested_filename"] == "August 9, 2026 - Purchase Receipt - Amazon.png"
    assert organization["folder_name"] == "Purchase & Reimbursement Receipts"


def test_receipt_organization_requires_service_date_confirmation_for_operations():
    period = SimpleNamespace(period_start=date(2026, 7, 1))
    structured = {
        "document_type": "payment_app_screenshot",
        "shared_fields": {"payment_date": "2026-07-23"},
        "entries": [{"category": "maintenance", "item_name": "Maintenance"}],
    }

    organization = build_receipt_organization(structured, "payment.jpg", period)

    assert organization["status"] == "needs_review"
    assert organization["date_source"] == "payment_date"
    assert organization["missing_fields"] == ["service date"]


def test_human_receipt_review_builds_an_approved_filename():
    period = SimpleNamespace(period_start=date(2026, 8, 1))

    organization = update_receipt_organization(
        {},
        {
            "receipt_date": "2026-09-21",
            "document_group": "reimbursement",
            "store_name": "Walmart",
            "status": "approved",
        },
        "IMG_0991.JPG",
        period,
        approved_by=2,
    )

    assert organization["status"] == "approved"
    assert organization["approved_filename"] == "September 21, 2026 - Reimbursement Receipt - Walmart.jpg"
    assert organization["approved_by"] == 2
    assert organization["relative_path"].startswith("2026/9. September 2026/")


def test_receipt_reprocessing_preserves_human_reviewed_name_and_context():
    period = SimpleNamespace(period_start=date(2026, 8, 1))
    reviewed = update_receipt_organization(
        {},
        {
            "receipt_date": "2026-07-21",
            "document_group": "operations",
            "expense_type": "Deep Cleaning",
            "filename": "July 21, 2026 - Deep Cleaning - PT300.jpg",
            "status": "suggested",
        },
        "payment.jpg",
        period,
    )
    changed_extraction = {
        "document_type": "payment_app_screenshot",
        "shared_fields": {"payment_date": "2026-07-23"},
        "entries": [{"category": "maintenance", "service_date": "2026-07-23"}],
    }

    organization = build_receipt_organization(
        changed_extraction,
        "payment.jpg",
        period,
        existing=reviewed,
    )

    assert organization["status"] == "suggested"
    assert organization["receipt_date"] == "2026-07-21"
    assert organization["expense_type"] == "Deep Cleaning"
    assert organization["effective_filename"] == "July 21, 2026 - Deep Cleaning - PT300.jpg"


def test_drive_sync_uses_reviewed_receipt_name_and_folder_hierarchy(monkeypatch, tmp_path):
    source_file = tmp_path / "receipt.jpg"
    source_file.write_bytes(b"receipt-bytes")
    folder_calls = []
    created_files = []

    class FakeRequest:
        def __init__(self, payload):
            self.payload = payload

        def execute(self):
            return self.payload

    class FakeFiles:
        def create(self, **kwargs):
            created_files.append(kwargs)
            return FakeRequest(
                {
                    "id": "receipt-file-id",
                    "name": kwargs["body"]["name"],
                    "webViewLink": "https://drive.google.com/file/d/receipt-file-id/view",
                }
            )

        def update(self, **kwargs):
            raise AssertionError(f"Unexpected update: {kwargs}")

    class FakeDriveService:
        def files(self):
            return FakeFiles()

    def fake_ensure_folder(_service, parent_id, folder_name):
        folder_calls.append((parent_id, folder_name))
        return {"id": f"folder-{len(folder_calls)}", "name": folder_name}

    monkeypatch.setattr(bookkeeping_service.config, "GOOGLE_DRIVE_BOOKKEEPING_ROOT_FOLDER_ID", "readonly-test-root")
    monkeypatch.setattr(bookkeeping_service, "get_bookkeeping_google_drive_service", lambda **_: (FakeDriveService(), {"mode": "test", "google_email": None}))
    monkeypatch.setattr(bookkeeping_service, "get_upload_absolute_path", lambda _: source_file)
    monkeypatch.setattr(bookkeeping_service, "_ensure_google_drive_folder", fake_ensure_folder)
    monkeypatch.setattr(bookkeeping_service, "_list_google_drive_children", lambda *_: {})
    monkeypatch.setattr(bookkeeping_service, "MediaFileUpload", lambda *_, **__: object())

    upload = SimpleNamespace(
        stage="expense",
        stored_path="unused",
        original_filename="IMG_4421.jpg",
        content_type="image/jpeg",
        summary={
            "receipt_organization": {
                "status": "approved",
                "effective_filename": "July 21, 2026 - Cleaning.jpg",
                "folder_name": "Cleanings, Maintenance & Misc. Receipts",
                "year_folder": "2026",
                "month_folder": "7. July 2026",
            }
        },
    )
    portfolio = SimpleNamespace(name="PT300", code="PT300")
    period = SimpleNamespace(period_start=date(2026, 8, 1), name="August 2026")

    result = bookkeeping_service.sync_bookkeeping_uploads_to_google_drive(portfolio, period, [upload])

    assert folder_calls == [
        ("readonly-test-root", "PT300"),
        ("folder-1", "2026"),
        ("folder-2", "7. July 2026"),
        ("folder-3", "Cleanings, Maintenance & Misc. Receipts"),
    ]
    assert created_files[0]["body"] == {
        "name": "July 21, 2026 - Cleaning.jpg",
        "parents": ["folder-4"],
    }
    assert upload.summary["drive_sync"]["drive_filename"] == "July 21, 2026 - Cleaning.jpg"
    assert result["folder_path"] == "PT300/2026/7. July 2026"


def test_expense_item_from_payload_preserves_upload_id_when_edit_payload_omits_it():
    item = SimpleNamespace(category="supplies", upload_id=88, created_by=7)

    bookkeeping_routes._expense_item_from_payload(
        item,
        {
            "category": "supplies",
            "item_name": "Replacement sheets",
            "amount": 42.50,
        },
        created_by=7,
    )

    assert item.upload_id == 88


def test_expense_item_from_payload_allows_explicit_upload_unlink():
    item = SimpleNamespace(category="supplies", upload_id=88, created_by=7)

    bookkeeping_routes._expense_item_from_payload(
        item,
        {
            "category": "supplies",
            "upload_id": "",
        },
        created_by=7,
    )

    assert item.upload_id is None


def test_detect_source_from_headers_prefers_known_signatures():
    headers = ["Reservation number", "Arrival", "Departure", "Commission %", "Property name"]
    assert detect_source_from_headers(headers) == "booking_com"


def test_summarize_revenue_rows_ignores_total_rows():
    rows = [
        {
            "Reservation ID": 1,
            "Check-In Date": "2026-01-08",
            "Check-Out Date": "2026-01-10",
            "Property": "PT300-15J-IG",
            "Guest name": "Mike Lorii",
            "Total Price": 338.55,
            "Hostaway application fee": 3.39,
            "Stripe processing fees": 10.12,
            "Vrbo Commission": 15.3,
        },
        {
            "Reservation ID": 2,
            "Check-In Date": "2026-01-10",
            "Check-Out Date": "2026-01-11",
            "Property": "PT300-10i-IG",
            "Guest name": "Joseph Watts",
            "Total Price": 205.77,
            "Hostaway application fee": 2.06,
            "Stripe processing fees": 6.27,
            "Vrbo Commission": 9.3,
        },
        {
            "Reservation ID": None,
            "Property": None,
            "Guest name": None,
            "Total Price": 544.32,
            "Hostaway application fee": 5.45,
            "Stripe processing fees": 16.39,
            "Vrbo Commission": 24.6,
        },
    ]

    summary = summarize_revenue_rows("vrbo", rows)

    assert summary["record_count"] == 2
    assert summary["gross_total"] == 544.32
    assert summary["hostaway_fee_total"] == 5.45
    assert summary["stripe_fee_total"] == 16.39
    assert summary["commission_total"] == 24.6
    assert len(summary["processing_fee_rows"]) == 2


def test_extract_expense_evidence_bundle_retries_dense_retail_receipt_after_timeout(monkeypatch):
    calls = []

    def fake_request_openai_json(*, system_prompt, user_content, timeout_seconds):
        calls.append(
            {
                "timeout_seconds": timeout_seconds,
                "prompt": user_content[0]["text"],
            }
        )
        if len(calls) == 1:
            raise RuntimeError("Request timed out.")
        return {
            "document_type": "retail_receipt",
            "overall_confidence": 0.91,
            "support_only": False,
            "explicit_guest_refund_language": False,
            "guest_refund_phrase": None,
            "shared_fields": {"store_name": "dd's DISCOUNTS", "payment_method": "card"},
            "reimbursement_proof": {},
            "entries": [
                {
                    "category": "supplies",
                    "confidence": 0.91,
                    "item_name": "Battery pack",
                    "vendor": "dd's DISCOUNTS",
                    "property_code": None,
                    "scope": "portfolio",
                    "amount": 9.99,
                    "total": 9.99,
                    "service_date": None,
                    "payment_date": None,
                    "payment_method": "card",
                    "account_holder": None,
                    "account_number": "0043",
                    "purchase_type": "retail",
                    "store_name": "dd's DISCOUNTS",
                    "quantity": 1,
                    "unit_amount": 9.99,
                    "subtotal": None,
                    "discount": None,
                    "shipping": None,
                    "tax": None,
                    "reimbursement_method": None,
                    "reimbursement_date": None,
                    "details": "Battery pack",
                    "review_reason": None,
                }
            ],
        }

    monkeypatch.setattr(bookkeeping_service, "_request_openai_json", fake_request_openai_json)

    result = bookkeeping_service.extract_expense_evidence_bundle(
        file_bytes=b"fake-image-bytes",
        filename="March 17, 2026 - Purchase Receipt - dd's Discount.jpeg",
        mime_type="image/jpeg",
        parsed_summary={"preview_text": None, "page_count": None},
        property_alias_map=None,
    )

    assert result is not None
    assert result["document_type"] == "retail_receipt"
    assert len(result["entries"]) == 1
    assert [call["timeout_seconds"] for call in calls] == [
        bookkeeping_service.OPENAI_VISION_TIMEOUT_SECONDS,
        max(
            bookkeeping_service.OPENAI_VISION_SLOW_FALLBACK_TIMEOUT_SECONDS,
            bookkeeping_service.OPENAI_VISION_TIMEOUT_SECONDS,
        ),
    ]
    assert "Fallback mode: this is likely a dense itemized retail receipt." in calls[1]["prompt"]


def test_extract_expense_evidence_bundle_keeps_non_receipt_timeouts_fast(monkeypatch):
    calls = []

    def fake_request_openai_json(*, system_prompt, user_content, timeout_seconds):
        calls.append(timeout_seconds)
        raise RuntimeError("Request timed out.")

    monkeypatch.setattr(bookkeeping_service, "_request_openai_json", fake_request_openai_json)

    with pytest.raises(RuntimeError, match="Request timed out"):
        bookkeeping_service.extract_expense_evidence_bundle(
            file_bytes=b"fake-image-bytes",
            filename="March 12, 2026 - PT300 Bedsheets.jpeg",
            mime_type="image/jpeg",
            parsed_summary={"preview_text": None, "page_count": None},
            property_alias_map=None,
        )

    assert calls == [bookkeeping_service.OPENAI_VISION_TIMEOUT_SECONDS]


def test_build_sheet_views_shows_configured_software_rows_in_live_tab():
    portfolio = SimpleNamespace(
        code="PT300",
        listing_count=16,
        hostaway_price_per_listing=Decimal("46.32"),
        pricelabs_price_per_listing=Decimal("10.01"),
        management_fee_percentage=Decimal("0.00"),
        revenue_channels=[],
    )

    sheet_views = build_sheet_views(
        portfolio,
        period=SimpleNamespace(),
        uploads=[],
        revenue_items=[],
        expense_items=[],
    )

    software_sheet = next(sheet for sheet in sheet_views if sheet["key"] == "expense_software_fee")

    assert software_sheet["editable"] is False
    assert [row["software_name"] for row in software_sheet["rows"]] == ["Hostaway", "Pricelabs"]
    assert [row["effective_total"] for row in software_sheet["rows"]] == [741.12, 160.16]
    assert [row["entry_type"] for row in software_sheet["rows"]] == ["Configured", "Configured"]


def test_build_sheet_views_keeps_manual_software_rows_editable():
    portfolio = SimpleNamespace(
        code="PT300",
        listing_count=16,
        hostaway_price_per_listing=Decimal("46.32"),
        pricelabs_price_per_listing=Decimal("10.01"),
        management_fee_percentage=Decimal("0.00"),
        revenue_channels=[],
    )
    manual_item = DummyExpenseItem(
        bookkeeping_expense_item_id=55,
        category="software_fee",
        item_name="Owner portal add-on",
        vendor="Hostfully",
        property_code="PT300",
        payment_method="Card",
        total=25.0,
        amount=25.0,
        updated_at=None,
        service_date=None,
        payment_date=None,
        needs_review=False,
    )

    sheet_views = build_sheet_views(
        portfolio,
        period=SimpleNamespace(),
        uploads=[],
        revenue_items=[],
        expense_items=[manual_item],
    )

    software_sheet = next(sheet for sheet in sheet_views if sheet["key"] == "expense_software_fee")
    manual_row = software_sheet["rows"][-1]

    assert software_sheet["editable"] is True
    assert [row["software_name"] for row in software_sheet["rows"][:2]] == ["Hostaway", "Pricelabs"]
    assert manual_row["row_id"] == 55
    assert manual_row["row_type"] == "expense_item"
    assert manual_row["entry_type"] == "Manual"
    assert manual_row["effective_total"] == 25.0


def test_live_expense_sheet_exposes_bulk_edit_fields_for_visible_columns():
    portfolio = SimpleNamespace(
        code="PT300",
        listing_count=0,
        hostaway_price_per_listing=Decimal("0"),
        pricelabs_price_per_listing=Decimal("0"),
        management_fee_percentage=Decimal("0"),
        revenue_channels=[],
    )

    sheet_views = build_sheet_views(
        portfolio,
        period=SimpleNamespace(),
        uploads=[],
        revenue_items=[],
        expense_items=[],
    )

    expense_sheet = next(sheet for sheet in sheet_views if sheet["key"] == "expenses_all")
    edit_fields = {column["key"]: column.get("edit_field") for column in expense_sheet["columns"]}

    assert edit_fields["service_date"] == "service_date"
    assert edit_fields["category"] == "category"
    assert edit_fields["effective_total"] == "total"
    assert edit_fields["needs_review"] == "needs_review"


def test_revenue_source_headers_map_to_canonical_bulk_edit_fields():
    assert bookkeeping_service.revenue_edit_field_for_header("airbnb", "Guest") == "guest_name"
    assert bookkeeping_service.revenue_edit_field_for_header("vrbo", "Total Price") == "gross_amount"
    assert bookkeeping_service.revenue_edit_field_for_header("booking_com", "Commission %") is None


def test_bulk_edit_payload_changes_only_requested_field():
    item = SimpleNamespace(
        to_dict=lambda: {
            "guest_name": "Original Guest",
            "property_code": "PT300-2i",
            "gross_amount": 65.0,
        }
    )

    payload = bookkeeping_routes._bulk_edit_item_payload(
        "revenue_item",
        item,
        "property_code",
        "PT300-6N",
    )

    assert payload == {
        "guest_name": "Original Guest",
        "property_code": "PT300-6N",
        "gross_amount": 65.0,
    }


def test_bulk_edit_payload_rejects_non_bulk_field():
    item = SimpleNamespace(to_dict=lambda: {})

    with pytest.raises(ValueError, match="cannot be bulk edited"):
        bookkeeping_routes._bulk_edit_item_payload("revenue_item", item, "source", "vrbo")


def test_build_bookkeeping_workbook_creates_owner_statement_and_tabs():
    portfolio = SimpleNamespace(
        code="PT300",
        name="Peachtree Towers",
        property_name="Peachtree Towers",
        property_address="300 Peachtree Street, NE, Atlanta, GA 30308",
        owner_share_percentage=Decimal("100.00"),
        management_fee_percentage=Decimal("20.00"),
        listing_count=16,
        hostaway_price_per_listing=Decimal("47.66"),
        pricelabs_price_per_listing=Decimal("8.767"),
    )
    period = SimpleNamespace(
        period_start=__import__("datetime").date(2026, 1, 1),
        name="January 2026",
    )
    uploads = [
        SimpleNamespace(
            bookkeeping_upload_id=101,
            stage="revenue",
            source="booking_com",
            headers=["Reservation number", "Final amount", "Commission amount", "Property name"],
            parsed_rows=[
                {"Reservation number": 1, "Final amount": 219.0, "Commission amount": 32.85, "Property name": "PT300-10M-KW"},
                {"Reservation number": 2, "Final amount": 280.0, "Commission amount": 42.0, "Property name": "PT300-2M-KW"},
            ],
            summary={"gross_total": 499.0, "commission_total": 74.85, "hostaway_fee_total": 0.0, "stripe_fee_total": 0.0, "processing_fee_rows": []},
        ),
        SimpleNamespace(
            bookkeeping_upload_id=102,
            stage="revenue",
            source="direct_bookings",
            headers=["Reservation ID", "Guest name", "Property", "Total Price", "Hostaway application fee", "Stripe processing fees"],
            parsed_rows=[
                {
                    "Reservation ID": 10,
                    "Guest name": "Alana Nolen",
                    "Property": "PT300-20M-KW",
                    "Total Price": 488.64,
                    "Hostaway application fee": 4.89,
                    "Stripe processing fees": 14.17,
                }
            ],
            summary={
                "gross_total": 488.64,
                "commission_total": 0.0,
                "hostaway_fee_total": 4.89,
                "stripe_fee_total": 14.17,
                "processing_fee_rows": [
                    {
                        "guest": "Alana Nolen",
                        "booking_platform": "Direct Bookings",
                        "listing": "PT300-20M-KW",
                        "stripe": 14.17,
                        "hostaway": 4.89,
                    }
                ],
            },
        ),
        SimpleNamespace(
            bookkeeping_upload_id=500,
            stage="expense",
            source="expense_evidence",
            summary={
                "drive_sync": {
                    "file_url": "https://drive.google.com/file/d/test-cleaning-receipt/view?usp=drive_link",
                },
            },
        ),
    ]
    expense_items = [
        DummyExpenseItem(category="cleaning", total=65.0, amount=65.0, item_name="Cleaning", vendor="Cecillia", property_code="PT300-15K-IG", service_date=None, payment_method="Zelle", account_holder="Alma Ramirez", account_number="4042", payment_date=None, upload_id=500, details=None, description=None, needs_review=False, review_reason=None),
        DummyExpenseItem(category="software_fee", total=25.0, amount=25.0, item_name="Extra Software", vendor="Vendor", property_code="PT300", service_date=None, payment_method="Card", account_holder=None, account_number=None, payment_date=None, upload_id=None, details=None, description=None),
    ]

    workbook_bytes = build_bookkeeping_workbook(portfolio, period, uploads, expense_items)
    workbook = load_workbook(filename=__import__("io").BytesIO(workbook_bytes), data_only=True)

    assert "Owner Statement (January 2026)" in workbook.sheetnames
    assert "Official Booking.com Report for"[:31] in workbook.sheetnames
    assert "Official Direct Bookings Report"[:31] in workbook.sheetnames

    owner_sheet = workbook["Owner Statement (January 2026)"]
    assert owner_sheet["L12"].value == 499
    assert owner_sheet["L15"].value == 488.64
    assert owner_sheet["L22"].value == 65
    assert owner_sheet["L28"].value == 74.85


def test_build_bookkeeping_workbook_matches_pt300_owner_statement_layout_and_receipt_links():
    portfolio = SimpleNamespace(
        code="PT300",
        name="Peachtree Towers",
        property_name="Peachtree Towers",
        property_address="300 Peachtree Street, NE, Atlanta, GA 30308",
        owner_share_percentage=Decimal("100.00"),
        management_fee_percentage=Decimal("20.00"),
        listing_count=16,
        hostaway_price_per_listing=Decimal("47.66"),
        pricelabs_price_per_listing=Decimal("8.767"),
    )
    period = SimpleNamespace(
        period_start=__import__("datetime").date(2026, 1, 1),
        name="January 2026",
    )
    uploads = [
        SimpleNamespace(
            bookkeeping_upload_id=201,
            stage="expense",
            source="expense_evidence",
            summary={
                "drive_sync": {
                    "file_url": "https://drive.google.com/file/d/receipt-201/view?usp=drive_link",
                },
            },
        ),
    ]
    expense_items = [
        DummyExpenseItem(
            category="cleaning",
            total=65.0,
            amount=65.0,
            item_name="Cleaning",
            vendor="Cecillia",
            property_code="PT300-15K-IG",
            service_date=None,
            payment_method="Zelle",
            account_holder="Alma Ramirez",
            account_number="4042",
            payment_date=None,
            upload_id=201,
            details=None,
            description=None,
            needs_review=False,
            review_reason=None,
        ),
    ]

    workbook_bytes = build_bookkeeping_workbook(portfolio, period, uploads, expense_items, revenue_items=[])
    workbook = load_workbook(filename=__import__("io").BytesIO(workbook_bytes))

    owner_sheet = workbook["Owner Statement (January 2026)"]
    merged_ranges = {str(cell_range) for cell_range in owner_sheet.merged_cells.ranges}
    assert "A1:N2" in merged_ranges
    assert "B3:M3" in merged_ranges
    assert "B6:E8" in merged_ranges
    assert "G6:M8" in merged_ranges
    assert "B11:K11" in merged_ranges
    assert "L11:M11" in merged_ranges
    assert "B32:K32" in merged_ranges
    assert "L34:M34" in merged_ranges
    assert owner_sheet["A1"].value == "OWNER STATEMENT (PT300)"
    assert owner_sheet["B3"].value == "For January 2026"
    assert owner_sheet["B5"].value == "PROPERTY NAME"
    assert owner_sheet["G5"].value == "PROPERTY ADDRESS"
    assert owner_sheet["L19"].value == "=SUM(L11:M17)"
    assert owner_sheet["L34"].value == "=L19-L32"
    assert owner_sheet["A1"].fill.fgColor.rgb == "FF000000"

    cleaning_sheet = workbook["Cleaning Expenses"]
    assert cleaning_sheet["J2"].value == "Click/Tap Here"
    assert cleaning_sheet["J2"].hyperlink is not None
    assert cleaning_sheet["J2"].hyperlink.target == "https://drive.google.com/file/d/receipt-201/view?usp=drive_link"


def test_build_revenue_item_payloads_uses_listing_mapping_aliases():
    mapping = SimpleNamespace(
        bookkeeping_listing_mapping_id=7,
        official_name="MIDDLEFORK-RIDGE",
        listing_name="Middlefork Ridge",
        internal_listing_name="MF Ridge",
        aliases=["The Ridge"],
        is_active=True,
        listing_id=101,
    )
    upload = SimpleNamespace(
        stage="revenue",
        source="booking_com",
        headers=["Reservation number", "Arrival", "Departure", "Booker name", "Final amount", "Commission amount", "Property name"],
        parsed_rows=[
            {
                "Reservation number": "ABC-1",
                "Arrival": "2026-01-02",
                "Departure": "2026-01-04",
                "Booker name": "Jane Doe",
                "Final amount": "420.00",
                "Commission amount": "63.00",
                "Property name": "The Ridge",
            }
        ],
        summary={"gross_total": 420.0, "commission_total": 63.0},
        sheet_name="Booking.com",
        stored_path="unused",
        original_filename="booking.csv",
    )

    payloads = build_revenue_item_payloads(
        upload,
        property_alias_map={"THERIDGE": "MIDDLEFORK-RIDGE"},
        listing_lookup={"alias_to_mapping": {"THERIDGE": mapping}},
    )

    assert len(payloads) == 1
    assert payloads[0]["property_code"] == "MIDDLEFORK-RIDGE"
    assert payloads[0]["listing_mapping_id"] == 7
    assert payloads[0]["normalized_data"]["Property name"] == "MIDDLEFORK-RIDGE"


def test_build_workspace_summary_exposes_sheet_views_and_revenue_items():
    portfolio = SimpleNamespace(
        bookkeeping_portfolio_id=1,
        code="PT300",
        name="Peachtree Towers",
        property_name="Peachtree Towers",
        property_address="Atlanta",
        owner_share_percentage=Decimal("100.00"),
        management_fee_percentage=Decimal("20.00"),
        listing_count=2,
        hostaway_price_per_listing=Decimal("47.66"),
        pricelabs_price_per_listing=Decimal("8.77"),
        to_dict=lambda: {"bookkeeping_portfolio_id": 1, "code": "PT300", "name": "Peachtree Towers", "property_name": "Peachtree Towers", "property_address": "Atlanta"},
    )
    period = SimpleNamespace(
        bookkeeping_period_id=1,
        name="January 2026",
        period_start=__import__("datetime").date(2026, 1, 1),
        period_end=__import__("datetime").date(2026, 1, 31),
        status="draft",
        to_dict=lambda: {"bookkeeping_period_id": 1, "name": "January 2026", "status": "draft", "period_start": "2026-01-01", "period_end": "2026-01-31"},
    )
    revenue_item = SimpleNamespace(
        bookkeeping_revenue_item_id=11,
        source="direct_bookings",
        row_index=0,
        reservation_identifier="R-1",
        confirmation_code=None,
        guest_name="Alana Nolen",
        property_code="PT300-20M-KW",
        raw_listing_name="PT300-20M-KW",
        transaction_type=None,
        currency="USD",
        transaction_date=None,
        booking_date=None,
        start_date=__import__("datetime").date(2026, 1, 3),
        end_date=__import__("datetime").date(2026, 1, 5),
        nights=2,
        gross_amount=Decimal("488.64"),
        paid_out_amount=None,
        commission_amount=Decimal("0"),
        hostaway_fee_amount=Decimal("4.89"),
        stripe_fee_amount=Decimal("14.17"),
        cleaning_fee_amount=None,
        tax_amount=None,
        refund_amount=None,
        details=None,
        normalized_data={"Property": "PT300-20M-KW", "Guest name": "Alana Nolen", "Total Price": 488.64},
        raw_data={"Property": "PT300-20M-KW", "Guest name": "Alana Nolen", "Total Price": 488.64},
        needs_review=False,
        review_reason=None,
        updated_at=__import__("datetime").datetime(2026, 1, 10, 12, 0, 0),
        to_dict=lambda: {
            "bookkeeping_revenue_item_id": 11,
            "source": "direct_bookings",
            "guest_name": "Alana Nolen",
            "property_code": "PT300-20M-KW",
            "gross_amount": 488.64,
            "hostaway_fee_amount": 4.89,
            "stripe_fee_amount": 14.17,
            "normalized_data": {"Property": "PT300-20M-KW", "Guest name": "Alana Nolen", "Total Price": 488.64},
            "updated_at": "2026-01-10T12:00:00",
            "needs_review": False,
        },
    )
    expense_item = DummyExpenseItem(
        bookkeeping_expense_item_id=21,
        category="cleaning",
        total=65.0,
        amount=65.0,
        item_name="Cleaning",
        vendor="Cecillia",
        property_code="PT300-15K-IG",
        service_date=None,
        payment_method="Zelle",
        account_holder="Alma Ramirez",
        account_number="4042",
        payment_date=None,
        upload_id=None,
        details=None,
        description=None,
        needs_review=False,
        review_reason=None,
        updated_at=__import__("datetime").datetime(2026, 1, 11, 9, 0, 0),
    )

    summary = build_workspace_summary(
        portfolio,
        period,
        uploads=[],
        expense_items=[expense_item],
        revenue_items=[revenue_item],
        listing_mappings=[],
        change_proposals=[],
        revisions=[],
    )

    assert summary["summary_cards"]["owner_revenue"] == 488.64
    assert summary["summary_cards"]["owner_expenses"] > 65
    assert any(sheet["key"] == "owner_statement" for sheet in summary["sheet_views"])
    assert any(sheet["key"] == "revenue_all" for sheet in summary["sheet_views"])


def test_build_workspace_summary_uses_portfolio_revenue_channels_for_coverage():
    portfolio = SimpleNamespace(
        bookkeeping_portfolio_id=1,
        code="PT300",
        name="PT300",
        property_name="PT300",
        property_address="Atlanta",
        revenue_channels=["airbnb", "vrbo"],
        owner_share_percentage=Decimal("100.00"),
        management_fee_percentage=Decimal("20.00"),
        listing_count=2,
        hostaway_price_per_listing=Decimal("47.66"),
        pricelabs_price_per_listing=Decimal("8.77"),
        to_dict=lambda: {
            "bookkeeping_portfolio_id": 1,
            "code": "PT300",
            "name": "PT300",
            "property_name": "PT300",
            "property_address": "Atlanta",
            "revenue_channels": ["airbnb", "vrbo"],
        },
    )
    period = SimpleNamespace(
        bookkeeping_period_id=1,
        name="January 2026",
        period_start=__import__("datetime").date(2026, 1, 1),
        period_end=__import__("datetime").date(2026, 1, 31),
        status="draft",
        to_dict=lambda: {"bookkeeping_period_id": 1, "name": "January 2026", "status": "draft", "period_start": "2026-01-01", "period_end": "2026-01-31"},
    )
    airbnb_upload = SimpleNamespace(
        stage="revenue",
        source="airbnb",
        headers=[],
        parsed_rows=[],
        summary={"gross_total": 1200.0, "commission_total": 0.0, "hostaway_fee_total": 0.0, "stripe_fee_total": 0.0, "processing_fee_rows": []},
    )

    summary = build_workspace_summary(
        portfolio,
        period,
        uploads=[airbnb_upload],
        expense_items=[],
        revenue_items=[],
        listing_mappings=[],
        change_proposals=[],
        revisions=[],
    )

    assert summary["revenue_progress"] == {"completed": 1, "total": 2, "missing": 1}
    checklist_by_source = {entry["source"]: entry for entry in summary["revenue_checklist"]}
    assert checklist_by_source["airbnb"]["status"] == "uploaded"
    assert checklist_by_source["vrbo"]["status"] == "missing"
    assert "booking_com" not in checklist_by_source


def test_build_workspace_revision_snapshot_stays_compact():
    portfolio = SimpleNamespace(
        bookkeeping_portfolio_id=1,
        code="PT300",
        name="PT300",
        listing_tag="pt300",
        property_name="PT300",
        property_address="Atlanta",
        revenue_channels=["airbnb", "vrbo"],
        management_fee_percentage=Decimal("20.00"),
        listing_count=2,
        hostaway_price_per_listing=Decimal("47.66"),
        pricelabs_price_per_listing=Decimal("8.77"),
        to_dict=lambda: {
            "bookkeeping_portfolio_id": 1,
            "code": "PT300",
            "name": "PT300",
            "listing_tag": "pt300",
            "portfolio_tag": "pt300",
            "property_name": "PT300",
            "property_address": "Atlanta",
            "revenue_channels": ["airbnb", "vrbo"],
            "listing_count": 2,
            "management_fee_percentage": 20.0,
        },
    )
    period = SimpleNamespace(
        bookkeeping_period_id=2,
        name="January 2026",
        status="approved",
        period_start=__import__("datetime").date(2026, 1, 1),
        period_end=__import__("datetime").date(2026, 1, 31),
        to_dict=lambda: {
            "bookkeeping_period_id": 2,
            "name": "January 2026",
            "status": "approved",
            "period_start": "2026-01-01",
            "period_end": "2026-01-31",
        },
    )
    uploads = [
        SimpleNamespace(
            stage="expense",
            source="expense_evidence",
            upload_status="processed",
            original_filename="receipt-1.pdf",
        ),
        SimpleNamespace(
            stage="revenue",
            source="airbnb",
            upload_status="processed",
            original_filename="airbnb.csv",
            summary={
                "gross_total": 1200.0,
                "commission_total": 0.0,
                "hostaway_fee_total": 0.0,
                "stripe_fee_total": 0.0,
                "processing_fee_rows": [],
            },
        ),
    ]
    revenue_item = SimpleNamespace(
        source="airbnb",
        gross_amount=Decimal("1200.00"),
        commission_amount=Decimal("0.00"),
        hostaway_fee_amount=Decimal("0.00"),
        stripe_fee_amount=Decimal("0.00"),
        needs_review=False,
    )
    expense_item = DummyExpenseItem(
        category="cleaning",
        total=125.0,
        amount=125.0,
        needs_review=True,
    )
    processing_batch = SimpleNamespace(
        bookkeeping_processing_batch_id=9,
        stage="expense",
        status="completed",
        total_uploads=2,
        processed_uploads=2,
        successful_uploads=2,
        failed_uploads=0,
        created_at=__import__("datetime").datetime(2026, 1, 31, 16, 0, 0),
    )

    snapshot = build_workspace_revision_snapshot(
        portfolio,
        period,
        uploads=uploads,
        expense_items=[expense_item],
        revenue_items=[revenue_item],
        listing_mappings=[],
        change_proposals=[],
        processing_batches=[processing_batch],
    )

    assert "sheet_views" not in snapshot
    assert "uploads" not in snapshot
    assert "expense_items" not in snapshot
    assert "revenue_items" not in snapshot
    assert snapshot["summary_cards"]["owner_revenue"] == 1200.0
    assert snapshot["summary_cards"]["owner_expenses"] > 125.0
    assert snapshot["upload_counts"]["expense"] == 1
    assert snapshot["upload_counts"]["revenue"] == 1
