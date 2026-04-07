#!/usr/bin/env python3
"""
Bookkeeping parsing, summarization, and export services.
"""

import base64
import csv
import mimetypes
import io
import json
import logging
import os
import re
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from functools import lru_cache
from itertools import combinations
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image

import dashboard.config as config
from dashboard.auth.models import get_google_drive_credential_for_user, save_google_drive_credential
from dashboard.bookkeeping.models import (
    DEFAULT_REVENUE_CHANNELS,
    EXPENSE_CATEGORIES,
    REVENUE_SOURCES,
    SPECIAL_UPLOAD_SOURCES,
    normalize_revenue_channels,
)

try:
    from google.auth.transport.requests import Request as GoogleAuthRequest
    from google.oauth2 import credentials as user_credentials
    from google.oauth2 import service_account
    from googleapiclient.discovery import build as build_google_api
    from googleapiclient.http import MediaFileUpload
except ImportError:  # pragma: no cover - optional until Drive sync is configured
    GoogleAuthRequest = None
    user_credentials = None
    service_account = None
    build_google_api = None
    MediaFileUpload = None


BOOKKEEPING_FILES_DIR = Path(config.PROJECT_ROOT) / "data" / "bookkeeping"
AUTO_CATEGORIZE_CONFIDENCE_THRESHOLD = 0.88
OPENAI_VISION_TIMEOUT_SECONDS = float(os.getenv("BOOKKEEPING_OPENAI_VISION_TIMEOUT_SECONDS", "90"))
OPENAI_VISION_SLOW_FALLBACK_TIMEOUT_SECONDS = float(
    os.getenv("BOOKKEEPING_OPENAI_VISION_SLOW_FALLBACK_TIMEOUT_SECONDS", "180")
)
OPENAI_VISION_MAX_RETRIES = int(os.getenv("BOOKKEEPING_OPENAI_VISION_MAX_RETRIES", "0"))
REVIEW_ROW_FILL_COLOR = "FFF2CC"
WORKSPACE_PREVIEW_TEXT_LIMIT = 600
VISION_IMAGE_MAX_DIMENSION = 1600
VISION_IMAGE_JPEG_QUALITY = 82
SUPPORTED_VISION_MIME_TYPES = {
    "image/png": ".png",
    "image/jpeg": ".jpg",
    "image/jfif": ".jpg",
    "image/pjpeg": ".jpg",
    "image/pipeg": ".jpg",
    "image/webp": ".webp",
}
IMAGE_EXTENSION_TO_MIME = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".jfif": "image/jpeg",
    ".webp": "image/webp",
}
RETAIL_RECEIPT_FILENAME_HINTS = (
    "receipt",
    "purchase",
    "invoice",
    "order",
    "dollar tree",
    "sam's club",
    "sams club",
    "home depot",
    "target",
    "walmart",
    "costco",
    "ross",
    "discount",
)
PAYMENT_APP_FILENAME_HINTS = (
    "venmo",
    "zelle",
    "cash app",
    "cashapp",
    "paypal",
    "apple cash",
)
GOOGLE_DRIVE_FOLDER_MIME_TYPE = "application/vnd.google-apps.folder"
GOOGLE_DRIVE_BOOKKEEPING_SCOPES = ("https://www.googleapis.com/auth/drive",)
GOOGLE_DRIVE_EVIDENCE_SUBFOLDER = "evidences"
RECEIPT_LINK_TEXT = "Click/Tap Here"
OWNER_STATEMENT_HEADER_FILL = "FF000000"
OWNER_STATEMENT_BAND_FILL = "FFF3F3F3"
THIN_BLACK_SIDE = Side(style="thin", color="FF000000")
THIN_BLACK_BORDER = Border(
    left=THIN_BLACK_SIDE,
    right=THIN_BLACK_SIDE,
    top=THIN_BLACK_SIDE,
    bottom=THIN_BLACK_SIDE,
)
OWNER_STATEMENT_COLUMN_WIDTHS = {
    "A": 3.38,
    "B": 3.88,
    "C": 8.13,
    "D": 13.0,
    "E": 10.13,
    "F": 1.63,
    "G": 8.13,
    "H": 1.63,
    "I": 9.88,
    "J": 13.0,
    "K": 1.63,
    "L": 9.88,
    "M": 13.0,
    "N": 3.38,
}
OWNER_STATEMENT_LABEL_NUMBER_FORMAT = '[$$]#,##0.00;-[$$]#,##0.00;;'
OWNER_STATEMENT_VALUE_NUMBER_FORMAT = '"$"#,##0.00'
REVENUE_SOURCE_LABELS = {
    "airbnb": "Airbnb",
    "booking_com": "Booking.com",
    "vrbo": "Vrbo",
    "hopper": "Hopper Homes",
    "direct_bookings": "Direct Bookings",
    "google": "Google",
    "direct_refund": "Direct Refund",
    "expense_evidence": "Expense Evidence",
    "bank_statement": "Bank Statement",
    "credit_card_statement": "Credit Card Statement",
    "stripe_statement": "Stripe Statement",
}

SHEET_LABEL_TEMPLATES = {
    "airbnb": "Offical Airbnb Report for {month_label}",
    "booking_com": "Official Booking.com Report for {month_label}",
    "vrbo": "Official Vrbo Report for {month_label}",
    "hopper": "Official Hopper Homes Report for {month_label}",
    "direct_bookings": "Official Direct Bookings Report for {month_label}",
    "google": "Official Google Report for {month_label}",
    "direct_refund": "Direct Refund",
}

SOURCE_SHEET_KEYWORDS = [
    ("direct_refund", ("direct refund",)),
    ("direct_bookings", ("direct booking", "direct bookings", "direct csv", "direct")),
    ("booking_com", ("booking.com", "official booking.com")),
    ("airbnb", ("airbnb",)),
    ("vrbo", ("vrbo",)),
    ("hopper", ("hopper",)),
    ("google", ("google",)),
]

SOURCE_HEADER_SIGNATURES = {
    "airbnb": {"confirmation code", "gross earnings", "service fee"},
    "booking_com": {"reservation number", "commission %", "property name"},
    "vrbo": {"vrbo commission", "hostaway application fee", "stripe processing fees"},
    "hopper": {"hopper homes commission", "listing number"},
    "direct_bookings": {"guest channel fee", "hostaway application fee", "stripe processing fees"},
    "google": {"guest channel fee", "hostaway application fee", "stripe processing fees"},
    "direct_refund": {"platform", "type", "refund receipt"},
}

PROCESSING_FEE_SOURCE_HEADERS = {
    "vrbo": {
        "guest": "Guest name",
        "listing": "Property",
        "stripe": "Stripe processing fees",
        "hostaway": "Hostaway application fee",
    },
    "direct_bookings": {
        "guest": "Guest name",
        "listing": "Property",
        "stripe": "Stripe processing fees",
        "hostaway": "Hostaway application fee",
    },
    "google": {
        "guest": "Guest name",
        "listing": "Property",
        "stripe": "Stripe processing fees",
        "hostaway": "Hostaway application fee",
    },
}

CSV_FALLBACK_HEADERS = {
    "vrbo": [
        "Reservation ID",
        "Check-In Date",
        "Check-Out Date",
        "Property",
        "Listing Number",
        "Guest name",
        "Base Rate",
        "Discount",
        "Cleaning Fee",
        "Lodging Tax",
        "Refund",
        "Total Price",
        "Hostaway application fee",
        "Stripe processing fees",
        "Vrbo Commission",
    ],
    "hopper": [
        "Reservation ID",
        "Check-In Date",
        "Check-Out Date",
        "Property",
        "Listing Number",
        "Guest name",
        "Base Rate",
        "Discount",
        "Cleaning Fee",
        "Lodging Tax",
        "Hopper Homes Commission",
        "Total Price",
    ],
    "direct_bookings": [
        "Reservation ID",
        "Check-In Date",
        "Check-Out Date",
        "Property",
        "Listing Number",
        "Guest name",
        "Base Rate",
        "Discount",
        "Cleaning Fee",
        "Lodging Tax",
        "Guest Channel Fee",
        "Total Price",
        "Hostaway application fee",
        "Stripe processing fees",
    ],
}

PROPERTY_CODE_PATTERN = re.compile(r"\b([A-Z0-9]+-\d+[A-Z0-9]-[A-Z]{2})\b", re.IGNORECASE)
PAYMENT_MEMO_TOKEN_PATTERN = re.compile(r"\b[A-Z0-9]{2,4}\b", re.IGNORECASE)
MASKED_ACCOUNT_NUMBER_PATTERN = re.compile(r"^(?:X+|[*xX]+)?\d{2,4}$")
MAINTENANCE_MEMO_KEYWORDS = (
    "unclog",
    "repair",
    "replace",
    "fix",
    "toilet",
    "plumbing",
    "leak",
    "maintenance",
)
PT300_CLEANING_RATE_BY_UNIT = {
    "10H": 65.0,
    "10I": 65.0,
    "10M": 60.0,
    "15I": 65.0,
    "15J": 65.0,
    "15K": 65.0,
    "18C": 65.0,
    "19M": 60.0,
    "20M": 60.0,
    "21M": 60.0,
    "23J": 65.0,
    "2I": 65.0,
    "2M": 60.0,
    "3E": 65.0,
    "6N": 65.0,
    "8H": 65.0,
}
PT300_FULL_CODE_BY_UNIT = {
    "10H": "PT300-10H-IG",
    "10I": "PT300-10I-IG",
    "10M": "PT300-10M-KW",
    "15I": "PT300-15I-IG",
    "15J": "PT300-15J-IG",
    "15K": "PT300-15K-IG",
    "18C": "PT300-18C-KW",
    "19M": "PT300-19M-KW",
    "20M": "PT300-20M-KW",
    "21M": "PT300-21M-KW",
    "23J": "PT300-23J-KW",
    "2I": "PT300-2I-IG",
    "2M": "PT300-2M-KW",
    "3E": "PT300-3E-IG",
    "6N": "PT300-6N-IG",
    "8H": "PT300-8H-IG",
}
logger = logging.getLogger(__name__)
STANDARD_REVENUE_HEADERS = {
    "airbnb": [
        "Date",
        "Arriving by date",
        "Type",
        "Confirmation code",
        "Booking date",
        "Start date",
        "End date",
        "Nights",
        "Guest",
        "Listing",
        "Details",
        "Reference code",
        "Currency",
        "Amount",
        "Paid out",
        "Service fee",
        "Fast pay fee",
        "Cleaning fee",
        "Gross earnings",
        "Occupancy taxes",
        "Earnings year",
    ],
    "booking_com": [
        "Reservation number",
        "Arrival",
        "Departure",
        "Booker name",
        "Commission %",
        "Original amount",
        "Final amount",
        "Commission amount",
        "Status",
        "Property name",
    ],
    "vrbo": CSV_FALLBACK_HEADERS["vrbo"],
    "hopper": CSV_FALLBACK_HEADERS["hopper"],
    "direct_bookings": CSV_FALLBACK_HEADERS["direct_bookings"],
    "google": [
        "Reservation ID",
        "Check-In Date",
        "Check-Out Date",
        "Property",
        "Listing Number",
        "Guest name",
        "Base Rate",
        "Cleaning Fee",
        "Lodging Tax",
        "Guest Channel Fee",
        "Total Price",
        "Hostaway application fee",
        "Stripe processing fees",
    ],
}

REVENUE_ROW_FIELD_MAP = {
    "reservation_identifier": {
        "booking_com": "Reservation number",
        "vrbo": "Reservation ID",
        "hopper": "Reservation ID",
        "direct_bookings": "Reservation ID",
        "google": "Reservation ID",
    },
    "confirmation_code": {
        "airbnb": "Confirmation code",
    },
    "guest_name": {
        "airbnb": "Guest",
        "booking_com": "Booker name",
        "vrbo": "Guest name",
        "hopper": "Guest name",
        "direct_bookings": "Guest name",
        "google": "Guest name",
    },
    "property_code": {
        "airbnb": "Listing",
        "booking_com": "Property name",
        "vrbo": "Property",
        "hopper": "Property",
        "direct_bookings": "Property",
        "google": "Property",
        "direct_refund": "Listing",
    },
    "transaction_type": {
        "airbnb": "Type",
        "direct_refund": "Type",
    },
    "transaction_date": {
        "airbnb": "Date",
        "direct_refund": "Date",
    },
    "booking_date": {
        "airbnb": "Booking date",
    },
    "start_date": {
        "airbnb": "Start date",
        "booking_com": "Arrival",
        "vrbo": "Check-In Date",
        "hopper": "Check-In Date",
        "direct_bookings": "Check-In Date",
        "google": "Check-In Date",
    },
    "end_date": {
        "airbnb": "End date",
        "booking_com": "Departure",
        "vrbo": "Check-Out Date",
        "hopper": "Check-Out Date",
        "direct_bookings": "Check-Out Date",
        "google": "Check-Out Date",
    },
    "nights": {
        "airbnb": "Nights",
    },
    "currency": {
        "airbnb": "Currency",
    },
    "gross_amount": {
        "airbnb": "Amount",
        "booking_com": "Final amount",
        "vrbo": "Total Price",
        "hopper": "Total Price",
        "direct_bookings": "Total Price",
        "google": "Total Price",
        "direct_refund": "Amount",
    },
    "paid_out_amount": {
        "airbnb": "Paid out",
    },
    "commission_amount": {
        "booking_com": "Commission amount",
        "vrbo": "Vrbo Commission",
        "hopper": "Hopper Homes Commission",
    },
    "hostaway_fee_amount": {
        "vrbo": "Hostaway application fee",
        "direct_bookings": "Hostaway application fee",
        "google": "Hostaway application fee",
    },
    "stripe_fee_amount": {
        "vrbo": "Stripe processing fees",
        "direct_bookings": "Stripe processing fees",
        "google": "Stripe processing fees",
    },
    "cleaning_fee_amount": {
        "airbnb": "Cleaning fee",
        "vrbo": "Cleaning Fee",
        "hopper": "Cleaning Fee",
        "direct_bookings": "Cleaning Fee",
        "google": "Cleaning Fee",
    },
    "tax_amount": {
        "airbnb": "Occupancy taxes",
        "vrbo": "Lodging Tax",
        "hopper": "Lodging Tax",
        "direct_bookings": "Lodging Tax",
        "google": "Lodging Tax",
    },
    "refund_amount": {
        "vrbo": "Refund",
    },
    "details": {
        "airbnb": "Details",
        "direct_refund": "Details",
    },
}

REVENUE_COMMON_GRID_COLUMNS = [
    ("source", "Source"),
    ("reservation_identifier", "Reservation"),
    ("guest_name", "Guest"),
    ("property_code", "Property"),
    ("start_date", "Start"),
    ("end_date", "End"),
    ("gross_amount", "Gross"),
    ("commission_amount", "Commission"),
    ("hostaway_fee_amount", "Hostaway"),
    ("stripe_fee_amount", "Stripe"),
    ("needs_review", "Review"),
]
REVENUE_DATE_HEADERS = {
    "Date",
    "Arriving by date",
    "Booking date",
    "Start date",
    "End date",
    "Arrival",
    "Departure",
    "Check-In Date",
    "Check-Out Date",
}
REVENUE_PERCENT_HEADERS = {"Commission %"}
REVENUE_INTEGER_HEADERS = {"Reservation ID", "Reservation number", "Listing Number", "Nights", "Earnings year"}
REVENUE_MONEY_HEADERS = {
    "Amount",
    "Paid out",
    "Service fee",
    "Fast pay fee",
    "Cleaning fee",
    "Gross earnings",
    "Occupancy taxes",
    "Original amount",
    "Final amount",
    "Commission amount",
    "Base Rate",
    "Discount",
    "Cleaning Fee",
    "Lodging Tax",
    "Refund",
    "Total Price",
    "Hostaway application fee",
    "Stripe processing fees",
    "Vrbo Commission",
    "Hopper Homes Commission",
    "Guest Channel Fee",
}


def ensure_bookkeeping_storage_dir() -> Path:
    BOOKKEEPING_FILES_DIR.mkdir(parents=True, exist_ok=True)
    return BOOKKEEPING_FILES_DIR


def safe_filename(filename: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", (filename or "").strip())
    cleaned = re.sub(r"-{2,}", "-", cleaned).strip("-")
    return cleaned or "upload"


def logical_month_label(period_start: date) -> str:
    return period_start.strftime("%B %Y")


def normalize_header(header: Any) -> str:
    return str(header or "").strip()


def normalize_property_token(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def is_empty_value(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def json_safe_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def money_value(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip().replace(",", "").replace("$", "")
        if not stripped:
            return 0.0
        try:
            return float(stripped)
        except ValueError:
            return 0.0
    return 0.0


def percentage_value(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float, Decimal)):
        numeric = float(value)
        return numeric / 100.0 if numeric > 1 else numeric
    if isinstance(value, str):
        stripped = value.strip().replace(",", "").replace("%", "")
        if not stripped:
            return 0.0
        try:
            numeric = float(stripped)
            return numeric / 100.0 if numeric > 1 else numeric
        except ValueError:
            return 0.0
    return 0.0


def decimal_or_none(value: Any) -> Optional[Decimal]:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def parse_date_or_none(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d", "%b %d, %Y", "%B %d, %Y"):
            try:
                return datetime.strptime(stripped, fmt).date()
            except ValueError:
                continue
    return None


def source_label(source: str) -> str:
    return REVENUE_SOURCE_LABELS.get(source, source.replace("_", " ").title())


def configured_revenue_sources(portfolio: Any) -> List[str]:
    configured = normalize_revenue_channels(getattr(portfolio, "revenue_channels", None))
    return configured or list(DEFAULT_REVENUE_CHANNELS)


def sheet_title_for_source(source: str, period_start: date) -> str:
    template = SHEET_LABEL_TEMPLATES.get(source, "{month_label}")
    title = template.format(month_label=logical_month_label(period_start))
    return title[:31]


def detect_source_from_sheet_name(sheet_name: str) -> Optional[str]:
    normalized = (sheet_name or "").strip().lower()
    for source, keywords in SOURCE_SHEET_KEYWORDS:
        if any(keyword in normalized for keyword in keywords):
            return source
    return None


def detect_source_from_filename(filename: str) -> Optional[str]:
    stem = Path(filename or "").stem.replace("-", " ").replace("_", " ")
    return detect_source_from_sheet_name(stem)


def detect_source_from_headers(headers: Sequence[str], declared_source: str = "auto") -> Optional[str]:
    if declared_source and declared_source != "auto":
        return declared_source

    normalized_headers = {normalize_header(header).lower() for header in headers if normalize_header(header)}
    if not normalized_headers:
        return None

    for source in ("airbnb", "booking_com", "vrbo", "hopper", "direct_refund", "direct_bookings"):
        signature = SOURCE_HEADER_SIGNATURES[source]
        if signature.issubset(normalized_headers):
            return source

    return None


def save_upload_bytes(period_id: int, stage: str, original_filename: str, file_bytes: bytes) -> Tuple[str, int]:
    storage_root = ensure_bookkeeping_storage_dir()
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
    target_dir = storage_root / str(period_id) / stage
    target_dir.mkdir(parents=True, exist_ok=True)
    safe_name = safe_filename(original_filename)
    relative_path = Path(str(period_id)) / stage / f"{timestamp}-{safe_name}"
    full_path = storage_root / relative_path
    full_path.write_bytes(file_bytes)
    return str(relative_path), len(file_bytes)


def get_upload_absolute_path(stored_path: str) -> Path:
    return ensure_bookkeeping_storage_dir() / stored_path


def read_upload_as_base64(stored_path: str) -> str:
    return base64.b64encode(get_upload_absolute_path(stored_path).read_bytes()).decode("utf-8")


def _drive_sync_payload(upload: Any) -> Dict[str, Any]:
    summary = dict(getattr(upload, "summary", None) or {})
    payload = summary.get("drive_sync")
    return dict(payload) if isinstance(payload, dict) else {}


def _set_drive_sync_payload(upload: Any, payload: Dict[str, Any]) -> None:
    summary = dict(getattr(upload, "summary", None) or {})
    summary["drive_sync"] = payload
    upload.summary = summary


def upload_drive_file_url(upload: Any) -> Optional[str]:
    payload = _drive_sync_payload(upload)
    return _string_or_none(payload.get("file_url"))


def _drive_query_escape(value: str) -> str:
    return (value or "").replace("\\", "\\\\").replace("'", "\\'")


def _sanitize_drive_name(value: Optional[str], fallback: str) -> str:
    cleaned = (value or "").strip().replace("/", "-")
    return cleaned or fallback


@lru_cache(maxsize=1)
def _get_service_account_google_drive_service():
    if service_account is None or build_google_api is None or MediaFileUpload is None:
        raise RuntimeError(
            "Google Drive support is not installed. Install google-api-python-client and google-auth to enable bookkeeping evidence sync."
        )

    if config.GOOGLE_DRIVE_SERVICE_ACCOUNT_JSON:
        service_account_info = json.loads(config.GOOGLE_DRIVE_SERVICE_ACCOUNT_JSON)
        credentials = service_account.Credentials.from_service_account_info(
            service_account_info,
            scopes=list(GOOGLE_DRIVE_BOOKKEEPING_SCOPES),
        )
    elif config.GOOGLE_DRIVE_SERVICE_ACCOUNT_FILE:
        credentials = service_account.Credentials.from_service_account_file(
            config.GOOGLE_DRIVE_SERVICE_ACCOUNT_FILE,
            scopes=list(GOOGLE_DRIVE_BOOKKEEPING_SCOPES),
        )
    else:
        raise RuntimeError(
            "Google Drive bookkeeping sync is not configured. Set GOOGLE_DRIVE_SERVICE_ACCOUNT_FILE or GOOGLE_DRIVE_SERVICE_ACCOUNT_JSON and share the bookkeeping Drive folder with that service account."
        )

    return build_google_api("drive", "v3", credentials=credentials, cache_discovery=False)


def _build_user_google_drive_credentials(user: Any):
    if GoogleAuthRequest is None or user_credentials is None or build_google_api is None or MediaFileUpload is None:
        raise RuntimeError(
            "Google Drive user authorization support is not installed. Install google-api-python-client and google-auth to enable bookkeeping evidence sync."
        )

    if not user or not getattr(user, "user_id", None):
        return None

    stored_credential = get_google_drive_credential_for_user(user.user_id)
    if not stored_credential:
        return None

    if not config.GOOGLE_CLIENT_ID or not config.GOOGLE_CLIENT_SECRET:
        raise RuntimeError("Google OAuth credentials are not configured for Drive access.")

    scopes = stored_credential.scopes() or list(GOOGLE_DRIVE_BOOKKEEPING_SCOPES)
    credentials = user_credentials.Credentials(
        token=stored_credential.access_token,
        refresh_token=stored_credential.refresh_token,
        token_uri=stored_credential.token_uri or "https://oauth2.googleapis.com/token",
        client_id=config.GOOGLE_CLIENT_ID,
        client_secret=config.GOOGLE_CLIENT_SECRET,
        scopes=scopes,
    )
    if stored_credential.expires_at:
        expiry = stored_credential.expires_at
        if expiry.tzinfo is not None:
            expiry = expiry.astimezone(timezone.utc).replace(tzinfo=None)
        credentials.expiry = expiry

    if credentials.expired:
        if not credentials.refresh_token:
            raise RuntimeError("Google Drive authorization expired. Reconnect Drive from Cotton Candy and try again.")
        credentials.refresh(GoogleAuthRequest())
        save_google_drive_credential(
            user.user_id,
            access_token=credentials.token,
            refresh_token=credentials.refresh_token or stored_credential.refresh_token,
            token_uri=credentials.token_uri,
            scopes=credentials.scopes or scopes,
            expires_at=(
                credentials.expiry.astimezone(timezone.utc).replace(tzinfo=None)
                if getattr(credentials.expiry, "tzinfo", None) is not None
                else credentials.expiry
            ) if credentials.expiry else None,
            google_email=stored_credential.google_email,
        )

    return credentials, stored_credential


def get_bookkeeping_google_drive_service(*, user: Any = None):
    user_credentials_bundle = _build_user_google_drive_credentials(user)
    if user_credentials_bundle:
        credentials, stored_credential = user_credentials_bundle
        service = build_google_api("drive", "v3", credentials=credentials, cache_discovery=False)
        return service, {
            "mode": "user_authorized",
            "google_email": getattr(stored_credential, "google_email", None) or getattr(user, "email", None),
        }

    service = _get_service_account_google_drive_service()
    return service, {
        "mode": "service_account",
        "google_email": None,
    }


def _find_google_drive_child(service: Any, parent_id: str, name: str, mime_type: Optional[str] = None) -> Optional[Dict[str, Any]]:
    query_parts = [
        f"'{parent_id}' in parents",
        f"name = '{_drive_query_escape(name)}'",
        "trashed = false",
    ]
    if mime_type:
        query_parts.append(f"mimeType = '{mime_type}'")
    response = service.files().list(
        q=" and ".join(query_parts),
        fields="files(id, name, webViewLink)",
        includeItemsFromAllDrives=True,
        supportsAllDrives=True,
        corpora="allDrives",
        pageSize=10,
    ).execute()
    files = response.get("files") or []
    return files[0] if files else None


def _list_google_drive_children(service: Any, parent_id: str) -> Dict[str, Dict[str, Any]]:
    page_token = None
    children: Dict[str, Dict[str, Any]] = {}
    while True:
        response = service.files().list(
            q=f"'{parent_id}' in parents and trashed = false",
            fields="nextPageToken, files(id, name, webViewLink, mimeType)",
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            corpora="allDrives",
            pageSize=1000,
            pageToken=page_token,
        ).execute()
        for entry in response.get("files") or []:
            name = entry.get("name")
            if name and name not in children:
                children[name] = entry
        page_token = response.get("nextPageToken")
        if not page_token:
            break
    return children


def _ensure_google_drive_folder(service: Any, parent_id: str, folder_name: str) -> Dict[str, Any]:
    existing = _find_google_drive_child(service, parent_id, folder_name, mime_type=GOOGLE_DRIVE_FOLDER_MIME_TYPE)
    if existing:
        return existing
    return service.files().create(
        body={
            "name": folder_name,
            "mimeType": GOOGLE_DRIVE_FOLDER_MIME_TYPE,
            "parents": [parent_id],
        },
        fields="id, name, webViewLink",
        supportsAllDrives=True,
    ).execute()


def sync_bookkeeping_uploads_to_google_drive(
    portfolio: Any,
    period: Any,
    uploads: Sequence[Any],
    *,
    month_label: Optional[str] = None,
    stages: Optional[Sequence[str]] = None,
    user: Any = None,
) -> Optional[Dict[str, Any]]:
    relevant_stages = set(stages or ("expense", "corroboration"))
    relevant_uploads = [upload for upload in uploads if getattr(upload, "stage", None) in relevant_stages]
    if not relevant_uploads:
        return None

    root_folder_id = (config.GOOGLE_DRIVE_BOOKKEEPING_ROOT_FOLDER_ID or "").strip()
    if not root_folder_id:
        raise RuntimeError("GOOGLE_DRIVE_BOOKKEEPING_ROOT_FOLDER_ID is not configured.")

    service, auth_context = get_bookkeeping_google_drive_service(user=user)
    portfolio_folder_name = _sanitize_drive_name(getattr(portfolio, "name", None) or getattr(portfolio, "code", None), "Portfolio")
    month_folder_name = _sanitize_drive_name(month_label or getattr(period, "name", None), "Month")
    portfolio_folder = _ensure_google_drive_folder(service, root_folder_id, portfolio_folder_name)
    month_folder = _ensure_google_drive_folder(service, portfolio_folder["id"], month_folder_name)
    evidence_folder = _ensure_google_drive_folder(service, month_folder["id"], GOOGLE_DRIVE_EVIDENCE_SUBFOLDER)
    evidence_children = _list_google_drive_children(service, evidence_folder["id"])
    synced_at = datetime.utcnow().isoformat()

    for upload in relevant_uploads:
        file_path = get_upload_absolute_path(getattr(upload, "stored_path"))
        if not file_path.exists():
            raise FileNotFoundError(f"Bookkeeping upload file is missing on disk: {file_path}")

        drive_name = getattr(upload, "original_filename", file_path.name)
        existing = evidence_children.get(drive_name)
        prior_payload = _drive_sync_payload(upload)
        if (
            existing
            and prior_payload.get("file_id") == existing.get("id")
            and prior_payload.get("authorization_mode") == auth_context.get("mode")
            and prior_payload.get("evidence_folder_id") == evidence_folder.get("id")
            and prior_payload.get("month_folder_id") == month_folder.get("id")
            and prior_payload.get("portfolio_folder_id") == portfolio_folder.get("id")
            and prior_payload.get("root_folder_id") == root_folder_id
            and prior_payload.get("stored_path") == getattr(upload, "stored_path", None)
            and prior_payload.get("original_filename") == drive_name
        ):
            existing_url = existing.get("webViewLink") or prior_payload.get("file_url") or (
                f"https://drive.google.com/file/d/{existing.get('id')}/view?usp=drive_link" if existing.get("id") else None
            )
            if existing_url != prior_payload.get("file_url"):
                _set_drive_sync_payload(
                    upload,
                    {
                        **prior_payload,
                        "file_url": existing_url,
                        "synced_at": synced_at,
                    },
                )
            continue

        media = MediaFileUpload(
            str(file_path),
            mimetype=getattr(upload, "content_type", None) or mimetypes.guess_type(getattr(upload, "original_filename", ""))[0] or "application/octet-stream",
            resumable=False,
        )
        if existing:
            uploaded = service.files().update(
                fileId=existing["id"],
                media_body=media,
                body={"name": drive_name},
                fields="id, name, webViewLink",
                supportsAllDrives=True,
            ).execute()
        else:
            uploaded = service.files().create(
                body={
                    "name": drive_name,
                    "parents": [evidence_folder["id"]],
                },
                media_body=media,
                fields="id, name, webViewLink",
                supportsAllDrives=True,
            ).execute()

        file_id = uploaded.get("id")
        file_url = uploaded.get("webViewLink") or (f"https://drive.google.com/file/d/{file_id}/view?usp=drive_link" if file_id else None)
        if drive_name:
            evidence_children[drive_name] = uploaded
        _set_drive_sync_payload(
            upload,
            {
                "provider": "google_drive",
                "authorization_mode": auth_context.get("mode"),
                "google_email": auth_context.get("google_email"),
                "root_folder_id": root_folder_id,
                "portfolio_folder_id": portfolio_folder.get("id"),
                "month_folder_id": month_folder.get("id"),
                "evidence_folder_id": evidence_folder.get("id"),
                "folder_path": f"{portfolio_folder_name}/{month_folder_name}/{GOOGLE_DRIVE_EVIDENCE_SUBFOLDER}",
                "stored_path": getattr(upload, "stored_path", None),
                "original_filename": drive_name,
                "file_id": file_id,
                "file_url": file_url,
                "synced_at": synced_at,
            },
        )

    return {
        "portfolio_folder_id": portfolio_folder.get("id"),
        "month_folder_id": month_folder.get("id"),
        "evidence_folder_id": evidence_folder.get("id"),
        "folder_path": f"{portfolio_folder_name}/{month_folder_name}/{GOOGLE_DRIVE_EVIDENCE_SUBFOLDER}",
        "authorization_mode": auth_context.get("mode"),
        "google_email": auth_context.get("google_email"),
    }


def _csv_rows_from_bytes(
    file_bytes: bytes,
    fallback_headers: Optional[Sequence[str]] = None,
) -> Tuple[List[str], List[Dict[str, Any]]]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)
    if not rows:
        return [], []

    first_row = [normalize_header(header) for header in rows[0]]
    use_fallback_headers = False
    if fallback_headers:
        fallback_set = {normalize_header(header).lower() for header in fallback_headers}
        first_row_set = {header.lower() for header in first_row if header}
        use_fallback_headers = bool(first_row) and not fallback_set.intersection(first_row_set)

    headers = [normalize_header(header) for header in (fallback_headers if use_fallback_headers else first_row)]
    data_rows = rows if use_fallback_headers else rows[1:]
    parsed_rows: List[Dict[str, Any]] = []
    for raw_row in data_rows:
        padded = list(raw_row) + [""] * max(0, len(headers) - len(raw_row))
        row_dict = {}
        for index, header in enumerate(headers):
            row_dict[header] = padded[index] if index < len(padded) else ""
        if any(not is_empty_value(value) for value in row_dict.values()):
            parsed_rows.append(row_dict)

    return headers, parsed_rows


def _worksheet_rows(worksheet) -> Tuple[List[str], List[Dict[str, Any]]]:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [normalize_header(cell) for cell in rows[0]]
    parsed_rows: List[Dict[str, Any]] = []
    for raw_row in rows[1:]:
        row_dict = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = raw_row[index] if index < len(raw_row) else None
            row_dict[header] = json_safe_value(value)
        if any(not is_empty_value(value) for value in row_dict.values()):
            parsed_rows.append(row_dict)
    return headers, parsed_rows


def _is_data_row_for_summary(source: str, row: Dict[str, Any]) -> bool:
    if not row:
        return False
    values = [value for value in row.values() if not is_empty_value(value)]
    if not values:
        return False
    if len(values) == 1 and isinstance(values[0], str):
        return False
    if any(str(value).strip().upper() == "TOTAL" for value in values if isinstance(value, str)):
        return False

    source_checks = {
        "airbnb": ("Type", "Guest", "Listing", "Date"),
        "booking_com": ("Reservation number", "Booker name", "Property name"),
        "vrbo": ("Reservation ID", "Guest name", "Property"),
        "hopper": ("Reservation ID", "Guest name", "Property"),
        "direct_bookings": ("Reservation ID", "Guest name", "Property"),
        "google": ("Reservation ID", "Guest name", "Property"),
        "direct_refund": ("Date", "Type", "Listing"),
    }
    required_fields = source_checks.get(source, tuple(row.keys()))
    return any(not is_empty_value(row.get(field)) for field in required_fields)


def build_processing_fee_rows(source: str, rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    mapping = PROCESSING_FEE_SOURCE_HEADERS.get(source)
    if not mapping:
        return []

    results = []
    for row in rows:
        if not _is_data_row_for_summary(source, row):
            continue
        stripe_total = money_value(row.get(mapping["stripe"]))
        hostaway_total = money_value(row.get(mapping["hostaway"]))
        if stripe_total == 0 and hostaway_total == 0:
            continue

        results.append(
            {
                "guest": _string_or_none(row.get(mapping["guest"])),
                "booking_platform": "Direct Booking" if source == "direct_bookings" else source_label(source),
                "listing": _string_or_none(row.get(mapping["listing"])),
                "stripe": stripe_total,
                "hostaway": hostaway_total,
            }
        )
    return results


def summarize_revenue_rows(source: str, rows: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
    data_rows = [row for row in rows if _is_data_row_for_summary(source, row)]
    summary = {
        "source": source,
        "record_count": len(data_rows),
        "gross_total": 0.0,
        "commission_total": 0.0,
        "hostaway_fee_total": 0.0,
        "stripe_fee_total": 0.0,
        "processing_fee_rows": build_processing_fee_rows(source, data_rows),
    }

    gross_column = {
        "airbnb": "Amount",
        "booking_com": "Final amount",
        "vrbo": "Total Price",
        "hopper": "Total Price",
        "direct_bookings": "Total Price",
        "google": "Total Price",
        "direct_refund": "Amount",
    }.get(source, "Amount")

    commission_column = {
        "booking_com": "Commission amount",
        "vrbo": "Vrbo Commission",
        "hopper": "Hopper Homes Commission",
    }.get(source)

    for row in data_rows:
        summary["gross_total"] += money_value(row.get(gross_column))
        if source == "vrbo":
            if not is_empty_value(row.get("Vrbo Commission")):
                summary["commission_total"] += money_value(row.get("Vrbo Commission"))
            else:
                summary["commission_total"] += (money_value(row.get("Base Rate")) + money_value(row.get("Cleaning Fee"))) * 0.05
        elif commission_column:
            summary["commission_total"] += money_value(row.get(commission_column))
        if source in ("vrbo", "direct_bookings", "google"):
            summary["hostaway_fee_total"] += money_value(row.get("Hostaway application fee"))
            summary["stripe_fee_total"] += money_value(row.get("Stripe processing fees"))

    return {
        **summary,
        "gross_total": round(summary["gross_total"], 2),
        "commission_total": round(summary["commission_total"], 2),
        "hostaway_fee_total": round(summary["hostaway_fee_total"], 2),
        "stripe_fee_total": round(summary["stripe_fee_total"], 2),
    }


def parse_revenue_file(file_bytes: bytes, filename: str, declared_source: str = "auto") -> List[Dict[str, Any]]:
    extension = Path(filename).suffix.lower()
    logical_uploads: List[Dict[str, Any]] = []

    if extension == ".csv":
        inferred_source = (
            declared_source if declared_source and declared_source != "auto"
            else detect_source_from_filename(filename)
        )
        headers, rows = _csv_rows_from_bytes(file_bytes, fallback_headers=CSV_FALLBACK_HEADERS.get(inferred_source))
        detected_source = detect_source_from_headers(headers, inferred_source or declared_source)
        detected_source = detected_source or inferred_source or declared_source or "unknown"
        logical_uploads.append(
            {
                "stage": "revenue",
                "source": detected_source,
                "detected_source": detected_source,
                "sheet_name": None,
                "headers": headers,
                "rows": rows,
                "preview_rows": rows[:8],
                "row_count": len(rows),
                "summary": summarize_revenue_rows(detected_source, rows),
            }
        )
        return logical_uploads

    if extension != ".xlsx":
        raise ValueError("Revenue uploads must be CSV or XLSX files.")

    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    for worksheet in workbook.worksheets:
        headers, rows = _worksheet_rows(worksheet)
        if not headers:
            continue
        detected_source = detect_source_from_sheet_name(worksheet.title) or detect_source_from_headers(headers, declared_source)
        if not detected_source:
            continue

        logical_uploads.append(
            {
                "stage": "revenue",
                "source": detected_source,
                "detected_source": detected_source,
                "sheet_name": worksheet.title,
                "headers": headers,
                "rows": rows,
                "preview_rows": rows[:8],
                "row_count": len(rows),
                "summary": summarize_revenue_rows(detected_source, rows),
            }
        )

    if not logical_uploads:
        raise ValueError("No supported revenue sheets were found in the workbook.")

    return logical_uploads


def parse_supporting_file(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    extension = Path(filename).suffix.lower()
    summary = {
        "preview_text": None,
        "page_count": None,
    }

    if extension == ".pdf":
        text_chunks: List[str] = []
        page_count = 0
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            page_count = len(pdf.pages)
            for page in pdf.pages[:3]:
                text = page.extract_text() or ""
                if text:
                    text_chunks.append(text.strip())
        preview_text = "\n".join(text_chunks)[:2000] or None
        summary["page_count"] = page_count
        summary["preview_text"] = preview_text

    return summary


def _normalized_image_mime_type(filename: str, mime_type: Optional[str]) -> Optional[str]:
    if mime_type in SUPPORTED_VISION_MIME_TYPES:
        return mime_type
    return IMAGE_EXTENSION_TO_MIME.get(Path(filename).suffix.lower())


def _optimize_image_bytes_for_vision(
    file_bytes: bytes,
    normalized_mime_type: Optional[str],
) -> Tuple[bytes, Optional[str]]:
    if not file_bytes or not normalized_mime_type or not normalized_mime_type.startswith("image/"):
        return file_bytes, normalized_mime_type

    try:
        with Image.open(io.BytesIO(file_bytes)) as image:
            if getattr(image, "is_animated", False):
                return file_bytes, normalized_mime_type

            original_size = image.size
            if image.mode in ("RGBA", "LA", "P"):
                background = Image.new("RGB", image.size, (255, 255, 255))
                if image.mode == "P":
                    image = image.convert("RGBA")
                background.paste(image, mask=image.split()[-1] if image.mode in ("RGBA", "LA") else None)
                image = background
            elif image.mode != "RGB":
                image = image.convert("RGB")

            if image.width > VISION_IMAGE_MAX_DIMENSION or image.height > VISION_IMAGE_MAX_DIMENSION:
                image.thumbnail((VISION_IMAGE_MAX_DIMENSION, VISION_IMAGE_MAX_DIMENSION), Image.Resampling.LANCZOS)

            output = io.BytesIO()
            image.save(output, "JPEG", quality=VISION_IMAGE_JPEG_QUALITY, optimize=True)
            optimized_bytes = output.getvalue()

            if (
                normalized_mime_type == "image/jpeg"
                and image.size == original_size
                and len(optimized_bytes) >= len(file_bytes)
            ):
                return file_bytes, normalized_mime_type
            return optimized_bytes, "image/jpeg"
    except Exception:
        return file_bytes, normalized_mime_type


def _float_or_none(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _coerce_scope(scope: Any, property_code: Optional[str]) -> str:
    normalized = str(scope or "").strip().lower()
    if normalized in {"property", "portfolio"}:
        return normalized
    return "property" if property_code else "portfolio"


def _string_or_none(value: Any) -> Optional[str]:
    normalized = str(value or "").strip()
    return normalized or None


def _looks_masked_account_number(value: Any) -> bool:
    account_number = _string_or_none(value)
    if not account_number:
        return False
    return bool(MASKED_ACCOUNT_NUMBER_PATTERN.fullmatch(account_number))


def _looks_generic_payment_method(value: Any) -> bool:
    payment_method = normalize_property_token(value)
    return payment_method in {
        "",
        "PAYMENTAPP",
        "PERSONTOPERSONPAYMENT",
        "BANKTRANSFER",
        "BANK",
        "DEBIT",
        "DEBITCARD",
        "CARD",
        "CREDIT",
        "CREDITCARD",
    }


def _is_payment_screenshot_document(document_type: Any) -> bool:
    normalized = normalize_property_token(document_type)
    return "PAYMENT" in normalized and ("SCREENSHOT" in normalized or "APP" in normalized)


def _needs_payment_screenshot_context(extracted: Optional[Dict[str, Any]]) -> bool:
    if not extracted:
        return False
    shared_fields = extracted.get("shared_fields") if isinstance(extracted.get("shared_fields"), dict) else {}
    payment_method = shared_fields.get("payment_method")
    vendor = _string_or_none(shared_fields.get("vendor"))
    account_holder = _string_or_none(shared_fields.get("account_holder"))
    account_number = _string_or_none(shared_fields.get("account_number"))
    if _looks_generic_payment_method(payment_method):
        return True
    if not vendor and not account_holder:
        return True
    if not account_number or _looks_masked_account_number(account_number):
        return True
    return False


def _title_case_words(value: Any) -> Optional[str]:
    raw_value = _string_or_none(value)
    if not raw_value:
        return None
    if raw_value.startswith("@"):
        return raw_value
    words = [word for word in re.split(r"\s+", raw_value) if word]
    if not words:
        return None
    return " ".join(word.capitalize() for word in words)


def _add_property_alias(alias_map: Dict[str, str], alias: Optional[str], canonical: Optional[str]) -> None:
    normalized_alias = normalize_property_token(alias)
    normalized_canonical = _string_or_none(canonical)
    if not normalized_alias or not normalized_canonical:
        return
    alias_map[normalized_alias] = normalized_canonical


def _canonical_property_display_label(value: Any) -> Optional[str]:
    raw_value = _string_or_none(value)
    if not raw_value:
        return None

    if PROPERTY_CODE_PATTERN.search(raw_value):
        return PROPERTY_CODE_PATTERN.search(raw_value).group(1).upper()

    label = re.sub(r"\s+", " ", raw_value).strip(" -")
    if ":" in label:
        label = label.split(":", 1)[0].strip()
    label = re.sub(r"\bMiddlefork\b", "", label, flags=re.IGNORECASE).strip(" -")
    label = re.sub(r"\s+", " ", label).strip()
    if not label:
        return None
    if not label.lower().startswith("the "):
        label = f"The {label}"
    return label.strip()


def _portfolio_property_name_aliases(portfolio: Any) -> List[str]:
    raw_name = _string_or_none(getattr(portfolio, "property_name", None))
    if not raw_name:
        return []

    tokens = [
        token.strip()
        for token in re.split(r"\s*(?:\+|/|&|,)\s*", raw_name)
        if token.strip()
    ]
    aliases: List[str] = []
    for token in tokens:
        if token.lower().startswith("the "):
            aliases.append(token)
            aliases.append(token[4:])
        else:
            aliases.append(token)
            aliases.append(f"The {token}")
    return aliases


def build_property_alias_map(portfolio: Any, uploads: Sequence[Any]) -> Dict[str, str]:
    alias_map: Dict[str, str] = {}
    known_codes = set()
    listing_labels: List[str] = []
    expense_labels: List[str] = []
    listing_mappings = getattr(portfolio, "listing_mappings", None) or []

    for mapping in listing_mappings:
        canonical_label = _string_or_none(getattr(mapping, "official_name", None))
        if not canonical_label or not getattr(mapping, "is_active", True):
            continue
        for alias in [
            canonical_label,
            getattr(mapping, "listing_name", None),
            getattr(mapping, "internal_listing_name", None),
            *list(getattr(mapping, "aliases", None) or []),
        ]:
            _add_property_alias(alias_map, alias, canonical_label)
        if getattr(mapping, "listing_id", None) not in (None, ""):
            _add_property_alias(alias_map, str(mapping.listing_id), canonical_label)

    for upload in uploads:
        payload = normalized_revenue_upload_payload(upload) if getattr(upload, "stage", None) == "revenue" else {"rows": upload.parsed_rows or []}
        for row in payload["rows"] or []:
            for value in row.values():
                if not isinstance(value, str):
                    continue
                known_codes.update(match.upper() for match in PROPERTY_CODE_PATTERN.findall(value))
            for listing_key in ("Property", "Listing", "Property name"):
                canonical_label = _canonical_property_display_label(row.get(listing_key))
                if canonical_label:
                    listing_labels.append(canonical_label)
                    _add_property_alias(alias_map, row.get(listing_key), canonical_label)
        structured = ((getattr(upload, "summary", None) or {}).get("structured_extraction")) or {}
        for entry in structured.get("entries") or []:
            canonical_label = _canonical_property_display_label(entry.get("property_code"))
            if canonical_label:
                expense_labels.append(canonical_label)

    portfolio_prefix = normalize_property_token(getattr(portfolio, "code", None))
    for full_code in sorted(known_codes):
        _add_property_alias(alias_map, full_code, full_code)
        parts = full_code.split("-")
        if len(parts) < 3:
            continue
        unit_token = normalize_property_token(parts[1])
        _add_property_alias(alias_map, unit_token, full_code)
        _add_property_alias(alias_map, unit_token.replace("I", "1"), full_code)
        _add_property_alias(alias_map, unit_token.replace("1", "I"), full_code)
        if portfolio_prefix:
            _add_property_alias(alias_map, f"{portfolio_prefix}{unit_token}", full_code)

    for label in sorted(set(listing_labels)):
        _add_property_alias(alias_map, label, label)
        if label.lower().startswith("the "):
            _add_property_alias(alias_map, label[4:], label)
        for token in re.findall(r"[A-Za-z]{3,}", label):
            if token.lower() == "the":
                continue
            _add_property_alias(alias_map, token, label)

    for label in sorted(set(expense_labels)):
        _add_property_alias(alias_map, label, label)
        if label.lower().startswith("the "):
            _add_property_alias(alias_map, label[4:], label)
        for token in re.findall(r"[A-Za-z]{3,}", label):
            if token.lower() == "the":
                continue
            _add_property_alias(alias_map, token, label)

    unmatched_listing_labels = sorted(set(listing_labels) - set(expense_labels))
    unmatched_expense_labels = sorted(set(expense_labels) - set(listing_labels))
    if len(unmatched_listing_labels) == len(unmatched_expense_labels) == 1:
        source_label = unmatched_listing_labels[0]
        target_label = unmatched_expense_labels[0]
        _add_property_alias(alias_map, source_label, target_label)
        if source_label.lower().startswith("the "):
            _add_property_alias(alias_map, source_label[4:], target_label)
        for token in re.findall(r"[A-Za-z]{3,}", source_label):
            if token.lower() == "the":
                continue
            _add_property_alias(alias_map, token, target_label)

    for alias in _portfolio_property_name_aliases(portfolio):
        canonical_label = _canonical_property_display_label(alias) or alias
        _add_property_alias(alias_map, alias, canonical_label)
        if canonical_label.lower().startswith("the "):
            _add_property_alias(alias_map, canonical_label[4:], canonical_label)

    return alias_map


def canonicalize_property_code(value: Any, property_alias_map: Optional[Dict[str, str]]) -> Optional[str]:
    raw_value = _string_or_none(value)
    if not raw_value:
        return None

    if not property_alias_map:
        return raw_value

    direct_match = property_alias_map.get(normalize_property_token(raw_value))
    if direct_match:
        return direct_match

    first_token = normalize_property_token(raw_value.split()[0])
    return property_alias_map.get(first_token, raw_value)


def _parse_date_sequence_from_filename(filename: str) -> List[date]:
    stem = Path(filename or "").stem
    multi_match = re.search(r"([A-Za-z]+)\s+((?:\d{1,2}\s*,\s*)+)(\d{4})", stem)
    if multi_match:
        month_name, day_fragment, year_value = multi_match.groups()
        dates: List[date] = []
        for day_value in re.findall(r"\d{1,2}", day_fragment):
            normalized = f"{month_name} {day_value}, {year_value}"
            for fmt in ("%B %d, %Y", "%b %d, %Y"):
                try:
                    dates.append(datetime.strptime(normalized, fmt).date())
                    break
                except ValueError:
                    continue
        if dates:
            return dates

    patterns = (
        r"([A-Za-z]+)\s+(\d{1,2}),\s*(\d{4})",
        r"([A-Za-z]+)-(\d{1,2})-(\d{4})",
    )
    for pattern in patterns:
        match = re.search(pattern, stem)
        if not match:
            continue
        month_name, day_value, year_value = match.groups()
        normalized = f"{month_name} {day_value}, {year_value}"
        for fmt in ("%B %d, %Y", "%b %d, %Y"):
            try:
                return [datetime.strptime(normalized, fmt).date()]
            except ValueError:
                continue
    return []


def _parse_date_from_filename(filename: str) -> Optional[date]:
    parsed_dates = _parse_date_sequence_from_filename(filename)
    return parsed_dates[0] if parsed_dates else None


def _unique_property_codes(property_alias_map: Optional[Dict[str, str]]) -> List[str]:
    if not property_alias_map:
        return []
    return sorted({code for code in property_alias_map.values() if code})


def _token_variants(raw_token: str) -> List[str]:
    normalized = normalize_property_token(raw_token)
    variants = {normalized}
    for index, character in enumerate(normalized):
        if character == "1":
            variants.add(normalized[:index] + "I" + normalized[index + 1 :])
        elif character == "I":
            variants.add(normalized[:index] + "1" + normalized[index + 1 :])
    if normalized.isdigit() and len(normalized) == 2:
        variants.add(f"{normalized}M")
        variants.add(normalized[:-1] + "I")
    if len(normalized) == 3 and normalized.endswith("1"):
        variants.add(normalized[:-1] + "I")
    return [variant for variant in variants if variant]


def _candidate_property_codes_for_token(raw_token: str, property_alias_map: Optional[Dict[str, str]]) -> List[str]:
    unique_codes = _unique_property_codes(property_alias_map)
    token_variants = set(_token_variants(raw_token))
    matches: List[str] = []
    if property_alias_map:
        for variant in token_variants:
            direct_match = property_alias_map.get(variant)
            if direct_match and direct_match not in matches:
                matches.append(direct_match)
    for full_code in unique_codes:
        parts = full_code.split("-")
        unit_token = normalize_property_token(parts[1]) if len(parts) >= 2 else normalize_property_token(full_code)
        full_token = normalize_property_token(full_code)
        if unit_token in token_variants or full_token in token_variants:
            matches.append(full_code)
            continue
        if any(variant == unit_token.replace("I", "1") for variant in token_variants):
            matches.append(full_code)
            continue
        if any(variant == unit_token.replace("1", "I") for variant in token_variants):
            matches.append(full_code)
    for variant in token_variants:
        synthesized = PT300_FULL_CODE_BY_UNIT.get(variant)
        if synthesized and synthesized not in matches:
            matches.append(synthesized)
    return matches


def _cleaning_rate_for_property(property_code: Optional[str]) -> Optional[float]:
    normalized_code = (property_code or "").upper()
    if not normalized_code:
        return None

    unit_token = ""
    parts = normalized_code.split("-")
    if len(parts) >= 2:
        unit_token = normalize_property_token(parts[1])

    if unit_token in PT300_CLEANING_RATE_BY_UNIT:
        return PT300_CLEANING_RATE_BY_UNIT[unit_token]
    if normalized_code.endswith("-IG"):
        return 65.0
    if normalized_code.endswith("-KW"):
        return 60.0
    return 65.0


def _resolve_property_assignments_for_total(
    token_candidates: Sequence[Dict[str, Any]],
    target_total: float,
) -> Optional[List[Tuple[str, float]]]:
    if not token_candidates:
        return []

    normalized_target = round(target_total, 2)
    candidate_options: List[List[Tuple[str, float]]] = []
    for token in token_candidates:
        options: List[Tuple[str, float]] = []
        for property_code in token.get("candidates") or []:
            rate = _cleaning_rate_for_property(property_code)
            if rate is None:
                continue
            options.append((property_code, rate))
        deduped: List[Tuple[str, float]] = []
        seen = set()
        for property_code, rate in options:
            key = (property_code, round(rate, 2))
            if key in seen:
                continue
            seen.add(key)
            deduped.append((property_code, rate))
        if not deduped:
            return None
        candidate_options.append(deduped)

    min_suffix_sums = [0.0] * (len(candidate_options) + 1)
    max_suffix_sums = [0.0] * (len(candidate_options) + 1)
    for index in range(len(candidate_options) - 1, -1, -1):
        min_suffix_sums[index] = min(rate for _, rate in candidate_options[index]) + min_suffix_sums[index + 1]
        max_suffix_sums[index] = max(rate for _, rate in candidate_options[index]) + max_suffix_sums[index + 1]

    def backtrack(index: int, remaining: float, chosen: List[Tuple[str, float]]) -> Optional[List[Tuple[str, float]]]:
        rounded_remaining = round(remaining, 2)
        if index == len(candidate_options):
            return list(chosen) if abs(rounded_remaining) <= 0.02 else None

        if rounded_remaining < round(min_suffix_sums[index], 2) - 0.02:
            return None
        if rounded_remaining > round(max_suffix_sums[index], 2) + 0.02:
            return None

        for property_code, rate in sorted(candidate_options[index], key=lambda option: (option[1], option[0])):
            chosen.append((property_code, rate))
            resolved = backtrack(index + 1, rounded_remaining - rate, chosen)
            if resolved is not None:
                return resolved
            chosen.pop()
        return None

    return backtrack(0, normalized_target, [])


def _memo_override_entries_for_payment_screenshot(
    document_type: str,
    filename: str,
    shared_fields: Dict[str, Any],
    property_alias_map: Optional[Dict[str, str]],
    existing_entries: Optional[Sequence[Dict[str, Any]]] = None,
) -> Optional[List[Dict[str, Any]]]:
    if not _is_payment_screenshot_document(document_type):
        return None

    message_text = _string_or_none(shared_fields.get("message_text"))
    shared_total = _float_or_none(shared_fields.get("total"))
    if shared_total is None and existing_entries:
        shared_total = round(
            sum(
                _float_or_none(entry.get("total"))
                or _float_or_none(entry.get("amount"))
                or 0.0
                for entry in existing_entries
            ),
            2,
        )
    if not message_text or shared_total is None:
        return None

    memo_text = " ".join(message_text.split())
    memo_lower = memo_text.lower()
    service_date = parse_date_or_none(shared_fields.get("service_date")) or _parse_date_from_filename(filename)
    payment_date = parse_date_or_none(shared_fields.get("payment_date"))
    vendor = _string_or_none(shared_fields.get("vendor"))

    if "parking reimbursement" in memo_lower:
        portfolio_code = normalize_property_token(filename).split("EXPENSE", 1)[0] or None
        return [
            {
                "category": "misc",
                "confidence": 0.99,
                "property_code": None,
                "scope": "portfolio",
                "item_name": "Staff Parking",
                "vendor": vendor,
                "amount": round(shared_total, 2),
                "total": round(shared_total, 2),
                "service_date": service_date,
                "payment_date": payment_date,
                "payment_method": _string_or_none(shared_fields.get("payment_method")),
                "account_holder": _string_or_none(shared_fields.get("account_holder")),
                "account_number": _string_or_none(shared_fields.get("account_number")),
                "purchase_type": None,
                "store_name": None,
                "quantity": None,
                "unit_amount": None,
                "subtotal": None,
                "discount": None,
                "shipping": None,
                "tax": None,
                "reimbursement_method": None,
                "reimbursement_date": None,
                "details": memo_text,
                "review_reason": None,
                "line_index": 0,
                "group_key": safe_filename(filename),
            }
        ]

    segments = [segment.strip() for segment in re.split(r"\s*\+\s*", memo_text) if segment.strip()]
    base_segment = segments[0] if segments else memo_text
    base_has_deep_cleaning = "deep cleaning" in base_segment.lower()

    def extract_property_tokens(segment_text: str) -> List[Dict[str, Any]]:
        tokens: List[Dict[str, Any]] = []
        for raw_token in PAYMENT_MEMO_TOKEN_PATTERN.findall(segment_text):
            candidates = _candidate_property_codes_for_token(raw_token, property_alias_map)
            if candidates:
                tokens.append({"raw": raw_token, "candidates": candidates})
        return tokens

    def parse_deep_cleaning_segment(segment_text: str) -> Optional[Dict[str, Any]]:
        if "deep cleaning" not in segment_text.lower():
            return None
        amount_match = re.search(r"\(\s*(\d+(?:\.\d+)?)\s*\)", segment_text)
        cleaned_segment = re.sub(r"(?i)deep cleaning", "", segment_text)
        cleaned_segment = re.sub(r"\(\s*\d+(?:\.\d+)?\s*\)", "", cleaned_segment).strip()
        tokens = extract_property_tokens(cleaned_segment)
        if not tokens:
            return None
        return {
            "tokens": tokens,
            "total_amount": round(float(amount_match.group(1)), 2) if amount_match else None,
        }

    property_tokens: List[Dict[str, Any]] = [] if base_has_deep_cleaning else extract_property_tokens(base_segment)
    deep_cleaning_specs: List[Dict[str, Any]] = []
    if base_has_deep_cleaning:
        parsed_deep_segment = parse_deep_cleaning_segment(base_segment)
        if parsed_deep_segment:
            deep_cleaning_specs.append(parsed_deep_segment)

    special_specs: List[Dict[str, Any]] = []
    special_total = 0.0
    for extra_segment in segments[1:]:
        guest_service_match = re.search(r"(?:guest\s+)?services?\s*\(?\s*(\d+(?:\.\d+)?)\s*\)?", extra_segment, re.IGNORECASE)
        if guest_service_match:
            amount = round(float(guest_service_match.group(1)), 2)
            special_specs.append({"category": "misc", "item_name": "Guest Service", "amount": amount})
            special_total += amount
            continue

        deep_cleaning_spec = parse_deep_cleaning_segment(extra_segment)
        if deep_cleaning_spec:
            deep_cleaning_specs.append(deep_cleaning_spec)
            if deep_cleaning_spec["total_amount"] is not None:
                special_total += deep_cleaning_spec["total_amount"]
            continue

        maintenance_amount_match = re.search(r"\(?\s*(\d+(?:\.\d+)?)\s*\)?", extra_segment)
        if maintenance_amount_match and any(keyword in extra_segment.lower() for keyword in MAINTENANCE_MEMO_KEYWORDS):
            amount = round(float(maintenance_amount_match.group(1)), 2)
            maintenance_tokens = extract_property_tokens(extra_segment)
            explicit_property_code = None
            review_reason = None
            if len(maintenance_tokens) == 1:
                candidates = maintenance_tokens[0].get("candidates") or []
                if len(candidates) == 1:
                    explicit_property_code = candidates[0]
            elif maintenance_tokens:
                review_reason = "Multiple unit references appear in the maintenance note. Review ops conversation to assign the property."

            if explicit_property_code is None and review_reason is None:
                review_reason = "Unit not specified in the payment memo. Review ops conversation to assign the property."

            details = extra_segment
            for token in maintenance_tokens:
                raw_token = token.get("raw")
                if raw_token:
                    details = re.sub(rf"\b{re.escape(raw_token)}\b", "", details, flags=re.IGNORECASE)
            details = re.sub(r"\(?\s*\d+(?:\.\d+)?\s*\)?", "", details).strip(" -")
            special_specs.append(
                {
                    "category": "maintenance",
                    "item_name": details.title() if details else "Maintenance",
                    "amount": amount,
                    "property_code": explicit_property_code,
                    "scope": "property" if explicit_property_code else "portfolio",
                    "review_reason": review_reason,
                }
            )
            special_total += amount

    remaining_total = round(shared_total - special_total, 2)
    if remaining_total < -0.02:
        return None

    assignments: Optional[List[Tuple[str, float]]] = []
    if property_tokens:
        assignments = _resolve_property_assignments_for_total(property_tokens, remaining_total)
        if assignments is None:
            return None

    normalized_entries: List[Dict[str, Any]] = []
    for line_index, (property_code, amount) in enumerate(assignments or []):
        normalized_entries.append(
            {
                "category": "cleaning",
                "confidence": 0.99,
                "property_code": property_code,
                "scope": "property",
                "item_name": "Cleaning",
                "vendor": vendor,
                "amount": round(amount, 2),
                "total": round(amount, 2),
                "service_date": service_date,
                "payment_date": payment_date,
                "payment_method": _string_or_none(shared_fields.get("payment_method")),
                "account_holder": _string_or_none(shared_fields.get("account_holder")),
                "account_number": _string_or_none(shared_fields.get("account_number")),
                "purchase_type": None,
                "store_name": None,
                "quantity": None,
                "unit_amount": None,
                "subtotal": None,
                "discount": None,
                "shipping": None,
                "tax": None,
                "reimbursement_method": None,
                "reimbursement_date": None,
                "details": memo_text,
                "review_reason": None,
                "line_index": line_index,
                "group_key": safe_filename(filename),
            }
        )

    unresolved_deep_specs = [spec for spec in deep_cleaning_specs if spec.get("total_amount") is None]
    if unresolved_deep_specs:
        if len(unresolved_deep_specs) == 1 and not property_tokens:
            unresolved_deep_specs[0]["total_amount"] = remaining_total
        else:
            return None

    for deep_spec in deep_cleaning_specs:
        deep_total = round(deep_spec["total_amount"] or 0.0, 2)
        deep_tokens = deep_spec["tokens"]
        if deep_total <= 0 or not deep_tokens:
            continue
        per_property = round(deep_total / len(deep_tokens), 2)
        for token in deep_tokens:
            candidates = token.get("candidates") or []
            if not candidates:
                return None
            normalized_entries.append(
                {
                    "category": "cleaning",
                    "confidence": 0.99,
                    "property_code": candidates[0],
                    "scope": "property",
                    "item_name": "deep cleaning",
                    "vendor": vendor,
                    "amount": per_property,
                    "total": per_property,
                    "service_date": service_date,
                    "payment_date": payment_date,
                    "payment_method": _string_or_none(shared_fields.get("payment_method")),
                    "account_holder": _string_or_none(shared_fields.get("account_holder")),
                    "account_number": _string_or_none(shared_fields.get("account_number")),
                    "purchase_type": None,
                    "store_name": None,
                    "quantity": None,
                    "unit_amount": None,
                    "subtotal": None,
                    "discount": None,
                    "shipping": None,
                    "tax": None,
                    "reimbursement_method": None,
                    "reimbursement_date": None,
                    "details": memo_text,
                    "review_reason": None,
                    "line_index": len(normalized_entries),
                    "group_key": safe_filename(filename),
                }
            )

    anchor_property = normalized_entries[-1]["property_code"] if normalized_entries else None
    for offset, special_spec in enumerate(special_specs, start=len(normalized_entries)):
        property_code = special_spec.get("property_code", anchor_property)
        scope = special_spec.get("scope") or ("property" if property_code else "portfolio")
        normalized_entries.append(
            {
                "category": special_spec["category"],
                "confidence": 0.99,
                "property_code": property_code,
                "scope": scope,
                "item_name": special_spec["item_name"],
                "vendor": vendor,
                "amount": special_spec["amount"],
                "total": special_spec["amount"],
                "service_date": service_date,
                "payment_date": payment_date,
                "payment_method": _string_or_none(shared_fields.get("payment_method")),
                "account_holder": _string_or_none(shared_fields.get("account_holder")),
                "account_number": _string_or_none(shared_fields.get("account_number")),
                "purchase_type": None,
                "store_name": None,
                "quantity": None,
                "unit_amount": None,
                "subtotal": None,
                "discount": None,
                "shipping": None,
                "tax": None,
                "reimbursement_method": None,
                "reimbursement_date": None,
                "details": memo_text,
                "review_reason": special_spec.get("review_reason"),
                "line_index": offset,
                "group_key": safe_filename(filename),
            }
        )

    return normalized_entries


def normalize_structured_expense_extraction(
    extracted: Dict[str, Any],
    filename: str,
    property_alias_map: Optional[Dict[str, str]] = None,
) -> Optional[Dict[str, Any]]:
    if not extracted:
        return None

    document_type = _string_or_none(extracted.get("document_type")) or "expense_evidence"
    if _is_payment_screenshot_document(document_type):
        document_type = "payment_screenshot"
    support_only = bool(extracted.get("support_only"))
    explicit_guest_refund_language = bool(extracted.get("explicit_guest_refund_language"))
    shared_fields = extracted.get("shared_fields") if isinstance(extracted.get("shared_fields"), dict) else {}
    reimbursement_proof = extracted.get("reimbursement_proof") if isinstance(extracted.get("reimbursement_proof"), dict) else {}
    raw_entries = extracted.get("entries") if isinstance(extracted.get("entries"), list) else []
    normalized_entries: List[Dict[str, Any]] = []

    if document_type == "payment_screenshot":
        shared_payment_date = parse_date_or_none(shared_fields.get("payment_date"))
        shared_service_date = parse_date_or_none(shared_fields.get("service_date")) or _parse_date_from_filename(filename)
        if shared_payment_date and shared_service_date and shared_payment_date == shared_service_date and len(_parse_date_sequence_from_filename(filename)) >= 2:
            shared_fields = {**shared_fields}
            shared_fields.pop("payment_date", None)

    for index, raw_entry in enumerate(raw_entries):
        if not isinstance(raw_entry, dict):
            continue
        category = _string_or_none(raw_entry.get("category"))
        if category not in EXPENSE_CATEGORIES:
            continue

        confidence = max(0.0, min(1.0, _float_or_none(raw_entry.get("confidence")) or _float_or_none(extracted.get("overall_confidence")) or 0.0))
        property_code = canonicalize_property_code(raw_entry.get("property_code"), property_alias_map)
        amount = _float_or_none(raw_entry.get("amount"))
        total = _float_or_none(raw_entry.get("total"))
        subtotal = _float_or_none(raw_entry.get("subtotal"))
        if total is None:
            total = amount if amount is not None else subtotal
        scope = _coerce_scope(raw_entry.get("scope"), property_code)
        review_reason = _string_or_none(raw_entry.get("review_reason"))

        if category == "direct_refund" and not explicit_guest_refund_language:
            confidence = min(confidence, 0.45)
            review_reason = review_reason or "Direct refund proof needs explicit guest refund wording in the evidence."

        normalized_entries.append(
            {
                "category": category,
                "confidence": round(confidence, 4),
                "property_code": property_code,
                "scope": scope,
                "item_name": _string_or_none(raw_entry.get("item_name")),
                "vendor": _string_or_none(raw_entry.get("vendor")) or _string_or_none(shared_fields.get("vendor")),
                "amount": amount,
                "total": total,
                "service_date": parse_date_or_none(raw_entry.get("service_date")),
                "payment_date": parse_date_or_none(raw_entry.get("payment_date")) or parse_date_or_none(shared_fields.get("payment_date")),
                "payment_method": _string_or_none(raw_entry.get("payment_method")) or _string_or_none(shared_fields.get("payment_method")),
                "account_holder": _string_or_none(raw_entry.get("account_holder")) or _string_or_none(shared_fields.get("account_holder")),
                "account_number": _string_or_none(raw_entry.get("account_number")) or _string_or_none(shared_fields.get("account_number")),
                "purchase_type": _string_or_none(raw_entry.get("purchase_type")) or _string_or_none(shared_fields.get("purchase_type")),
                "store_name": _string_or_none(raw_entry.get("store_name")) or _string_or_none(shared_fields.get("store_name")),
                "quantity": _float_or_none(raw_entry.get("quantity")),
                "unit_amount": _float_or_none(raw_entry.get("unit_amount")),
                "subtotal": subtotal,
                "discount": _float_or_none(raw_entry.get("discount")),
                "shipping": _float_or_none(raw_entry.get("shipping")),
                "tax": _float_or_none(raw_entry.get("tax")),
                "reimbursement_method": _string_or_none(raw_entry.get("reimbursement_method")) or _string_or_none(shared_fields.get("reimbursement_method")),
                "reimbursement_date": parse_date_or_none(raw_entry.get("reimbursement_date")) or parse_date_or_none(shared_fields.get("reimbursement_date")),
                "details": _string_or_none(raw_entry.get("details")),
                "review_reason": review_reason,
                "line_index": index,
                "group_key": _string_or_none(raw_entry.get("group_key")) or safe_filename(filename),
            }
        )

    memo_override_entries = _memo_override_entries_for_payment_screenshot(
        document_type=document_type,
        filename=filename,
        shared_fields=shared_fields,
        property_alias_map=property_alias_map,
        existing_entries=normalized_entries,
    )
    if memo_override_entries:
        normalized_entries = memo_override_entries

    shared_total = _float_or_none(shared_fields.get("total"))
    if normalized_entries and shared_total is not None:
        populated_totals = [
            round((entry.get("total") if entry.get("total") is not None else entry.get("amount")), 2)
            for entry in normalized_entries
            if entry.get("total") is not None or entry.get("amount") is not None
        ]
        if (
            document_type == "payment_screenshot"
            and len(populated_totals) >= 3
            and len(set(populated_totals)) == 1
            and abs((populated_totals[0] * len(populated_totals)) - round(shared_total, 2)) <= 0.05
        ):
            for entry in normalized_entries:
                entry["amount"] = None
                entry["total"] = None

        remaining = round(shared_total, 2)
        unresolved_entries = []
        for entry in normalized_entries:
            current_total = entry.get("total") if entry.get("total") is not None else entry.get("amount")
            if current_total is not None:
                remaining = round(remaining - current_total, 2)
            else:
                unresolved_entries.append(entry)

        for entry in list(unresolved_entries):
            property_code = entry.get("property_code") or ""
            if entry["category"] != "cleaning" or "deep" in (entry.get("item_name") or "").lower():
                continue
            if property_code.endswith("-IG"):
                rate = 65.0
            elif property_code.endswith("-KW"):
                rate = 60.0
            else:
                rate = None
            if rate is None:
                continue
            entry["amount"] = rate
            entry["total"] = rate
            remaining = round(remaining - rate, 2)
            unresolved_entries.remove(entry)

        if unresolved_entries:
            if len(unresolved_entries) == 1:
                entry = unresolved_entries[0]
                entry["amount"] = remaining
                entry["total"] = remaining
            else:
                split_amount = round(remaining / len(unresolved_entries), 2) if len(unresolved_entries) else 0.0
                running_total = 0.0
                for index, entry in enumerate(unresolved_entries):
                    allocated = split_amount if index < len(unresolved_entries) - 1 else round(remaining - running_total, 2)
                    entry["amount"] = allocated
                    entry["total"] = allocated
                    running_total = round(running_total + allocated, 2)

    auto_entries = [
        entry for entry in normalized_entries
        if (entry.get("amount") is not None or entry.get("total") is not None)
        and not (entry["category"] == "direct_refund" and not explicit_guest_refund_language)
    ]

    top_entry = normalized_entries[0] if normalized_entries else None
    return {
        "document_type": document_type,
        "support_only": support_only,
        "overall_confidence": round(max((_float_or_none(extracted.get("overall_confidence")) or 0.0), *(entry["confidence"] for entry in normalized_entries)) if normalized_entries else (_float_or_none(extracted.get("overall_confidence")) or 0.0), 4),
        "explicit_guest_refund_language": explicit_guest_refund_language,
        "guest_refund_phrase": _string_or_none(extracted.get("guest_refund_phrase")),
        "shared_fields": {
            key: (value.isoformat() if isinstance(value, date) else value)
            for key, value in {
                "vendor": _string_or_none(shared_fields.get("vendor")),
                "payment_method": _string_or_none(shared_fields.get("payment_method")),
                "account_holder": _string_or_none(shared_fields.get("account_holder")),
                "account_number": _string_or_none(shared_fields.get("account_number")),
                "payment_date": parse_date_or_none(shared_fields.get("payment_date")),
                "purchase_type": _string_or_none(shared_fields.get("purchase_type")),
                "store_name": _string_or_none(shared_fields.get("store_name")),
                "message_text": _string_or_none(shared_fields.get("message_text")),
                "service_date": parse_date_or_none(shared_fields.get("service_date")) or _parse_date_from_filename(filename),
                "subtotal": _float_or_none(shared_fields.get("subtotal")),
                "discount": _float_or_none(shared_fields.get("discount")),
                "shipping": _float_or_none(shared_fields.get("shipping")),
                "tax": _float_or_none(shared_fields.get("tax")),
                "total": _float_or_none(shared_fields.get("total")),
                "reimbursement_method": _string_or_none(shared_fields.get("reimbursement_method")),
                "reimbursement_date": parse_date_or_none(shared_fields.get("reimbursement_date")),
            }.items()
            if value not in (None, "")
        },
        "reimbursement_proof": {
            key: (value.isoformat() if isinstance(value, date) else value)
            for key, value in {
                "amount": _float_or_none(reimbursement_proof.get("amount")),
                "payment_date": parse_date_or_none(reimbursement_proof.get("payment_date")),
                "payment_method": _string_or_none(reimbursement_proof.get("payment_method")),
                "account_holder": _string_or_none(reimbursement_proof.get("account_holder")),
                "account_number": _string_or_none(reimbursement_proof.get("account_number")),
            }.items()
            if value not in (None, "")
        },
        "entries": [
            {
                **entry,
                "service_date": entry["service_date"].isoformat() if entry.get("service_date") else None,
                "payment_date": entry["payment_date"].isoformat() if entry.get("payment_date") else None,
                "reimbursement_date": entry["reimbursement_date"].isoformat() if entry.get("reimbursement_date") else None,
            }
            for entry in normalized_entries
        ],
        "auto_entries_count": len(auto_entries),
        "can_auto_create": bool(auto_entries or support_only),
        "primary_entry": {
            "category": top_entry.get("category") if top_entry else None,
            "confidence": top_entry.get("confidence") if top_entry else 0.0,
            "item_name": top_entry.get("item_name") if top_entry else None,
            "vendor": top_entry.get("vendor") if top_entry else None,
            "property_code": top_entry.get("property_code") if top_entry else None,
            "amount": top_entry.get("amount") if top_entry else None,
            "total": top_entry.get("total") if top_entry else None,
            "payment_date": top_entry.get("payment_date").isoformat() if top_entry and top_entry.get("payment_date") else None,
        },
    }


def _request_openai_json(
    *,
    system_prompt: str,
    user_content: Any,
    timeout_seconds: float,
) -> Dict[str, Any]:
    from openai import OpenAI

    client = OpenAI(
        api_key=config.OPENAI_API_KEY,
        timeout=timeout_seconds,
        max_retries=OPENAI_VISION_MAX_RETRIES,
    )
    response = client.chat.completions.create(
        model=config.OPENAI_MODEL,
        temperature=0,
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_content},
        ],
    )
    raw_content = (response.choices[0].message.content or "{}").strip()
    return json.loads(raw_content)


def _build_expense_extraction_prompt(*, dense_retail_fallback: bool = False) -> str:
    prompt = (
        "Extract the document into structured bookkeeping data. "
        "Return keys: document_type, overall_confidence, support_only, explicit_guest_refund_language, guest_refund_phrase, "
        "shared_fields, reimbursement_proof, entries. "
        "shared_fields can contain vendor, payment_method, account_holder, account_number, payment_date, service_date, message_text, purchase_type, store_name, subtotal, discount, shipping, tax, total, reimbursement_method, reimbursement_date. "
        "reimbursement_proof can contain amount, payment_date, payment_method, account_holder, account_number. "
        "Each entry must include category, confidence, item_name, vendor, property_code, scope, amount, total, service_date, payment_date, payment_method, account_holder, account_number, purchase_type, store_name, quantity, unit_amount, subtotal, discount, shipping, tax, reimbursement_method, reimbursement_date, details, review_reason. "
        "Do not aggregate multiple properties into one entry when the evidence clearly lists them separately. "
        "For retail receipts, one receipt line item must become one entry whenever the document shows distinct purchased items. "
        "For PT300 cleaning screenshots, always read and return the exact message memo in shared_fields.message_text. "
        "For Venmo or Zelle screenshots, also extract payment_date, vendor, account_holder, and handle information from the transfer card itself whenever visible."
    )
    if dense_retail_fallback:
        prompt += (
            " Fallback mode: this is likely a dense itemized retail receipt. "
            "Take the extra time needed to enumerate every distinct purchased line item that is legible. "
            "Prefer exhaustive line-item coverage over collapsing the receipt into category summaries. "
            "If some lines are partially unreadable, still return the legible ones and use review_reason on uncertain rows."
        )
    return prompt


def _build_expense_extraction_content(
    *,
    extraction_prompt: str,
    file_bytes: bytes,
    filename: str,
    normalized_mime_type: Optional[str],
    preview_text: Optional[str],
    property_alias_map: Optional[Dict[str, str]],
) -> List[Dict[str, Any]]:
    content: List[Dict[str, Any]] = [{"type": "text", "text": extraction_prompt}]
    if property_alias_map:
        known_codes = sorted({code for code in property_alias_map.values()})
        if known_codes:
            content.append(
                {
                    "type": "text",
                    "text": f"Known property codes for canonicalization: {', '.join(known_codes[:40])}",
                }
            )
    if normalized_mime_type:
        vision_bytes, vision_mime_type = _optimize_image_bytes_for_vision(file_bytes, normalized_mime_type)
        encoded = base64.b64encode(vision_bytes).decode("utf-8")
        content.append(
            {
                "type": "image_url",
                "image_url": {
                    "url": f"data:{vision_mime_type};base64,{encoded}",
                },
            }
        )
    if preview_text:
        content.append(
            {
                "type": "text",
                "text": f"Extracted text preview:\n{preview_text[:4000]}",
            }
        )
    content.append({"type": "text", "text": f"Filename context: {filename}"})
    return content


def _is_timeout_error(exc: Exception) -> bool:
    if "timed out" in str(exc).lower():
        return True
    error_name = type(exc).__name__.lower()
    return "timeout" in error_name


def _looks_dense_itemized_retail_receipt(
    filename: str,
    preview_text: Optional[str],
) -> bool:
    signal_text = f"{filename}\n{preview_text or ''}".lower()
    if any(hint in signal_text for hint in PAYMENT_APP_FILENAME_HINTS):
        return False
    if any(hint in signal_text for hint in RETAIL_RECEIPT_FILENAME_HINTS):
        return True
    if preview_text:
        line_count = len([line for line in preview_text.splitlines() if line.strip()])
        currency_count = len(re.findall(r"\d+\.\d{2}", preview_text))
        return line_count >= 12 and currency_count >= 6
    return False


def _should_retry_dense_retail_receipt_fallback(
    *,
    filename: str,
    preview_text: Optional[str],
    exc: Exception,
) -> bool:
    return _is_timeout_error(exc) and _looks_dense_itemized_retail_receipt(filename, preview_text)


def _extract_payment_screenshot_context(
    file_bytes: bytes,
    filename: str,
    mime_type: Optional[str],
) -> Optional[Dict[str, Any]]:
    normalized_mime_type = _normalized_image_mime_type(filename, mime_type)
    if not normalized_mime_type:
        return None
    optimized_bytes, optimized_mime_type = _optimize_image_bytes_for_vision(file_bytes, normalized_mime_type)

    encoded = base64.b64encode(optimized_bytes).decode("utf-8")
    try:
        return _request_openai_json(
            system_prompt=(
                "You identify consumer payment app screenshots. "
                "Return JSON only with keys app_name, payment_method, payee_name, payee_handle, funding_source_label. "
                "Use payment_method values like Venmo, Zelle, Cash App, PayPal, Apple Cash, bank_transfer, debit, credit_card. "
                "If the screenshot shows Venmo UI chrome such as Privacy/Private, the Venmo logo, or the 'Turn on for purchases' panel, set payment_method to Venmo even when the funding source is a bank card. "
                "Do not confuse the funding source with the payment app. "
                "Only return payee_handle when it is visible."
            ),
            user_content=[
                {"type": "text", "text": f"Filename context: {filename}"},
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:{optimized_mime_type};base64,{encoded}",
                    },
                },
            ],
            timeout_seconds=OPENAI_VISION_TIMEOUT_SECONDS,
        )
    except ImportError:
        return None
    except json.JSONDecodeError:
        return None


def _apply_payment_screenshot_context(
    extracted: Optional[Dict[str, Any]],
    payment_context: Optional[Dict[str, Any]],
) -> Optional[Dict[str, Any]]:
    if not extracted or not payment_context:
        return extracted

    shared_fields = extracted.setdefault("shared_fields", {})
    proof_fields = extracted.setdefault("reimbursement_proof", {})
    normalized_method = _string_or_none(payment_context.get("payment_method"))
    payee_name = _title_case_words(payment_context.get("payee_name"))
    payee_handle = _string_or_none(payment_context.get("payee_handle"))

    if normalized_method and (_looks_generic_payment_method(shared_fields.get("payment_method")) or normalized_method in {"Venmo", "Zelle", "Cash App", "PayPal", "Apple Cash"}):
        shared_fields["payment_method"] = normalized_method
        if proof_fields and (_looks_generic_payment_method(proof_fields.get("payment_method")) or not proof_fields.get("payment_method")):
            proof_fields["payment_method"] = normalized_method

    if payee_name:
        if not shared_fields.get("vendor") or shared_fields.get("vendor") == shared_fields.get("account_holder"):
            shared_fields["vendor"] = payee_name
        if not shared_fields.get("account_holder") or shared_fields.get("account_holder") == shared_fields.get("vendor"):
            shared_fields["account_holder"] = payee_name
        if proof_fields and not proof_fields.get("account_holder"):
            proof_fields["account_holder"] = payee_name

    if payee_handle and (not shared_fields.get("account_number") or _looks_masked_account_number(shared_fields.get("account_number"))):
        shared_fields["account_number"] = payee_handle
    if payee_handle and proof_fields and (not proof_fields.get("account_number") or _looks_masked_account_number(proof_fields.get("account_number"))):
        proof_fields["account_number"] = payee_handle

    for entry in extracted.get("entries") or []:
        if normalized_method and (_looks_generic_payment_method(entry.get("payment_method")) or not entry.get("payment_method")):
            entry["payment_method"] = normalized_method
        if payee_name and (not entry.get("vendor") or entry.get("vendor") == entry.get("account_holder")):
            entry["vendor"] = payee_name
        if payee_name and not entry.get("account_holder"):
            entry["account_holder"] = payee_name
        if payee_handle and (not entry.get("account_number") or _looks_masked_account_number(entry.get("account_number"))):
            entry["account_number"] = payee_handle

    return extracted


def _revenue_checkout_date(source: str, row: Dict[str, Any]) -> Optional[date]:
    if source == "airbnb":
        if _string_or_none(row.get("Type")) != "Reservation":
            return None
        return parse_date_or_none(row.get("End date"))
    if source == "booking_com":
        return parse_date_or_none(row.get("Check-Out Date")) or parse_date_or_none(row.get("Departure"))
    if source in {"vrbo", "hopper", "direct_bookings", "google"}:
        if not _is_data_row_for_summary(source, row):
            return None
        return parse_date_or_none(row.get("Check-Out Date"))
    return None


def _build_property_turnover_dates(
    uploads: Sequence[Any],
    property_alias_map: Optional[Dict[str, str]],
) -> Dict[str, List[date]]:
    turnover_dates: Dict[str, List[date]] = defaultdict(list)
    for upload in uploads:
        if getattr(upload, "stage", None) != "revenue":
            continue
        payload = normalized_revenue_upload_payload(upload)
        for row in payload["rows"] or []:
            checkout_date = _revenue_checkout_date(upload.source, row)
            if not checkout_date:
                continue
            raw_property = row.get("Property") or row.get("Listing") or row.get("Property name")
            property_code = canonicalize_property_code(raw_property, property_alias_map) or _canonical_property_display_label(raw_property)
            if not property_code:
                continue
            turnover_dates[property_code].append(checkout_date)
    return {
        property_code: sorted(values)
        for property_code, values in turnover_dates.items()
    }


def extract_expense_evidence_bundle(
    file_bytes: bytes,
    filename: str,
    mime_type: Optional[str],
    parsed_summary: Optional[Dict[str, Any]] = None,
    property_alias_map: Optional[Dict[str, str]] = None,
) -> Optional[Dict[str, Any]]:
    normalized_mime_type = _normalized_image_mime_type(filename, mime_type)
    preview_text = (parsed_summary or {}).get("preview_text")

    if not normalized_mime_type and not preview_text:
        return None

    system_prompt = (
        "You extract bookkeeping evidence for a short-term rental operator. "
        "Return JSON only. A single document may contain multiple bookkeeping entries. "
        "Allowed categories are cleaning, maintenance, supplies, misc, software_fee, direct_refund. "
        "Use support_only=true only for reimbursement proofs or supporting documents that should not create an expense by themselves. "
        "For direct_refund, set explicit_guest_refund_language=true only when the evidence explicitly says guest refund, refund to guest, guest reimbursement, or equivalent wording. "
        "When a cleaning screenshot lists multiple units, return one entry per unit with its own amount and service date. "
        "When a Venmo or Zelle memo names properties like Acorn, Oak, or unit nicknames, treat each named property occurrence as its own cleaning candidate instead of collapsing the payment into one row. "
        "For itemized store receipts and invoices, return one entry per purchased line item and do not collapse the receipt into a category summary. "
        "For PT300 cleaner Zelle screenshots, capture the exact payment memo in shared_fields.message_text. "
        "For person-to-person payment screenshots, extract the visible payee name or payment handle as vendor/account_holder when shown, and do not prefer the funding bank account name over the transfer recipient. "
        "If the screenshot is clearly inside Venmo, Zelle, Cash App, PayPal, or another consumer payment app, set shared_fields.payment_method to that app instead of the funding source. "
        "A Venmo payment screen can still be funded by Chase, debit, or bank balance; in that case the payment method is Venmo, not bank_transfer. "
        "For online order receipts, do not use the shipping recipient or ship-to name as account_holder unless the document explicitly identifies that person as the purchaser or cardholder. "
        "Known PT300 standard cleaning rates are: 10H=65, 10I=65, 10M=60, 15I=65, 15J=65, 15K=65, 18C=65, 19M=60, "
        "20M=60, 21M=60, 23J=65, 2I=65, 2M=60, 3E=65, 6N=65, 8H=65. "
        "guest service(5) is usually one $5 misc row, parking reimbursement is misc, and explicit repair notes like unclog toilet (30) are maintenance."
    )
    content = _build_expense_extraction_content(
        extraction_prompt=_build_expense_extraction_prompt(),
        file_bytes=file_bytes,
        filename=filename,
        normalized_mime_type=normalized_mime_type,
        preview_text=preview_text,
        property_alias_map=property_alias_map,
    )

    try:
        extracted = _request_openai_json(
            system_prompt=system_prompt,
            user_content=content,
            timeout_seconds=OPENAI_VISION_TIMEOUT_SECONDS,
        )
    except ImportError:
        return None
    except Exception as exc:
        if not _should_retry_dense_retail_receipt_fallback(
            filename=filename,
            preview_text=preview_text,
            exc=exc,
        ):
            raise

        logger.info(
            "Retrying dense retail receipt extraction for %s with %.1fs timeout fallback",
            filename,
            OPENAI_VISION_SLOW_FALLBACK_TIMEOUT_SECONDS,
        )
        fallback_content = _build_expense_extraction_content(
            extraction_prompt=_build_expense_extraction_prompt(dense_retail_fallback=True),
            file_bytes=file_bytes,
            filename=filename,
            normalized_mime_type=normalized_mime_type,
            preview_text=preview_text,
            property_alias_map=property_alias_map,
        )
        extracted = _request_openai_json(
            system_prompt=system_prompt,
            user_content=fallback_content,
            timeout_seconds=max(
                OPENAI_VISION_SLOW_FALLBACK_TIMEOUT_SECONDS,
                OPENAI_VISION_TIMEOUT_SECONDS,
            ),
        )

    document_type = _string_or_none(extracted.get("document_type")) or ""
    if _is_payment_screenshot_document(document_type) and _needs_payment_screenshot_context(extracted):
        payment_context = _extract_payment_screenshot_context(
            file_bytes=file_bytes,
            filename=filename,
            mime_type=normalized_mime_type,
        )
        extracted = _apply_payment_screenshot_context(extracted, payment_context)
    return normalize_structured_expense_extraction(extracted, filename, property_alias_map=property_alias_map)


def auto_extract_expense_evidence(
    file_bytes: bytes,
    filename: str,
    mime_type: Optional[str],
    parsed_summary: Optional[Dict[str, Any]] = None,
) -> Optional[Dict[str, Any]]:
    structured = extract_expense_evidence_bundle(
        file_bytes=file_bytes,
        filename=filename,
        mime_type=mime_type,
        parsed_summary=parsed_summary,
        property_alias_map=None,
    )
    return auto_extract_expense_evidence_from_structured(structured)


def auto_extract_expense_evidence_from_structured(
    structured: Optional[Dict[str, Any]],
) -> Optional[Dict[str, Any]]:
    if not structured:
        return None
    primary_entry = structured.get("primary_entry") or {}
    can_auto_create = bool(
        structured.get("can_auto_create")
        and primary_entry.get("category")
        and primary_entry.get("confidence", 0.0) >= AUTO_CATEGORIZE_CONFIDENCE_THRESHOLD
    )
    return {
        "category": primary_entry.get("category"),
        "confidence": round(_float_or_none(primary_entry.get("confidence")) or 0.0, 4),
        "can_auto_create": can_auto_create,
        "explicit_guest_refund_language": bool(structured.get("explicit_guest_refund_language")),
        "guest_refund_phrase": structured.get("guest_refund_phrase"),
        "item_name": primary_entry.get("item_name"),
        "vendor": primary_entry.get("vendor"),
        "property_code": primary_entry.get("property_code"),
        "scope": _coerce_scope(primary_entry.get("scope"), primary_entry.get("property_code")),
        "amount": _float_or_none(primary_entry.get("amount")),
        "total": _float_or_none(primary_entry.get("total")),
        "service_date": None,
        "payment_date": primary_entry.get("payment_date"),
        "payment_method": structured.get("shared_fields", {}).get("payment_method"),
        "account_holder": structured.get("shared_fields", {}).get("account_holder"),
        "account_number": structured.get("shared_fields", {}).get("account_number"),
        "purchase_type": structured.get("shared_fields", {}).get("purchase_type"),
        "store_name": structured.get("shared_fields", {}).get("store_name"),
        "quantity": None,
        "unit_amount": None,
        "subtotal": _float_or_none(structured.get("shared_fields", {}).get("subtotal")),
        "discount": _float_or_none(structured.get("shared_fields", {}).get("discount")),
        "shipping": _float_or_none(structured.get("shared_fields", {}).get("shipping")),
        "tax": _float_or_none(structured.get("shared_fields", {}).get("tax")),
        "reimbursement_method": structured.get("shared_fields", {}).get("reimbursement_method"),
        "reimbursement_date": structured.get("shared_fields", {}).get("reimbursement_date"),
        "details": None,
        "review_reason": (primary_entry or {}).get("review_reason"),
    }


def build_expense_item_payloads_from_extraction(
    structured_extraction: Optional[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    if not structured_extraction or structured_extraction.get("support_only"):
        return []

    shared_fields = structured_extraction.get("shared_fields") or {}
    explicit_guest_refund_language = bool(structured_extraction.get("explicit_guest_refund_language"))
    payloads: List[Dict[str, Any]] = []

    for entry in structured_extraction.get("entries") or []:
        category = entry.get("category")
        if category not in EXPENSE_CATEGORIES:
            continue
        if category == "direct_refund" and not explicit_guest_refund_language:
            continue

        amount = _float_or_none(entry.get("amount"))
        total = _float_or_none(entry.get("total"))
        subtotal = _float_or_none(entry.get("subtotal"))
        if amount is None and total is None and subtotal is None:
            continue

        confidence = _float_or_none(entry.get("confidence")) or 0.0
        needs_review = confidence < AUTO_CATEGORIZE_CONFIDENCE_THRESHOLD or bool(entry.get("review_reason"))
        payloads.append(
            {
                "category": category,
                "item_name": _string_or_none(entry.get("item_name")),
                "vendor": _string_or_none(entry.get("vendor")) or _string_or_none(shared_fields.get("vendor")),
                "property_code": _string_or_none(entry.get("property_code")),
                "scope": entry.get("scope") or _coerce_scope(entry.get("scope"), entry.get("property_code")),
                "amount": amount,
                "total": total,
                "service_date": parse_date_or_none(entry.get("service_date")),
                "payment_date": parse_date_or_none(entry.get("payment_date")) or parse_date_or_none(shared_fields.get("payment_date")),
                "payment_method": _string_or_none(entry.get("payment_method")) or _string_or_none(shared_fields.get("payment_method")),
                "account_holder": _string_or_none(entry.get("account_holder")) or _string_or_none(shared_fields.get("account_holder")),
                "account_number": _string_or_none(entry.get("account_number")) or _string_or_none(shared_fields.get("account_number")),
                "purchase_type": _string_or_none(entry.get("purchase_type")) or _string_or_none(shared_fields.get("purchase_type")),
                "store_name": _string_or_none(entry.get("store_name")) or _string_or_none(shared_fields.get("store_name")),
                "quantity": _float_or_none(entry.get("quantity")),
                "unit_amount": _float_or_none(entry.get("unit_amount")),
                "subtotal": subtotal,
                "discount": _float_or_none(entry.get("discount")),
                "shipping": _float_or_none(entry.get("shipping")),
                "tax": _float_or_none(entry.get("tax")),
                "reimbursement_method": _string_or_none(entry.get("reimbursement_method")) or _string_or_none(shared_fields.get("reimbursement_method")),
                "reimbursement_date": parse_date_or_none(entry.get("reimbursement_date")) or parse_date_or_none(shared_fields.get("reimbursement_date")),
                "details": _string_or_none(entry.get("details")),
                "needs_review": needs_review,
                "review_reason": _string_or_none(entry.get("review_reason")),
                "extraction_data": {
                    "source": "auto_expense_evidence",
                    "confidence": confidence,
                    "document_type": structured_extraction.get("document_type"),
                    "group_key": entry.get("group_key"),
                    "line_index": entry.get("line_index"),
                    "explicit_guest_refund_language": explicit_guest_refund_language,
                    "guest_refund_phrase": structured_extraction.get("guest_refund_phrase"),
                    "shared_fields": shared_fields,
                },
            }
        )

    return payloads


def _payment_screenshot_total(structured_extraction: Optional[Dict[str, Any]]) -> Optional[float]:
    if not structured_extraction:
        return None

    shared_total = _float_or_none((structured_extraction.get("shared_fields") or {}).get("total"))
    if shared_total is not None:
        return round(shared_total, 2)

    totals = [
        _float_or_none(entry.get("total")) if _float_or_none(entry.get("total")) is not None else _float_or_none(entry.get("amount"))
        for entry in structured_extraction.get("entries") or []
    ]
    totals = [value for value in totals if value is not None]
    return round(sum(totals), 2) if totals else None


def _normalized_person_label(value: Any) -> Optional[str]:
    return _title_case_words(value)


def _match_property_alias_in_segment(segment_text: str, property_alias_map: Optional[Dict[str, str]]) -> Optional[str]:
    if not property_alias_map:
        return None

    matches: List[Tuple[int, str]] = []
    for raw_token in PAYMENT_MEMO_TOKEN_PATTERN.findall(segment_text):
        for candidate in _candidate_property_codes_for_token(raw_token, property_alias_map):
            matches.append((len(normalize_property_token(raw_token)), candidate))

    normalized_segment = normalize_property_token(segment_text)
    for alias, canonical in property_alias_map.items():
        if len(alias) < 3 or alias.isdigit():
            continue
        if alias in normalized_segment:
            matches.append((len(alias), canonical))

    if not matches:
        return None

    matches.sort(key=lambda item: (-item[0], item[1]))
    best_length = matches[0][0]
    best_candidates = sorted({candidate for length, candidate in matches if length == best_length})
    return best_candidates[0] if len(best_candidates) == 1 else None


def _parse_named_property_payment_plan(
    upload: Any,
    property_alias_map: Optional[Dict[str, str]],
) -> Optional[Dict[str, Any]]:
    structured = ((getattr(upload, "summary", None) or {}).get("structured_extraction")) or {}
    if not _is_payment_screenshot_document(structured.get("document_type")):
        return None

    shared_fields = structured.get("shared_fields") or {}
    message_text = _string_or_none(shared_fields.get("message_text"))
    if not message_text:
        return None

    total = _payment_screenshot_total(structured)
    if total is None:
        return None

    segments = [segment.strip() for segment in re.split(r"\s*\+\s*", message_text) if segment.strip()]
    if not segments:
        return None

    cleaning_segments = []
    special_segments = []
    last_property = None

    for segment in segments:
        lower_segment = segment.lower()
        explicit_amount_match = re.search(r"\(\s*(\d+(?:\.\d+)?)\s*\)", segment)
        explicit_amount = round(float(explicit_amount_match.group(1)), 2) if explicit_amount_match else None

        if "trip" in lower_segment:
            special_segments.append(
                {
                    "category": "misc",
                    "item_name": "Property Trip",
                    "amount": explicit_amount,
                    "property_code": last_property,
                    "segment": segment,
                }
            )
            continue

        matched_property = _match_property_alias_in_segment(segment, property_alias_map)
        if not matched_property:
            return None

        quantity_match = re.search(r"(?:x|×)\s*(\d+)", segment, re.IGNORECASE)
        quantity = max(1, int(quantity_match.group(1))) if quantity_match else 1
        cleaning_segments.append(
            {
                "property_code": matched_property,
                "count": quantity,
                "segment": segment,
            }
        )
        last_property = matched_property

    if not cleaning_segments:
        return None

    return {
        "upload_id": getattr(upload, "bookkeeping_upload_id", None),
        "filename": getattr(upload, "original_filename", None),
        "message_text": message_text,
        "shared_fields": shared_fields,
        "total": total,
        "dates": _parse_date_sequence_from_filename(getattr(upload, "original_filename", None)),
        "cleaning_segments": cleaning_segments,
        "special_segments": special_segments,
    }


def _infer_named_property_cleaning_rates(
    plans: Sequence[Dict[str, Any]],
) -> Dict[str, float]:
    rate_map: Dict[str, float] = {}
    changed = True

    while changed:
        changed = False
        for plan in plans:
            property_counts = Counter()
            for segment in plan.get("cleaning_segments") or []:
                property_counts[segment["property_code"]] += int(segment.get("count") or 1)

            fixed_special_total = sum(
                round(float(segment.get("amount")), 2)
                for segment in plan.get("special_segments") or []
                if segment.get("amount") is not None
            )
            remaining_total = round(float(plan.get("total") or 0) - fixed_special_total, 2)
            unknown_properties = []

            for property_code, count in property_counts.items():
                if property_code in rate_map:
                    remaining_total = round(remaining_total - (rate_map[property_code] * count), 2)
                else:
                    unknown_properties.append(property_code)

            if len(unknown_properties) != 1:
                continue

            unresolved_specials = [segment for segment in plan.get("special_segments") or [] if segment.get("amount") is None]
            if unresolved_specials:
                continue

            property_code = unknown_properties[0]
            inferred_rate = round(remaining_total / property_counts[property_code], 2)
            if inferred_rate <= 0:
                continue
            previous_rate = rate_map.get(property_code)
            if previous_rate is None or abs(previous_rate - inferred_rate) > 0.02:
                rate_map[property_code] = inferred_rate
                changed = True

    return rate_map


def _build_named_property_payment_payloads(
    plan: Dict[str, Any],
    rate_map: Dict[str, float],
    turnover_dates_by_property: Optional[Dict[str, List[date]]] = None,
) -> Optional[List[Dict[str, Any]]]:
    property_counts = Counter()
    for segment in plan.get("cleaning_segments") or []:
        property_counts[segment["property_code"]] += int(segment.get("count") or 1)

    if any(property_code not in rate_map for property_code in property_counts):
        return None

    cleanings_total = round(sum(rate_map[property_code] * count for property_code, count in property_counts.items()), 2)
    remaining_total = round(float(plan.get("total") or 0) - cleanings_total, 2)
    special_segments = [dict(segment) for segment in plan.get("special_segments") or []]
    unresolved_specials = [segment for segment in special_segments if segment.get("amount") is None]

    if len(unresolved_specials) > 1:
        return None
    if len(unresolved_specials) == 1:
        unresolved_specials[0]["amount"] = remaining_total
        remaining_total = 0.0

    if abs(remaining_total) > 0.02 and not special_segments:
        return None

    shared_fields = plan.get("shared_fields") or {}
    vendor = _normalized_person_label(shared_fields.get("vendor"))
    payment_method = _string_or_none(shared_fields.get("payment_method"))
    account_holder = _normalized_person_label(shared_fields.get("account_holder")) or vendor
    account_number = _string_or_none(shared_fields.get("account_number"))
    payment_date = parse_date_or_none(shared_fields.get("payment_date"))
    dates = plan.get("dates") or []
    message_text = plan.get("message_text")
    filename = plan.get("filename")

    cleaning_occurrences: List[str] = []
    for segment in plan.get("cleaning_segments") or []:
        cleaning_occurrences.extend([segment["property_code"]] * int(segment.get("count") or 1))

    special_service_dates: List[Optional[date]] = []
    cleaning_dates = list(dates)
    for _ in special_segments:
        if not cleaning_dates:
            special_service_dates.append(None)
            continue
        special_index = min(len(cleaning_dates) // 2, len(cleaning_dates) - 1)
        special_service_dates.append(cleaning_dates.pop(special_index))

    property_to_turnover_dates = {
        property_code: list((turnover_dates_by_property or {}).get(property_code) or [])
        for property_code in property_counts
    }
    remaining_dates = list(cleaning_dates)
    assigned_dates_by_property: Dict[str, List[Optional[date]]] = {property_code: [] for property_code in property_counts}

    for property_code, count in property_counts.items():
        available_turnovers = set(property_to_turnover_dates.get(property_code) or [])
        exact_matches = [service_date for service_date in remaining_dates if service_date in available_turnovers][:count]
        assigned_dates_by_property[property_code].extend(exact_matches)
        unmatched_exact = Counter(exact_matches)
        updated_remaining_dates = []
        for service_date in remaining_dates:
            if unmatched_exact.get(service_date, 0):
                unmatched_exact[service_date] -= 1
                continue
            updated_remaining_dates.append(service_date)
        remaining_dates = updated_remaining_dates

    for property_code in cleaning_occurrences:
        if len(assigned_dates_by_property[property_code]) >= property_counts[property_code]:
            continue
        fallback_date = remaining_dates.pop(0) if remaining_dates else (dates[-1] if dates else None)
        assigned_dates_by_property[property_code].append(fallback_date)

    payloads: List[Dict[str, Any]] = []
    for index, property_code in enumerate(cleaning_occurrences):
        property_dates = assigned_dates_by_property.get(property_code) or []
        service_date = property_dates.pop(0) if property_dates else (dates[index] if index < len(dates) else (dates[-1] if dates else None))
        payloads.append(
            {
                "category": "cleaning",
                "item_name": "Cleaning",
                "vendor": vendor,
                "property_code": property_code,
                "scope": "property",
                "amount": round(rate_map[property_code], 2),
                "total": round(rate_map[property_code], 2),
                "service_date": service_date,
                "payment_date": payment_date,
                "payment_method": payment_method,
                "account_holder": account_holder,
                "account_number": account_number,
                "purchase_type": None,
                "store_name": None,
                "quantity": None,
                "unit_amount": None,
                "subtotal": None,
                "discount": None,
                "shipping": None,
                "tax": None,
                "reimbursement_method": None,
                "reimbursement_date": None,
                "details": message_text,
                "needs_review": False,
                "review_reason": None,
                "extraction_data": {
                    "source": "auto_expense_evidence",
                    "confidence": 0.98,
                    "document_type": "payment_screenshot",
                    "group_key": safe_filename(filename),
                    "line_index": index,
                    "explicit_guest_refund_language": False,
                    "guest_refund_phrase": None,
                    "shared_fields": shared_fields,
                },
            }
        )

    if special_segments:
        special_index_base = len(payloads)
        for offset, segment in enumerate(special_segments, start=special_index_base):
            amount = _float_or_none(segment.get("amount"))
            if amount is None or amount <= 0:
                continue
            payloads.append(
                {
                    "category": segment.get("category") or "misc",
                    "item_name": segment.get("item_name") or "Misc. Expense",
                    "vendor": vendor,
                    "property_code": segment.get("property_code"),
                    "scope": "property" if segment.get("property_code") else "portfolio",
                    "amount": round(amount, 2),
                    "total": round(amount, 2),
                    "service_date": special_service_dates[offset - special_index_base] if offset - special_index_base < len(special_service_dates) else (dates[-1] if dates else None),
                    "payment_date": payment_date,
                    "payment_method": payment_method,
                    "account_holder": account_holder,
                    "account_number": account_number,
                    "purchase_type": None,
                    "store_name": None,
                    "quantity": None,
                    "unit_amount": None,
                    "subtotal": None,
                    "discount": None,
                    "shipping": None,
                    "tax": None,
                    "reimbursement_method": None,
                    "reimbursement_date": None,
                    "details": message_text,
                    "needs_review": False,
                    "review_reason": None,
                    "extraction_data": {
                        "source": "auto_expense_evidence",
                        "confidence": 0.96,
                        "document_type": "payment_screenshot",
                        "group_key": safe_filename(filename),
                        "line_index": offset,
                        "explicit_guest_refund_language": False,
                        "guest_refund_phrase": None,
                        "shared_fields": shared_fields,
                    },
                }
            )

    return payloads or None


def reconcile_named_property_payment_uploads(
    uploads: Sequence[Any],
    property_alias_map: Optional[Dict[str, str]],
) -> Dict[int, List[Dict[str, Any]]]:
    plans: List[Dict[str, Any]] = []
    for upload in uploads:
        if getattr(upload, "stage", None) != "expense":
            continue
        plan = _parse_named_property_payment_plan(upload, property_alias_map)
        if plan:
            plans.append(plan)

    if not plans:
        return {}

    rate_map = _infer_named_property_cleaning_rates(plans)
    turnover_dates_by_property = _build_property_turnover_dates(uploads, property_alias_map)
    payloads_by_upload: Dict[int, List[Dict[str, Any]]] = {}
    for plan in plans:
        payloads = _build_named_property_payment_payloads(plan, rate_map, turnover_dates_by_property=turnover_dates_by_property)
        upload_id = plan.get("upload_id")
        if payloads and upload_id is not None:
            payloads_by_upload[int(upload_id)] = payloads
    return payloads_by_upload


def reconcile_reimbursement_receipts(uploads: Sequence[Any], expense_items: Sequence[Any]) -> None:
    grouped_supply_items: Dict[int, List[Any]] = defaultdict(list)
    for item in expense_items:
        if item.category == "supplies" and item.upload_id:
            grouped_supply_items[item.upload_id].append(item)

    support_uploads = []
    for upload in uploads:
        summary = upload.summary or {}
        structured = summary.get("structured_extraction") or {}
        if not structured or not structured.get("support_only"):
            continue
        reimbursement_proof = structured.get("reimbursement_proof") or {}
        amount = _float_or_none(reimbursement_proof.get("amount"))
        if amount is None:
            continue
        support_uploads.append(
            {
                "upload": upload,
                "amount": round(amount, 2),
                "payment_date": parse_date_or_none(reimbursement_proof.get("payment_date")),
                "payment_method": reimbursement_proof.get("payment_method"),
                "account_holder": reimbursement_proof.get("account_holder"),
                "account_number": reimbursement_proof.get("account_number"),
            }
        )

    supply_groups = []
    for upload_id, items in grouped_supply_items.items():
        first_item = items[0]
        shared_total = _float_or_none(_item_extraction_field(first_item, "total"))
        total = round(shared_total if shared_total is not None else sum(item.effective_total() for item in items), 2)
        if total <= 0:
            continue
        payment_date = max((item.payment_date for item in items if item.payment_date), default=None)
        supply_groups.append(
            {
                "upload_id": upload_id,
                "items": items,
                "amount": total,
                "payment_date": payment_date,
                "assigned": False,
            }
        )

    for proof in sorted(support_uploads, key=lambda item: ((item["payment_date"] or date.max), item["amount"])):
        eligible_groups = [
            group for group in supply_groups
            if not group["assigned"] and (proof["payment_date"] is None or group["payment_date"] is None or group["payment_date"] <= proof["payment_date"])
        ]
        matched_group_ids = None
        for size in range(1, min(len(eligible_groups), 6) + 1):
            for combo in combinations(eligible_groups, size):
                total = round(sum(group["amount"] for group in combo), 2)
                if abs(total - proof["amount"]) <= 0.02:
                    matched_group_ids = {group["upload_id"] for group in combo}
                    break
            if matched_group_ids:
                break

        if not matched_group_ids:
            continue

        for group in supply_groups:
            if group["upload_id"] not in matched_group_ids:
                continue
            group["assigned"] = True
            for item in group["items"]:
                item.reimbursement_method = proof["payment_method"] or item.reimbursement_method
                item.account_holder = proof["account_holder"] or item.account_holder
                item.account_number = proof["account_number"] or item.account_number
                item.reimbursement_date = proof["payment_date"] or item.reimbursement_date
                extraction_data = item.extraction_data or {}
                extraction_data["reimbursement_upload_id"] = proof["upload"].bookkeeping_upload_id
                item.extraction_data = extraction_data


def aggregate_uploads_by_source(uploads: Sequence[Any], stage: str) -> Dict[str, List[Any]]:
    grouped: Dict[str, List[Any]] = defaultdict(list)
    for upload in uploads:
        if upload.stage == stage:
            grouped[upload.source].append(upload)
    return grouped


def aggregate_expense_totals(expense_items: Sequence[Any]) -> Dict[str, float]:
    totals = {category: 0.0 for category in EXPENSE_CATEGORIES}
    grouped_supplies: Dict[Any, List[Any]] = defaultdict(list)
    for item in expense_items:
        if item.category == "supplies":
            group_key = _item_extraction_field(item, "group_key") or item.upload_id or item.bookkeeping_expense_item_id
            grouped_supplies[group_key].append(item)
            continue
        totals[item.category] = totals.get(item.category, 0.0) + float(item.effective_total())

    for grouped_items in grouped_supplies.values():
        first_item = grouped_items[0]
        shared_total = _float_or_none(_item_extraction_field(first_item, "total"))
        if shared_total is None:
            shared_total = sum(item.effective_total() for item in grouped_items)
        totals["supplies"] += shared_total
    return {key: round(value, 2) for key, value in totals.items()}


def _build_listing_mapping_lookup(listing_mappings: Optional[Sequence[Any]]) -> Dict[str, Any]:
    alias_to_mapping: Dict[str, Any] = {}
    by_id: Dict[int, Any] = {}
    for mapping in listing_mappings or []:
        mapping_id = getattr(mapping, "bookkeeping_listing_mapping_id", None)
        if mapping_id is not None:
            by_id[mapping_id] = mapping
        if not getattr(mapping, "is_active", True):
            continue
        aliases: List[str] = []
        for candidate in (
            getattr(mapping, "official_name", None),
            getattr(mapping, "internal_listing_name", None),
            getattr(mapping, "listing_name", None),
            getattr(mapping, "listing_id", None),
        ):
            if candidate not in (None, ""):
                aliases.append(str(candidate))
        aliases.extend(str(alias) for alias in (getattr(mapping, "aliases", None) or []) if str(alias).strip())
        for alias in aliases:
            alias_to_mapping.setdefault(normalize_property_token(alias), mapping)
    return {"alias_to_mapping": alias_to_mapping, "by_id": by_id}


def _match_listing_mapping(raw_value: Any, listing_lookup: Optional[Dict[str, Any]]) -> Optional[Any]:
    if not listing_lookup:
        return None
    alias_to_mapping = listing_lookup.get("alias_to_mapping") or {}
    normalized = normalize_property_token(raw_value)
    if not normalized:
        return None
    return alias_to_mapping.get(normalized)


def _revenue_field_header(source: str, field_name: str) -> Optional[str]:
    return (REVENUE_ROW_FIELD_MAP.get(field_name) or {}).get(source)


def _serialize_revenue_row(source: str, row: Dict[str, Any]) -> Dict[str, Any]:
    headers = STANDARD_REVENUE_HEADERS.get(source, list(row.keys()))
    serialized: Dict[str, Any] = {}
    for header in headers:
        serialized[header] = json_safe_value(row.get(header))
    return serialized


def _entity_to_dict(entity: Any) -> Dict[str, Any]:
    if entity is None:
        return {}
    if hasattr(entity, "to_dict"):
        return entity.to_dict()
    data = {}
    for key, value in getattr(entity, "__dict__", {}).items():
        if key.startswith("_"):
            continue
        data[key] = json_safe_value(value)
    return data


def _trim_preview_text(value: Any, limit: int = WORKSPACE_PREVIEW_TEXT_LIMIT) -> Optional[str]:
    text = _string_or_none(value)
    if not text:
        return None
    if len(text) <= limit:
        return text
    return f"{text[:limit].rstrip()}..."


def _compact_upload_summary(summary: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not isinstance(summary, dict) or not summary:
        return {}

    compact: Dict[str, Any] = {}
    preview_text = _trim_preview_text(summary.get("preview_text"))
    if preview_text:
        compact["preview_text"] = preview_text

    page_count = summary.get("page_count")
    if page_count not in (None, ""):
        compact["page_count"] = page_count

    auto_extraction = summary.get("auto_extraction")
    if isinstance(auto_extraction, dict) and auto_extraction:
        compact["auto_extraction"] = {
            key: json_safe_value(auto_extraction.get(key))
            for key in (
                "category",
                "confidence",
                "can_auto_create",
                "explicit_guest_refund_language",
                "guest_refund_phrase",
                "item_name",
                "vendor",
                "property_code",
                "scope",
                "amount",
                "total",
                "payment_date",
                "payment_method",
                "review_reason",
            )
            if auto_extraction.get(key) not in (None, "", [], {})
        }

    structured = summary.get("structured_extraction")
    if isinstance(structured, dict) and structured:
        shared_fields = structured.get("shared_fields") if isinstance(structured.get("shared_fields"), dict) else {}
        reimbursement_proof = (
            structured.get("reimbursement_proof")
            if isinstance(structured.get("reimbursement_proof"), dict)
            else {}
        )
        compact["structured_extraction"] = {
            key: json_safe_value(structured.get(key))
            for key in (
                "document_type",
                "support_only",
                "overall_confidence",
                "explicit_guest_refund_language",
                "guest_refund_phrase",
                "auto_entries_count",
                "can_auto_create",
                "primary_entry",
            )
            if structured.get(key) not in (None, "", [], {})
        }
        if shared_fields:
            compact["structured_extraction"]["shared_fields"] = {
                key: json_safe_value(shared_fields.get(key))
                for key in (
                    "vendor",
                    "payment_method",
                    "account_holder",
                    "account_number",
                    "payment_date",
                    "service_date",
                    "message_text",
                    "purchase_type",
                    "store_name",
                    "subtotal",
                    "discount",
                    "shipping",
                    "tax",
                    "total",
                    "reimbursement_method",
                    "reimbursement_date",
                )
                if shared_fields.get(key) not in (None, "", [], {})
            }
        if reimbursement_proof:
            compact["structured_extraction"]["reimbursement_proof"] = {
                key: json_safe_value(reimbursement_proof.get(key))
                for key in (
                    "amount",
                    "payment_date",
                    "payment_method",
                    "account_holder",
                    "account_number",
                )
                if reimbursement_proof.get(key) not in (None, "", [], {})
            }
        if isinstance(structured.get("entries"), list):
            compact["structured_extraction"]["entry_count"] = len(structured.get("entries") or [])

    return compact


def _workspace_upload_to_dict(upload: Any) -> Dict[str, Any]:
    return {
        "bookkeeping_upload_id": getattr(upload, "bookkeeping_upload_id", None),
        "period_id": getattr(upload, "period_id", None),
        "processing_batch_id": getattr(upload, "processing_batch_id", None),
        "stage": getattr(upload, "stage", None),
        "source": getattr(upload, "source", None),
        "detected_source": getattr(upload, "detected_source", None),
        "sheet_name": getattr(upload, "sheet_name", None),
        "upload_status": getattr(upload, "upload_status", None),
        "original_filename": getattr(upload, "original_filename", None),
        "original_relative_path": getattr(upload, "original_relative_path", None),
        "content_type": getattr(upload, "content_type", None),
        "file_extension": getattr(upload, "file_extension", None),
        "file_size": getattr(upload, "file_size", None),
        "row_count": getattr(upload, "row_count", None),
        "headers": list(getattr(upload, "headers", None) or []),
        "preview_rows": list(getattr(upload, "preview_rows", None) or []),
        "summary": _compact_upload_summary(getattr(upload, "summary", None) or {}),
        "processing_error": getattr(upload, "processing_error", None),
        "processing_started_at": getattr(getattr(upload, "processing_started_at", None), "isoformat", lambda: None)(),
        "processing_completed_at": getattr(getattr(upload, "processing_completed_at", None), "isoformat", lambda: None)(),
        "notes": getattr(upload, "notes", None),
        "uploaded_by": getattr(upload, "uploaded_by", None),
        "created_at": getattr(getattr(upload, "created_at", None), "isoformat", lambda: None)(),
    }


def _workspace_revenue_item_to_dict(item: Any) -> Dict[str, Any]:
    return {
        "bookkeeping_revenue_item_id": getattr(item, "bookkeeping_revenue_item_id", None),
        "period_id": getattr(item, "period_id", None),
        "upload_id": getattr(item, "upload_id", None),
        "listing_mapping_id": getattr(item, "listing_mapping_id", None),
        "source": getattr(item, "source", None),
        "row_index": getattr(item, "row_index", None),
        "reservation_identifier": getattr(item, "reservation_identifier", None),
        "confirmation_code": getattr(item, "confirmation_code", None),
        "guest_name": getattr(item, "guest_name", None),
        "property_code": getattr(item, "property_code", None),
        "raw_listing_name": getattr(item, "raw_listing_name", None),
        "transaction_type": getattr(item, "transaction_type", None),
        "currency": getattr(item, "currency", None),
        "transaction_date": getattr(getattr(item, "transaction_date", None), "isoformat", lambda: None)(),
        "booking_date": getattr(getattr(item, "booking_date", None), "isoformat", lambda: None)(),
        "start_date": getattr(getattr(item, "start_date", None), "isoformat", lambda: None)(),
        "end_date": getattr(getattr(item, "end_date", None), "isoformat", lambda: None)(),
        "nights": getattr(item, "nights", None),
        "gross_amount": money_value(getattr(item, "gross_amount", None)),
        "paid_out_amount": money_value(getattr(item, "paid_out_amount", None)),
        "commission_amount": money_value(getattr(item, "commission_amount", None)),
        "hostaway_fee_amount": money_value(getattr(item, "hostaway_fee_amount", None)),
        "stripe_fee_amount": money_value(getattr(item, "stripe_fee_amount", None)),
        "cleaning_fee_amount": money_value(getattr(item, "cleaning_fee_amount", None)),
        "tax_amount": money_value(getattr(item, "tax_amount", None)),
        "refund_amount": money_value(getattr(item, "refund_amount", None)),
        "details": getattr(item, "details", None),
        "normalized_data": dict(getattr(item, "normalized_data", None) or {}),
        "needs_review": bool(getattr(item, "needs_review", False)),
        "review_reason": getattr(item, "review_reason", None),
        "created_by": getattr(item, "created_by", None),
        "created_at": getattr(getattr(item, "created_at", None), "isoformat", lambda: None)(),
        "updated_at": getattr(getattr(item, "updated_at", None), "isoformat", lambda: None)(),
    }


def _workspace_expense_item_to_dict(item: Any) -> Dict[str, Any]:
    return {
        "bookkeeping_expense_item_id": getattr(item, "bookkeeping_expense_item_id", None),
        "period_id": getattr(item, "period_id", None),
        "upload_id": getattr(item, "upload_id", None),
        "category": getattr(item, "category", None),
        "item_name": getattr(item, "item_name", None),
        "vendor": getattr(item, "vendor", None),
        "property_code": getattr(item, "property_code", None),
        "scope": getattr(item, "scope", None),
        "description": getattr(item, "description", None),
        "amount": money_value(getattr(item, "amount", None)),
        "service_date": getattr(getattr(item, "service_date", None), "isoformat", lambda: None)(),
        "payment_date": getattr(getattr(item, "payment_date", None), "isoformat", lambda: None)(),
        "payment_method": getattr(item, "payment_method", None),
        "account_holder": getattr(item, "account_holder", None),
        "account_number": getattr(item, "account_number", None),
        "purchase_type": getattr(item, "purchase_type", None),
        "store_name": getattr(item, "store_name", None),
        "quantity": money_value(getattr(item, "quantity", None)),
        "unit_amount": money_value(getattr(item, "unit_amount", None)),
        "subtotal": money_value(getattr(item, "subtotal", None)),
        "discount": money_value(getattr(item, "discount", None)),
        "shipping": money_value(getattr(item, "shipping", None)),
        "tax": money_value(getattr(item, "tax", None)),
        "total": money_value(getattr(item, "total", None)),
        "effective_total": item.effective_total() if hasattr(item, "effective_total") else 0.0,
        "reimbursement_method": getattr(item, "reimbursement_method", None),
        "reimbursement_date": getattr(getattr(item, "reimbursement_date", None), "isoformat", lambda: None)(),
        "details": getattr(item, "details", None),
        "needs_review": bool(getattr(item, "needs_review", False)),
        "review_reason": getattr(item, "review_reason", None),
        "extraction_data": dict(getattr(item, "extraction_data", None) or {}),
        "created_by": getattr(item, "created_by", None),
        "created_at": getattr(getattr(item, "created_at", None), "isoformat", lambda: None)(),
        "updated_at": getattr(getattr(item, "updated_at", None), "isoformat", lambda: None)(),
    }


def build_revenue_item_payloads(
    upload: Any,
    property_alias_map: Optional[Dict[str, str]] = None,
    listing_lookup: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    if getattr(upload, "stage", None) != "revenue":
        return []

    payload = normalized_revenue_upload_payload(upload)
    source = getattr(upload, "source", None)
    results: List[Dict[str, Any]] = []

    for row_index, row in enumerate(payload.get("rows") or []):
        if not _is_data_row_for_summary(source, row):
            continue

        listing_header = _revenue_field_header(source, "property_code")
        reservation_header = _revenue_field_header(source, "reservation_identifier")
        confirmation_header = _revenue_field_header(source, "confirmation_code")
        guest_header = _revenue_field_header(source, "guest_name")
        transaction_type_header = _revenue_field_header(source, "transaction_type")
        transaction_date_header = _revenue_field_header(source, "transaction_date")
        booking_date_header = _revenue_field_header(source, "booking_date")
        start_date_header = _revenue_field_header(source, "start_date")
        end_date_header = _revenue_field_header(source, "end_date")
        nights_header = _revenue_field_header(source, "nights")
        currency_header = _revenue_field_header(source, "currency")
        details_header = _revenue_field_header(source, "details")

        raw_listing_name = row.get(listing_header) if listing_header else None
        canonical_property = canonicalize_property_code(raw_listing_name, property_alias_map) or _canonical_property_display_label(raw_listing_name) or _string_or_none(raw_listing_name)
        listing_mapping = _match_listing_mapping(canonical_property, listing_lookup) or _match_listing_mapping(raw_listing_name, listing_lookup)
        property_code = getattr(listing_mapping, "official_name", None) or canonical_property

        normalized_row = _serialize_revenue_row(source, row)
        if listing_header and property_code:
            normalized_row[listing_header] = property_code

        results.append(
            {
                "source": source,
                "row_index": row_index,
                "reservation_identifier": _string_or_none(row.get(reservation_header)) if reservation_header else None,
                "confirmation_code": _string_or_none(row.get(confirmation_header)) if confirmation_header else None,
                "guest_name": _string_or_none(row.get(guest_header)) if guest_header else None,
                "property_code": property_code,
                "raw_listing_name": _string_or_none(raw_listing_name),
                "listing_mapping_id": getattr(listing_mapping, "bookkeeping_listing_mapping_id", None),
                "transaction_type": _string_or_none(row.get(transaction_type_header)) if transaction_type_header else None,
                "currency": _string_or_none(row.get(currency_header)) if currency_header else None,
                "transaction_date": parse_date_or_none(row.get(transaction_date_header)) if transaction_date_header else None,
                "booking_date": parse_date_or_none(row.get(booking_date_header)) if booking_date_header else None,
                "start_date": parse_date_or_none(row.get(start_date_header)) if start_date_header else None,
                "end_date": parse_date_or_none(row.get(end_date_header)) if end_date_header else None,
                "nights": int(money_value(row.get(nights_header))) if nights_header and not is_empty_value(row.get(nights_header)) else None,
                "gross_amount": money_value(row.get(_revenue_field_header(source, "gross_amount"))),
                "paid_out_amount": money_value(row.get(_revenue_field_header(source, "paid_out_amount"))),
                "commission_amount": money_value(row.get(_revenue_field_header(source, "commission_amount"))),
                "hostaway_fee_amount": money_value(row.get(_revenue_field_header(source, "hostaway_fee_amount"))),
                "stripe_fee_amount": money_value(row.get(_revenue_field_header(source, "stripe_fee_amount"))),
                "cleaning_fee_amount": money_value(row.get(_revenue_field_header(source, "cleaning_fee_amount"))),
                "tax_amount": money_value(row.get(_revenue_field_header(source, "tax_amount"))),
                "refund_amount": money_value(row.get(_revenue_field_header(source, "refund_amount"))),
                "details": _string_or_none(row.get(details_header)) if details_header else None,
                "normalized_data": normalized_row,
                "raw_data": {key: json_safe_value(value) for key, value in row.items()},
                "needs_review": False,
                "review_reason": None,
            }
        )

    return results


def build_revenue_rows_by_source(
    revenue_items: Optional[Sequence[Any]] = None,
    uploads: Optional[Sequence[Any]] = None,
) -> Dict[str, List[Dict[str, Any]]]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    for item in revenue_items or []:
        row = dict(getattr(item, "normalized_data", None) or getattr(item, "raw_data", None) or {})
        for field_name in REVENUE_ROW_FIELD_MAP:
            header = _revenue_field_header(getattr(item, "source", None), field_name)
            if not header:
                continue
            value = getattr(item, field_name, None)
            if value not in (None, ""):
                row[header] = json_safe_value(value)
        grouped[item.source].append(row)

    if grouped:
        return {
            source: list(rows)
            for source, rows in grouped.items()
        }

    uploads_by_source = aggregate_uploads_by_source(uploads or [], "revenue")
    for source, source_uploads in uploads_by_source.items():
        _, rows = _merged_headers_and_rows(source_uploads)
        grouped[source].extend(rows)
    return grouped


def aggregate_revenue_totals(
    uploads: Sequence[Any],
    revenue_items: Optional[Sequence[Any]] = None,
) -> Dict[str, Dict[str, Any]]:
    totals: Dict[str, Dict[str, float]] = defaultdict(
        lambda: {
            "gross_total": 0.0,
            "commission_total": 0.0,
            "hostaway_fee_total": 0.0,
            "stripe_fee_total": 0.0,
            "processing_fee_rows": [],
        }
    )

    if revenue_items:
        for item in revenue_items:
            bucket = totals[item.source]
            bucket["gross_total"] += money_value(getattr(item, "gross_amount", None))
            bucket["commission_total"] += money_value(getattr(item, "commission_amount", None))
            bucket["hostaway_fee_total"] += money_value(getattr(item, "hostaway_fee_amount", None))
            bucket["stripe_fee_total"] += money_value(getattr(item, "stripe_fee_amount", None))
            stripe_total = money_value(getattr(item, "stripe_fee_amount", None))
            hostaway_total = money_value(getattr(item, "hostaway_fee_amount", None))
            if stripe_total or hostaway_total:
                bucket["processing_fee_rows"].append(
                    {
                        "guest": getattr(item, "guest_name", None),
                        "booking_platform": "Direct Booking" if item.source == "direct_bookings" else source_label(item.source),
                        "listing": getattr(item, "property_code", None) or getattr(item, "raw_listing_name", None),
                        "stripe": round(stripe_total, 2),
                        "hostaway": round(hostaway_total, 2),
                    }
                )
    else:
        for upload in uploads:
            if upload.stage != "revenue":
                continue
            summary = normalized_revenue_upload_payload(upload)["summary"] or {}
            bucket = totals[upload.source]
            bucket["gross_total"] += money_value(summary.get("gross_total"))
            bucket["commission_total"] += money_value(summary.get("commission_total"))
            bucket["hostaway_fee_total"] += money_value(summary.get("hostaway_fee_total"))
            bucket["stripe_fee_total"] += money_value(summary.get("stripe_fee_total"))
            bucket["processing_fee_rows"].extend(summary.get("processing_fee_rows") or [])

    return {
        source: {
            **{metric: round(value, 2) for metric, value in values.items() if metric != "processing_fee_rows"},
            "processing_fee_rows": list(values.get("processing_fee_rows") or []),
        }
        for source, values in totals.items()
    }


def build_review_queue_rows(
    revenue_items: Sequence[Any],
    expense_items: Sequence[Any],
    change_proposals: Optional[Sequence[Any]] = None,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for item in revenue_items:
        if getattr(item, "needs_review", False) or getattr(item, "review_reason", None):
            rows.append(
                {
                    "kind": "revenue_row",
                    "row_type": "revenue_item",
                    "row_id": item.bookkeeping_revenue_item_id,
                    "label": source_label(item.source),
                    "item": item.guest_name or item.reservation_identifier or "Revenue row",
                    "property": item.property_code or item.raw_listing_name,
                    "amount": round(money_value(item.gross_amount), 2),
                    "reason": item.review_reason or "Revenue row needs review.",
                }
            )
    for item in expense_items:
        if getattr(item, "needs_review", False) or getattr(item, "review_reason", None):
            rows.append(
                {
                    "kind": "expense_row",
                    "row_type": "expense_item",
                    "row_id": item.bookkeeping_expense_item_id,
                    "label": item.category.replace("_", " ").title(),
                    "item": item.item_name or item.vendor or "Expense row",
                    "property": item.property_code,
                    "amount": round(item.effective_total(), 2),
                    "reason": item.review_reason or "Expense row needs review.",
                }
            )
    for proposal in change_proposals or []:
        if getattr(proposal, "status", None) != "pending":
            continue
        rows.append(
            {
                "kind": "change_proposal",
                "row_type": getattr(proposal, "row_type", None),
                "row_id": getattr(proposal, "row_id", None),
                "proposal_id": getattr(proposal, "bookkeeping_ai_change_proposal_id", None),
                "label": "AI change proposal",
                "item": ", ".join(sorted((getattr(proposal, "proposed_values", None) or {}).keys())) or "Pending update",
                "property": None,
                "amount": None,
                "reason": getattr(proposal, "reason", None) or "AI found new evidence that conflicts with a human-edited row.",
            }
        )
    return rows


def _corroboration_upload_text(upload: Any) -> str:
    summary = getattr(upload, "summary", None) or {}
    preview_text = summary.get("preview_text") or ""
    structured_text = json.dumps(summary.get("structured_extraction") or {}, default=json_safe_value)
    auto_text = json.dumps(summary.get("auto_extraction") or {}, default=json_safe_value)
    return "\n".join(part for part in (preview_text, structured_text, auto_text) if part).lower()


def _corroboration_tokens_for_item(item: Any) -> List[str]:
    tokens: List[str] = []
    for candidate in (
        getattr(item, "vendor", None),
        getattr(item, "item_name", None),
        getattr(item, "store_name", None),
        getattr(item, "details", None),
        getattr(item, "property_code", None),
    ):
        raw_value = _string_or_none(candidate)
        if not raw_value:
            continue
        tokens.extend(token.lower() for token in re.findall(r"[A-Za-z0-9@#]+", raw_value) if len(token) > 2)
    return list(dict.fromkeys(tokens))


def build_corroboration_rows(uploads: Sequence[Any], expense_items: Sequence[Any]) -> Dict[str, Any]:
    corroboration_uploads = [upload for upload in uploads if getattr(upload, "stage", None) == "corroboration"]
    if not corroboration_uploads:
        return {
            "upload_count": 0,
            "matched_count": 0,
            "unmatched_count": 0,
            "rows": [],
        }

    upload_texts = [
        (upload, _corroboration_upload_text(upload))
        for upload in corroboration_uploads
    ]
    rows: List[Dict[str, Any]] = []
    matched_count = 0

    for item in expense_items:
        amount = round(item.effective_total(), 2)
        amount_tokens = {f"{amount:.2f}", str(int(amount)) if float(amount).is_integer() else None}
        amount_tokens = {token for token in amount_tokens if token}
        matched_upload = None
        for upload, preview_text in upload_texts:
            if not preview_text:
                continue
            has_amount = any(token in preview_text for token in amount_tokens)
            has_text_match = any(token in preview_text for token in _corroboration_tokens_for_item(item))
            if has_amount and has_text_match:
                matched_upload = upload
                break

        if matched_upload:
            matched_count += 1
        rows.append(
            {
                "expense_item_id": getattr(item, "bookkeeping_expense_item_id", None),
                "status": "matched" if matched_upload else "unmatched",
                "category": item.category,
                "item": item.item_name or item.vendor or item.category.replace("_", " ").title(),
                "property": item.property_code,
                "amount": amount,
                "statement": matched_upload.original_filename if matched_upload else None,
                "statement_upload_id": matched_upload.bookkeeping_upload_id if matched_upload else None,
            }
        )

    return {
        "upload_count": len(corroboration_uploads),
        "matched_count": matched_count,
        "unmatched_count": max(0, len(rows) - matched_count),
        "rows": rows,
    }


def build_sheet_views(
    portfolio: Any,
    period: Any,
    uploads: Sequence[Any],
    revenue_items: Sequence[Any],
    expense_items: Sequence[Any],
    change_proposals: Optional[Sequence[Any]] = None,
) -> List[Dict[str, Any]]:
    revenue_rows_by_source = build_revenue_rows_by_source(revenue_items=revenue_items, uploads=uploads)
    configured_sources = configured_revenue_sources(portfolio)
    visible_revenue_sources = list(configured_sources)
    for source in REVENUE_SOURCES:
        has_source_data = bool(revenue_rows_by_source.get(source)) or bool(
            any(upload.stage == "revenue" and upload.source == source for upload in uploads)
        )
        if source not in visible_revenue_sources and has_source_data:
            visible_revenue_sources.append(source)
    review_queue_rows = build_review_queue_rows(revenue_items, expense_items, change_proposals)
    corroboration_state = build_corroboration_rows(uploads, expense_items)
    revenue_totals = aggregate_revenue_totals(uploads, revenue_items)
    expense_totals = aggregate_expense_totals(expense_items)
    processing_fee_total = sum(
        values.get("hostaway_fee_total", 0.0) + values.get("stripe_fee_total", 0.0)
        for values in revenue_totals.values()
    )
    configured_software_total = (
        float((portfolio.hostaway_price_per_listing or Decimal("0")) * (portfolio.listing_count or 0))
        + float((portfolio.pricelabs_price_per_listing or Decimal("0")) * (portfolio.listing_count or 0))
    )
    manual_software_total = expense_totals.get("software_fee", 0.0)
    direct_refund_total = expense_totals.get("direct_refund", 0.0)
    total_revenue = sum(values.get("gross_total", 0.0) for values in revenue_totals.values()) - direct_refund_total
    total_expenses = (
        expense_totals.get("cleaning", 0.0)
        + expense_totals.get("maintenance", 0.0)
        + expense_totals.get("supplies", 0.0)
        + expense_totals.get("misc", 0.0)
        + configured_software_total
        + manual_software_total
        + processing_fee_total
        + revenue_totals.get("booking_com", {}).get("commission_total", 0.0)
        + revenue_totals.get("vrbo", {}).get("commission_total", 0.0)
        + revenue_totals.get("hopper", {}).get("commission_total", 0.0)
    )
    management_fee = round(total_revenue * (float(portfolio.management_fee_percentage or 0) / 100.0), 2)

    tabs: List[Dict[str, Any]] = [
        {
            "key": "owner_statement",
            "label": "Owner Statement",
            "editable": False,
            "columns": [
                {"key": "section", "label": "Section"},
                {"key": "line_item", "label": "Line Item"},
                {"key": "amount", "label": "Amount"},
            ],
            "rows": [
                {"section": "Revenue", "line_item": "Total Revenue", "amount": round(total_revenue, 2)},
                {"section": "Expenses", "line_item": "Cleaning", "amount": round(expense_totals.get("cleaning", 0.0), 2)},
                {"section": "Expenses", "line_item": "Maintenance", "amount": round(expense_totals.get("maintenance", 0.0), 2)},
                {"section": "Expenses", "line_item": "Supplies", "amount": round(expense_totals.get("supplies", 0.0), 2)},
                {"section": "Expenses", "line_item": "Software", "amount": round(configured_software_total + manual_software_total + processing_fee_total, 2)},
                {"section": "Expenses", "line_item": "Misc", "amount": round(expense_totals.get("misc", 0.0), 2)},
                {"section": "Expenses", "line_item": "Direct Refunds", "amount": round(direct_refund_total, 2)},
                {"section": "Fees", "line_item": "Management Fee", "amount": management_fee},
                {"section": "Net", "line_item": "Net Income", "amount": round(total_revenue - total_expenses - management_fee, 2)},
            ],
        },
        {
            "key": "revenue_summary",
            "label": "Revenue Summary",
            "editable": False,
            "columns": [
                {"key": "channel", "label": "Channel"},
                {"key": "records", "label": "Records"},
                {"key": "gross_total", "label": "Gross"},
                {"key": "commission_total", "label": "Commission"},
                {"key": "hostaway_fee_total", "label": "Hostaway"},
                {"key": "stripe_fee_total", "label": "Stripe"},
            ],
            "rows": [
                {
                    "channel": source_label(source),
                    "records": len(revenue_rows_by_source.get(source, [])),
                    "gross_total": round(revenue_totals.get(source, {}).get("gross_total", 0.0), 2),
                    "commission_total": round(revenue_totals.get(source, {}).get("commission_total", 0.0), 2),
                    "hostaway_fee_total": round(revenue_totals.get(source, {}).get("hostaway_fee_total", 0.0), 2),
                    "stripe_fee_total": round(revenue_totals.get(source, {}).get("stripe_fee_total", 0.0), 2),
                }
                for source in visible_revenue_sources
            ],
        },
    ]

    tabs.append(
        {
            "key": "revenue_all",
            "label": "All Revenue",
            "editable": True,
            "row_type": "revenue_item",
            "columns": [{"key": key, "label": label} for key, label in REVENUE_COMMON_GRID_COLUMNS],
            "rows": (
                [
                    {
                        "row_id": item.bookkeeping_revenue_item_id,
                        "row_type": "revenue_item",
                        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
                        "source": source_label(item.source),
                        "reservation_identifier": item.reservation_identifier or item.confirmation_code,
                        "guest_name": item.guest_name,
                        "property_code": item.property_code or item.raw_listing_name,
                        "start_date": item.start_date.isoformat() if item.start_date else None,
                        "end_date": item.end_date.isoformat() if item.end_date else None,
                        "gross_amount": round(money_value(item.gross_amount), 2),
                        "commission_amount": round(money_value(item.commission_amount), 2),
                        "hostaway_fee_amount": round(money_value(item.hostaway_fee_amount), 2),
                        "stripe_fee_amount": round(money_value(item.stripe_fee_amount), 2),
                        "needs_review": bool(item.needs_review),
                    }
                    for item in sorted(revenue_items, key=lambda item: (item.source or "", item.row_index or 0, item.start_date or date.max))
                ]
                if revenue_items else
                [
                    {
                        "row_id": None,
                        "row_type": "revenue_item",
                        "updated_at": None,
                        "source": source_label(source),
                        "reservation_identifier": row.get(_revenue_field_header(source, "reservation_identifier")) or row.get(_revenue_field_header(source, "confirmation_code")),
                        "guest_name": row.get(_revenue_field_header(source, "guest_name")),
                        "property_code": row.get(_revenue_field_header(source, "property_code")),
                        "start_date": row.get(_revenue_field_header(source, "start_date")),
                        "end_date": row.get(_revenue_field_header(source, "end_date")),
                        "gross_amount": round(money_value(row.get(_revenue_field_header(source, "gross_amount"))), 2),
                        "commission_amount": round(money_value(row.get(_revenue_field_header(source, "commission_amount"))), 2),
                        "hostaway_fee_amount": round(money_value(row.get(_revenue_field_header(source, "hostaway_fee_amount"))), 2),
                        "stripe_fee_amount": round(money_value(row.get(_revenue_field_header(source, "stripe_fee_amount"))), 2),
                        "needs_review": False,
                    }
                    for source in visible_revenue_sources
                    for row in revenue_rows_by_source.get(source, [])
                    if _is_data_row_for_summary(source, row)
                ]
            ),
        }
    )

    for source in visible_revenue_sources:
        tabs.append(
            {
                "key": f"revenue_{source}",
                "label": source_label(source),
                "editable": True,
                "row_type": "revenue_item",
                "columns": [{"key": header, "label": header} for header in STANDARD_REVENUE_HEADERS.get(source, [])],
                "rows": (
                    [
                        {
                            **row,
                            "row_id": getattr(item, "bookkeeping_revenue_item_id", None),
                            "row_type": "revenue_item",
                            "updated_at": getattr(item, "updated_at", None).isoformat() if getattr(item, "updated_at", None) else None,
                        }
                        for item, row in [
                            (item, dict(getattr(item, "normalized_data", None) or getattr(item, "raw_data", None) or {}))
                            for item in sorted(
                                [revenue_item for revenue_item in revenue_items if revenue_item.source == source],
                                key=lambda revenue_item: (revenue_item.row_index or 0, revenue_item.start_date or date.max),
                            )
                        ]
                    ]
                    if revenue_items else
                    [
                        {
                            **row,
                            "row_id": None,
                            "row_type": "revenue_item",
                            "updated_at": None,
                        }
                        for row in revenue_rows_by_source.get(source, [])
                        if _is_data_row_for_summary(source, row)
                    ]
                ),
            }
        )

    expense_columns = [
        {"key": "service_date", "label": "Service Date"},
        {"key": "category", "label": "Category"},
        {"key": "item_name", "label": "Item"},
        {"key": "vendor", "label": "Vendor"},
        {"key": "property_code", "label": "Property"},
        {"key": "effective_total", "label": "Amount"},
        {"key": "payment_method", "label": "Payment Method"},
        {"key": "needs_review", "label": "Review"},
    ]
    tabs.append(
        {
            "key": "expenses_all",
            "label": "All Expenses",
            "editable": True,
            "row_type": "expense_item",
            "columns": expense_columns,
            "rows": [
                {
                    "row_id": item.bookkeeping_expense_item_id,
                    "row_type": "expense_item",
                    "updated_at": item.updated_at.isoformat() if item.updated_at else None,
                    "service_date": item.service_date.isoformat() if item.service_date else None,
                    "category": item.category,
                    "item_name": item.item_name,
                    "vendor": item.vendor,
                    "property_code": item.property_code,
                    "effective_total": round(item.effective_total(), 2),
                    "payment_method": item.payment_method,
                    "needs_review": bool(item.needs_review),
                }
                for item in sorted(expense_items, key=_expense_sort_key)
            ],
        }
    )

    for category in EXPENSE_CATEGORIES:
        if category == "software_fee":
            category_items = sorted(
                [item for item in expense_items if item.category == category],
                key=_expense_sort_key,
            )
            software_columns = [
                {"key": "entry_type", "label": "Entry Type"},
                {"key": "software_name", "label": "Software"},
                {"key": "vendor", "label": "Vendor"},
                {"key": "price_per_listing", "label": "Price Per Listing"},
                {"key": "listing_count", "label": "Listings"},
                {"key": "effective_total", "label": "Amount"},
                {"key": "property_code", "label": "Property"},
                {"key": "needs_review", "label": "Review"},
            ]
            software_rows: List[Dict[str, Any]] = []
            configured_rows = [
                (
                    "Hostaway",
                    float(portfolio.hostaway_price_per_listing or 0),
                    portfolio.listing_count or 0,
                    float((portfolio.hostaway_price_per_listing or Decimal("0")) * (portfolio.listing_count or 0)),
                ),
                (
                    "Pricelabs",
                    float(portfolio.pricelabs_price_per_listing or 0),
                    portfolio.listing_count or 0,
                    float((portfolio.pricelabs_price_per_listing or Decimal("0")) * (portfolio.listing_count or 0)),
                ),
            ]
            for software_name, price_per_listing, listing_count, total in configured_rows:
                if price_per_listing == 0 and total == 0:
                    continue
                software_rows.append(
                    {
                        "row_id": None,
                        "row_type": None,
                        "updated_at": None,
                        "entry_type": "Configured",
                        "software_name": software_name,
                        "vendor": "Portfolio setting",
                        "price_per_listing": round(price_per_listing, 2),
                        "listing_count": listing_count,
                        "effective_total": round(total, 2),
                        "property_code": getattr(portfolio, "code", None),
                        "needs_review": False,
                    }
                )

            software_rows.extend(
                [
                    {
                        "row_id": item.bookkeeping_expense_item_id,
                        "row_type": "expense_item",
                        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
                        "entry_type": "Manual",
                        "software_name": item.item_name or item.vendor or "Manual Software Fee",
                        "vendor": item.vendor,
                        "price_per_listing": None,
                        "listing_count": None,
                        "effective_total": round(item.effective_total(), 2),
                        "property_code": item.property_code,
                        "needs_review": bool(item.needs_review),
                    }
                    for item in category_items
                ]
            )
            tabs.append(
                {
                    "key": f"expense_{category}",
                    "label": category.replace("_", " ").title(),
                    "editable": any(row.get("row_id") for row in software_rows),
                    "row_type": "expense_item",
                    "columns": software_columns,
                    "rows": software_rows,
                }
            )
            continue

        category_items = [item for item in expense_items if item.category == category]
        tabs.append(
            {
                "key": f"expense_{category}",
                "label": category.replace("_", " ").title(),
                "editable": True,
                "row_type": "expense_item",
                "columns": expense_columns,
                "rows": [
                    {
                        "row_id": item.bookkeeping_expense_item_id,
                        "row_type": "expense_item",
                        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
                        "service_date": item.service_date.isoformat() if item.service_date else None,
                        "category": item.category,
                        "item_name": item.item_name,
                        "vendor": item.vendor,
                        "property_code": item.property_code,
                        "effective_total": round(item.effective_total(), 2),
                        "payment_method": item.payment_method,
                        "needs_review": bool(item.needs_review),
                    }
                    for item in sorted(category_items, key=_expense_sort_key)
                ],
            }
        )

    tabs.extend(
        [
            {
                "key": "corroboration",
                "label": "Corroboration",
                "editable": False,
                "columns": [
                    {"key": "status", "label": "Status"},
                    {"key": "category", "label": "Category"},
                    {"key": "item", "label": "Item"},
                    {"key": "property", "label": "Property"},
                    {"key": "amount", "label": "Amount"},
                    {"key": "statement", "label": "Statement"},
                ],
                "rows": corroboration_state["rows"],
            },
            {
                "key": "review_queue",
                "label": "Review Queue",
                "editable": False,
                "columns": [
                    {"key": "kind", "label": "Kind"},
                    {"key": "label", "label": "Label"},
                    {"key": "item", "label": "Item"},
                    {"key": "property", "label": "Property"},
                    {"key": "amount", "label": "Amount"},
                    {"key": "reason", "label": "Reason"},
                ],
                "rows": review_queue_rows,
            },
        ]
    )

    return tabs


def build_workspace_summary(
    portfolio: Any,
    period: Any,
    uploads: Sequence[Any],
    expense_items: Sequence[Any],
    revenue_items: Optional[Sequence[Any]] = None,
    listing_mappings: Optional[Sequence[Any]] = None,
    change_proposals: Optional[Sequence[Any]] = None,
    processing_batches: Optional[Sequence[Any]] = None,
    revisions: Optional[Sequence[Any]] = None,
) -> Dict[str, Any]:
    revenue_items = list(revenue_items or [])
    listing_mappings = list(listing_mappings or [])
    change_proposals = list(change_proposals or [])
    processing_batches = list(processing_batches or [])
    revisions = list(revisions or [])
    revenue_totals = aggregate_revenue_totals(uploads, revenue_items)
    expense_totals = aggregate_expense_totals(expense_items)
    owner_share = 1.0
    management_fee_rate = float(portfolio.management_fee_percentage or 0) / 100.0
    configured_sources = configured_revenue_sources(portfolio)

    revenue_checklist = []
    revenue_completed = 0
    active_revenue_sources = list(configured_sources)
    for source in REVENUE_SOURCES:
        if source in active_revenue_sources:
            continue
        has_source_uploads = any(upload.stage == "revenue" and upload.source == source for upload in uploads)
        has_source_rows = any(getattr(item, "source", None) == source for item in revenue_items)
        if has_source_uploads or has_source_rows:
            active_revenue_sources.append(source)

    for source in active_revenue_sources:
        source_uploads = [upload for upload in uploads if upload.stage == "revenue" and upload.source == source]
        source_rows = [item for item in revenue_items if getattr(item, "source", None) == source]
        source_total = revenue_totals.get(source, {}).get("gross_total", 0.0)
        is_expected = source in configured_sources
        is_completed = bool(source_uploads or source_rows)
        if is_expected and is_completed:
            revenue_completed += 1
        if source_uploads:
            status = "uploaded"
        elif source_rows:
            status = "captured"
        elif is_expected:
            status = "missing"
        else:
            status = "unexpected"
        revenue_checklist.append(
            {
                "source": source,
                "label": source_label(source),
                "completed": is_completed,
                "expected": is_expected,
                "status": status,
                "missing": bool(is_expected and not is_completed),
                "upload_count": len(source_uploads),
                "row_count": len(source_rows),
                "gross_total": source_total,
            }
        )

    processing_fee_total = sum(
        values.get("hostaway_fee_total", 0.0) + values.get("stripe_fee_total", 0.0)
        for values in revenue_totals.values()
    )

    configured_software_total = (
        float((portfolio.hostaway_price_per_listing or Decimal("0")) * (portfolio.listing_count or 0))
        + float((portfolio.pricelabs_price_per_listing or Decimal("0")) * (portfolio.listing_count or 0))
    )
    manual_software_total = expense_totals.get("software_fee", 0.0)
    direct_refund_total = expense_totals.get("direct_refund", 0.0)
    total_revenue = sum(values.get("gross_total", 0.0) for values in revenue_totals.values()) - direct_refund_total
    expense_base_total = (
        expense_totals.get("cleaning", 0.0)
        + expense_totals.get("maintenance", 0.0)
        + expense_totals.get("supplies", 0.0)
        + expense_totals.get("misc", 0.0)
        + configured_software_total
        + manual_software_total
        + processing_fee_total
        + revenue_totals.get("booking_com", {}).get("commission_total", 0.0)
        + revenue_totals.get("vrbo", {}).get("commission_total", 0.0)
        + revenue_totals.get("hopper", {}).get("commission_total", 0.0)
    )
    owner_revenue = round(total_revenue * owner_share, 2)
    owner_expenses = round(expense_base_total * owner_share, 2)
    management_fee = round(owner_revenue * management_fee_rate, 2)

    return {
        "portfolio": _entity_to_dict(portfolio),
        "period": _entity_to_dict(period),
        "revenue_channels": configured_sources,
        "revenue_checklist": revenue_checklist,
        "revenue_progress": {
            "completed": revenue_completed,
            "total": len(configured_sources),
            "missing": max(0, len(configured_sources) - revenue_completed),
        },
        "uploads": [_workspace_upload_to_dict(upload) for upload in uploads],
        "listing_mappings": [_entity_to_dict(mapping) for mapping in listing_mappings],
        "revenue_items": [_workspace_revenue_item_to_dict(item) for item in revenue_items],
        "expense_items": [_workspace_expense_item_to_dict(item) for item in expense_items],
        "change_proposals": [_entity_to_dict(proposal) for proposal in change_proposals],
        "processing_batches": [
            _entity_to_dict(batch)
            for batch in processing_batches
            if getattr(batch, "status", None) in {"queued", "processing", "completed_with_errors"}
        ],
        "workspace_revisions": [
            {
                "bookkeeping_workspace_revision_id": getattr(revision, "bookkeeping_workspace_revision_id", None),
                "period_id": getattr(revision, "period_id", None),
                "status": getattr(revision, "status", None),
                "workbook_filename": getattr(revision, "workbook_filename", None),
                "created_by": getattr(revision, "created_by", None),
                "created_at": getattr(getattr(revision, "created_at", None), "isoformat", lambda: None)(),
            }
            for revision in revisions
        ],
        "revenue_totals": revenue_totals,
        "expense_totals": expense_totals,
        "corroboration_state": build_corroboration_rows(uploads, expense_items),
        "sheet_views": build_sheet_views(portfolio, period, uploads, revenue_items, expense_items, change_proposals),
        "review_queue": build_review_queue_rows(revenue_items, expense_items, change_proposals),
        "summary_cards": {
            "total_revenue": round(total_revenue, 2),
            "owner_revenue": owner_revenue,
            "owner_expenses": owner_expenses,
            "management_fee": management_fee,
            "net_income": round(owner_revenue - owner_expenses - management_fee, 2),
            "expense_items_needing_review": len([item for item in expense_items if item.needs_review]),
            "revenue_items_needing_review": len([item for item in revenue_items if item.needs_review]),
            "pending_change_proposals": len([proposal for proposal in change_proposals if proposal.status == "pending"]),
            "supporting_uploads": len([upload for upload in uploads if upload.stage != "revenue"]),
        },
        "software_totals": {
            "configured": round(configured_software_total, 2),
            "manual": round(manual_software_total, 2),
            "processing_fees": round(processing_fee_total, 2),
        },
    }


def build_workspace_revision_snapshot(
    portfolio: Any,
    period: Any,
    uploads: Sequence[Any],
    expense_items: Sequence[Any],
    revenue_items: Optional[Sequence[Any]] = None,
    listing_mappings: Optional[Sequence[Any]] = None,
    change_proposals: Optional[Sequence[Any]] = None,
    processing_batches: Optional[Sequence[Any]] = None,
) -> Dict[str, Any]:
    revenue_items = list(revenue_items or [])
    expense_items = list(expense_items or [])
    listing_mappings = list(listing_mappings or [])
    change_proposals = list(change_proposals or [])
    processing_batches = list(processing_batches or [])

    revenue_totals = aggregate_revenue_totals(uploads, revenue_items)
    expense_totals = aggregate_expense_totals(expense_items)
    configured_sources = configured_revenue_sources(portfolio)

    active_revenue_sources = list(configured_sources)
    for source in REVENUE_SOURCES:
        if source in active_revenue_sources:
            continue
        has_source_uploads = any(upload.stage == "revenue" and upload.source == source for upload in uploads)
        has_source_rows = any(getattr(item, "source", None) == source for item in revenue_items)
        if has_source_uploads or has_source_rows:
            active_revenue_sources.append(source)

    revenue_completed = 0
    revenue_checklist = []
    for source in active_revenue_sources:
        source_uploads = [upload for upload in uploads if upload.stage == "revenue" and upload.source == source]
        source_rows = [item for item in revenue_items if getattr(item, "source", None) == source]
        is_expected = source in configured_sources
        is_completed = bool(source_uploads or source_rows)
        if is_expected and is_completed:
            revenue_completed += 1
        revenue_checklist.append(
            {
                "source": source,
                "label": source_label(source),
                "expected": is_expected,
                "completed": is_completed,
                "upload_count": len(source_uploads),
                "row_count": len(source_rows),
                "gross_total": round(revenue_totals.get(source, {}).get("gross_total", 0.0), 2),
                "commission_total": round(revenue_totals.get(source, {}).get("commission_total", 0.0), 2),
                "hostaway_fee_total": round(revenue_totals.get(source, {}).get("hostaway_fee_total", 0.0), 2),
                "stripe_fee_total": round(revenue_totals.get(source, {}).get("stripe_fee_total", 0.0), 2),
            }
        )

    processing_fee_total = sum(
        values.get("hostaway_fee_total", 0.0) + values.get("stripe_fee_total", 0.0)
        for values in revenue_totals.values()
    )
    configured_software_total = (
        float((portfolio.hostaway_price_per_listing or Decimal("0")) * (portfolio.listing_count or 0))
        + float((portfolio.pricelabs_price_per_listing or Decimal("0")) * (portfolio.listing_count or 0))
    )
    manual_software_total = expense_totals.get("software_fee", 0.0)
    direct_refund_total = expense_totals.get("direct_refund", 0.0)
    total_revenue = sum(values.get("gross_total", 0.0) for values in revenue_totals.values()) - direct_refund_total
    total_expenses = (
        expense_totals.get("cleaning", 0.0)
        + expense_totals.get("maintenance", 0.0)
        + expense_totals.get("supplies", 0.0)
        + expense_totals.get("misc", 0.0)
        + configured_software_total
        + manual_software_total
        + processing_fee_total
        + revenue_totals.get("booking_com", {}).get("commission_total", 0.0)
        + revenue_totals.get("vrbo", {}).get("commission_total", 0.0)
        + revenue_totals.get("hopper", {}).get("commission_total", 0.0)
    )
    management_fee = round(total_revenue * (float(portfolio.management_fee_percentage or 0) / 100.0), 2)

    portfolio_data = _entity_to_dict(portfolio)
    period_data = _entity_to_dict(period)
    upload_counts = Counter((getattr(upload, "stage", None) or "unknown") for upload in uploads)
    upload_status_counts = Counter((getattr(upload, "upload_status", None) or "unknown") for upload in uploads)

    return {
        "portfolio": {
            key: portfolio_data.get(key)
            for key in (
                "bookkeeping_portfolio_id",
                "name",
                "code",
                "listing_tag",
                "portfolio_tag",
                "property_name",
                "property_address",
                "listing_count",
                "management_fee_percentage",
                "revenue_channels",
            )
            if portfolio_data.get(key) not in (None, "", [], {})
        },
        "period": {
            key: period_data.get(key)
            for key in (
                "bookkeeping_period_id",
                "name",
                "status",
                "period_start",
                "period_end",
            )
            if period_data.get(key) not in (None, "", [], {})
        },
        "revenue_progress": {
            "completed": revenue_completed,
            "total": len(configured_sources),
            "missing": max(0, len(configured_sources) - revenue_completed),
        },
        "revenue_checklist": revenue_checklist,
        "revenue_totals": revenue_totals,
        "expense_totals": expense_totals,
        "summary_cards": {
            "total_revenue": round(total_revenue, 2),
            "owner_revenue": round(total_revenue, 2),
            "owner_expenses": round(total_expenses, 2),
            "management_fee": management_fee,
            "net_income": round(total_revenue - total_expenses - management_fee, 2),
            "expense_items_needing_review": len([item for item in expense_items if getattr(item, "needs_review", False)]),
            "revenue_items_needing_review": len([item for item in revenue_items if getattr(item, "needs_review", False)]),
            "pending_change_proposals": len([proposal for proposal in change_proposals if getattr(proposal, "status", None) == "pending"]),
            "supporting_uploads": len([upload for upload in uploads if getattr(upload, "stage", None) != "revenue"]),
        },
        "software_totals": {
            "configured": round(configured_software_total, 2),
            "manual": round(manual_software_total, 2),
            "processing_fees": round(processing_fee_total, 2),
        },
        "upload_counts": {
            "total": len(uploads),
            "revenue": upload_counts.get("revenue", 0),
            "expense": upload_counts.get("expense", 0),
            "corroboration": upload_counts.get("corroboration", 0),
        },
        "upload_status_counts": dict(upload_status_counts),
        "listing_mapping_count": len(listing_mappings),
        "processing_batches": [
            {
                "bookkeeping_processing_batch_id": getattr(batch, "bookkeeping_processing_batch_id", None),
                "stage": getattr(batch, "stage", None),
                "status": getattr(batch, "status", None),
                "total_uploads": getattr(batch, "total_uploads", None),
                "processed_uploads": getattr(batch, "processed_uploads", None),
                "successful_uploads": getattr(batch, "successful_uploads", None),
                "failed_uploads": getattr(batch, "failed_uploads", None),
                "created_at": getattr(getattr(batch, "created_at", None), "isoformat", lambda: None)(),
            }
            for batch in processing_batches[-8:]
        ],
    }


def build_agent_context(
    portfolio: Any,
    period: Any,
    uploads: Sequence[Any],
    expense_items: Sequence[Any],
    revenue_items: Optional[Sequence[Any]] = None,
    listing_mappings: Optional[Sequence[Any]] = None,
    change_proposals: Optional[Sequence[Any]] = None,
) -> Dict[str, Any]:
    summary = build_workspace_summary(
        portfolio,
        period,
        uploads,
        expense_items,
        revenue_items=revenue_items,
        listing_mappings=listing_mappings,
        change_proposals=change_proposals,
    )
    preview_uploads = []
    for upload in uploads[:12]:
        preview_uploads.append(
            {
                "stage": upload.stage,
                "source": upload.source,
                "filename": upload.original_filename,
                "sheet_name": upload.sheet_name,
                "summary": upload.summary or {},
                "preview_rows": (upload.preview_rows or [])[:3],
            }
        )

    preview_revenue = []
    for item in (revenue_items or [])[:25]:
        preview_revenue.append(
            {
                "source": item.source,
                "reservation_identifier": item.reservation_identifier,
                "guest_name": item.guest_name,
                "property_code": item.property_code,
                "gross_amount": money_value(item.gross_amount),
                "needs_review": item.needs_review,
                "review_reason": item.review_reason,
            }
        )

    preview_expenses = []
    for item in expense_items[:25]:
        preview_expenses.append(
            {
                "category": item.category,
                "vendor": item.vendor,
                "item_name": item.item_name,
                "property_code": item.property_code,
                "effective_total": item.effective_total(),
                "needs_review": item.needs_review,
                "review_reason": item.review_reason,
                "details": item.details,
            }
        )

    return {
        "workspace_summary": summary["summary_cards"],
        "revenue_totals": summary["revenue_totals"],
        "expense_totals": summary["expense_totals"],
        "revenue_checklist": summary["revenue_checklist"],
        "listing_mappings": [mapping.to_dict() for mapping in (listing_mappings or [])[:25]],
        "uploads": preview_uploads,
        "revenue_items": preview_revenue,
        "expense_items": preview_expenses,
        "pending_change_proposals": [proposal.to_dict() for proposal in (change_proposals or [])[:10] if proposal.status == "pending"],
    }


def _write_sheet_headers(sheet, headers: Sequence[str], fill_color: str = "16324F") -> None:
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"


def _auto_width(sheet, max_width: int = 32) -> None:
    for column_cells in sheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        if not values:
            continue
        width = min(max(len(value) for value in values) + 2, max_width)
        sheet.column_dimensions[column_cells[0].column_letter].width = width


def _add_table(sheet, table_name: str, ref: str) -> None:
    if ref.split(":")[0] == ref.split(":")[1]:
        return
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _display_payment_method_label(value: Any) -> Optional[str]:
    raw_value = _string_or_none(value)
    if not raw_value:
        return None
    normalized = normalize_property_token(raw_value)
    mapping = {
        "BANKTRANSFER": "Bank Transfer",
        "DEBIT": "Debit Card",
        "DEBITCARD": "Debit Card",
        "CREDITCARD": "Credit Card",
        "ONLINE_RETAIL": "Online",
        "ONLINERETAIL": "Online",
    }
    return mapping.get(normalized, raw_value)


def _venmo_handle_for_item(item: Any) -> Optional[str]:
    payment_method = normalize_property_token(getattr(item, "payment_method", None))
    if payment_method != "VENMO":
        return None
    explicit_handle = _string_or_none(getattr(item, "account_number", None))
    if explicit_handle and explicit_handle.startswith("@"):
        return explicit_handle
    payee_name = _normalized_person_label(getattr(item, "account_holder", None) or getattr(item, "vendor", None))
    if not payee_name:
        return explicit_handle if explicit_handle and not _looks_masked_account_number(explicit_handle) else None
    compressed = re.sub(r"[^A-Za-z0-9]+", "", payee_name)
    return f"@{compressed}" if compressed else (explicit_handle if explicit_handle and not _looks_masked_account_number(explicit_handle) else None)


def _supply_purchase_type_label(value: Any) -> Optional[str]:
    normalized = normalize_property_token(value)
    mapping = {
        "RETAIL": "Store Purchase",
        "STOREPURCHASE": "Store Purchase",
        "INSTORE": "Store Purchase",
        "ONLINE": "Online",
        "ONLINERETAIL": "Online",
        "ONLINEORDER": "Online",
    }
    return mapping.get(normalized, _string_or_none(value))


def _compact_supply_item_name(value: Any) -> Optional[str]:
    raw_value = _string_or_none(value)
    if not raw_value:
        return None
    lowered = raw_value.lower()
    if "shampoo" in lowered:
        return "Shampoo"
    if "conditioner" in lowered:
        return "Conditioner"
    if "water" in lowered and "mineral" in lowered:
        return "Purified Water Minerals"
    if "propane" in lowered and "tank" in lowered:
        return "Propane New Spare Tank"
    return raw_value


def _supply_vendor_display(item: Any, upload_by_id: Optional[Dict[int, Any]] = None) -> Optional[str]:
    if getattr(item, "reimbursement_method", None):
        return _normalized_person_label(getattr(item, "account_holder", None) or getattr(item, "vendor", None))
    if upload_by_id:
        upload = upload_by_id.get(getattr(item, "upload_id", None))
        uploader = getattr(upload, "uploader", None) if upload else None
        uploader_name = _normalized_person_label(getattr(uploader, "name", None))
        if uploader_name:
            return uploader_name
    return _normalized_person_label(getattr(item, "account_holder", None) or getattr(item, "vendor", None))


def _display_payment_date(item: Any, upload_by_id: Optional[Dict[int, Any]] = None) -> Optional[date]:
    payment_date = getattr(item, "payment_date", None)
    if payment_date is None or not upload_by_id:
        return payment_date
    upload = upload_by_id.get(getattr(item, "upload_id", None))
    structured = ((getattr(upload, "summary", None) or {}).get("structured_extraction")) or {}
    if not _is_payment_screenshot_document(structured.get("document_type")):
        return payment_date
    filename_dates = _parse_date_sequence_from_filename(getattr(upload, "original_filename", None))
    if len(filename_dates) < 2:
        return payment_date
    if payment_date == getattr(item, "service_date", None):
        return None
    if filename_dates and payment_date <= min(filename_dates):
        return None
    return payment_date


def _row_needs_review(item: Any) -> bool:
    return bool(getattr(item, "needs_review", False) or getattr(item, "review_reason", None))


def _apply_review_highlight(sheet, row_index: int, max_column: int, review_reason: Optional[str] = None) -> None:
    review_fill = PatternFill("solid", fgColor=REVIEW_ROW_FILL_COLOR)
    for column_index in range(1, max_column + 1):
        sheet.cell(row=row_index, column=column_index).fill = review_fill
    if review_reason:
        sheet.cell(row=row_index, column=max_column).comment = Comment(review_reason, "Bookkeeping AI")


def normalized_revenue_upload_payload(upload: Any) -> Dict[str, Any]:
    cached = getattr(upload, "_normalized_revenue_payload", None)
    if cached is not None:
        return cached

    payload = {
        "source": getattr(upload, "source", None),
        "headers": list(getattr(upload, "headers", None) or []),
        "rows": list(getattr(upload, "parsed_rows", None) or []),
        "summary": dict(getattr(upload, "summary", None) or {}),
        "sheet_name": getattr(upload, "sheet_name", None),
    }

    needs_reparse = (
        getattr(upload, "stage", None) == "revenue"
        and getattr(upload, "source", None) in CSV_FALLBACK_HEADERS
        and payload["headers"]
        and payload["headers"][0] != CSV_FALLBACK_HEADERS[getattr(upload, "source")][0]
    )

    if needs_reparse:
        try:
            file_bytes = get_upload_absolute_path(getattr(upload, "stored_path")).read_bytes()
            reparsed = parse_revenue_file(
                file_bytes,
                getattr(upload, "original_filename", ""),
                getattr(upload, "source", None),
            )[0]
            payload = {
                "source": reparsed["source"],
                "headers": reparsed.get("headers") or [],
                "rows": reparsed.get("rows") or [],
                "summary": reparsed.get("summary") or {},
                "sheet_name": reparsed.get("sheet_name"),
            }
        except Exception:
            payload = {
                "source": getattr(upload, "source", None),
                "headers": list(getattr(upload, "headers", None) or []),
                "rows": list(getattr(upload, "parsed_rows", None) or []),
                "summary": dict(getattr(upload, "summary", None) or {}),
                "sheet_name": getattr(upload, "sheet_name", None),
            }

    setattr(upload, "_normalized_revenue_payload", payload)
    return payload


def _merged_headers_and_rows(uploads: Sequence[Any]) -> Tuple[List[str], List[Dict[str, Any]]]:
    headers: List[str] = []
    rows: List[Dict[str, Any]] = []
    for upload in uploads:
        payload = normalized_revenue_upload_payload(upload) if getattr(upload, "stage", None) == "revenue" else {
            "headers": upload.headers or [],
            "rows": upload.parsed_rows or [],
        }
        for header in payload["headers"] or []:
            if header and header not in headers:
                headers.append(header)
        rows.extend(payload["rows"] or [])
    return headers, rows


def infer_reporting_period_start(period: Any, uploads: Sequence[Any], expense_items: Sequence[Any]) -> date:
    month_counts: Dict[Tuple[int, int], int] = defaultdict(int)

    date_fields_by_source = {
        "airbnb": ("Arriving by date", "Start date", "Date"),
        "booking_com": ("Arrival",),
        "vrbo": ("Check-In Date",),
        "hopper": ("Check-In Date",),
        "direct_bookings": ("Check-In Date",),
        "google": ("Check-In Date",),
        "direct_refund": ("Date",),
    }

    for upload in uploads:
        if upload.stage != "revenue":
            continue
        payload = normalized_revenue_upload_payload(upload)
        for row in payload["rows"] or []:
            fields = date_fields_by_source.get(upload.source, tuple(row.keys()))
            parsed_date = None
            for field in fields:
                parsed_date = parse_date_or_none(row.get(field))
                if parsed_date:
                    break
            if parsed_date:
                month_counts[(parsed_date.year, parsed_date.month)] += 1

    for item in expense_items:
        for candidate in (item.service_date, item.payment_date):
            if candidate:
                month_counts[(candidate.year, candidate.month)] += 1

    if not month_counts:
        return period.period_start

    year, month = max(month_counts.items(), key=lambda item: item[1])[0]
    inferred = date(year, month, 1)
    return inferred


def _coerce_revenue_cell_value(header: str, value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, str) and value.startswith("="):
        return value
    if header in REVENUE_DATE_HEADERS:
        parsed_date = parse_date_or_none(value)
        if parsed_date:
            return datetime(parsed_date.year, parsed_date.month, parsed_date.day)
    if header in REVENUE_PERCENT_HEADERS:
        return percentage_value(value)
    if header in REVENUE_INTEGER_HEADERS:
        numeric = money_value(value)
        if numeric or str(value).strip() == "0":
            return int(round(numeric))
    if header in REVENUE_MONEY_HEADERS:
        return round(money_value(value), 2)
    return value


def _write_standardized_revenue_sheet(sheet, source: str, rows: Sequence[Dict[str, Any]]) -> None:
    headers = STANDARD_REVENUE_HEADERS.get(source)
    if not headers:
        merged_headers, merged_rows = _merged_headers_and_rows([])
        if not merged_headers:
            _write_sheet_headers(sheet, ["No uploaded records yet"])
            return
        _write_sheet_headers(sheet, merged_headers)
        for row_index, row in enumerate(merged_rows, start=2):
            for column_index, header in enumerate(merged_headers, start=1):
                sheet.cell(row=row_index, column=column_index, value=row.get(header))
        return

    _write_sheet_headers(sheet, headers)
    data_rows = [row for row in rows if _is_data_row_for_summary(source, row)]
    current_row = 2

    if source == "google" and not data_rows:
        current_row = 4
    else:
        for row in data_rows:
            for column_index, header in enumerate(headers, start=1):
                value = _coerce_revenue_cell_value(header, row.get(header))
                if source == "booking_com" and header == "Commission amount":
                    value = f"=G{current_row}*E{current_row}"
                elif source == "vrbo" and header == "Total Price":
                    value = f"=SUM(G{current_row}:K{current_row})"
                elif source == "vrbo" and header == "Vrbo Commission":
                    value = f"=(G{current_row}+I{current_row})*5%"
                elif source == "hopper" and header == "Total Price":
                    value = f"=SUM(G{current_row}:K{current_row})"
                elif source == "direct_bookings" and header == "Total Price":
                    value = f"=SUM(G{current_row}:K{current_row})"
                elif source == "direct_bookings" and header == "Stripe processing fees" and value is None:
                    value = f"=L{current_row}*0.029+0.3"
                sheet.cell(row=current_row, column=column_index, value=value)
            current_row += 1

    if source == "booking_com" and data_rows:
        sheet.cell(row=current_row, column=6, value=f"=SUM(F2:F{current_row - 1})")
        sheet.cell(row=current_row, column=7, value=f"=SUM(G2:G{current_row - 1})")
        sheet.cell(row=current_row, column=8, value=f"=SUM(H2:H{current_row - 1})")
    elif source == "vrbo" and data_rows:
        for column_letter in ("G", "H", "I", "J", "K", "L", "M", "N", "O"):
            sheet[f"{column_letter}{current_row}"] = f"=SUM({column_letter}2:{column_letter}{current_row - 1})"
    elif source == "hopper":
        start_row = 2
        footer_row = current_row
        for column_letter in ("G", "H", "I", "J", "K", "L"):
            sheet[f"{column_letter}{footer_row}"] = f"=SUM({column_letter}{start_row}:{column_letter}{max(start_row, footer_row - 1)})"
    elif source == "direct_bookings":
        start_row = 2
        footer_row = current_row
        for column_letter in ("G", "H", "I", "J", "K", "L", "M", "N"):
            sheet[f"{column_letter}{footer_row}"] = f"=SUM({column_letter}{start_row}:{column_letter}{max(start_row, footer_row - 1)})"
    elif source == "google":
        footer_row = current_row
        for column_letter in ("G", "H", "I", "J", "K", "L", "M"):
            sheet[f"{column_letter}{footer_row}"] = f"=SUM({column_letter}2:{column_letter}{footer_row - 1})"


def _expense_sort_key(item: Any) -> Tuple[Any, Any, Any, Any, Any, Any]:
    line_index = _item_extraction_field(item, "line_index", 0)
    return (
        item.service_date or date.max,
        item.payment_date or date.max,
        getattr(item, "upload_id", None) if getattr(item, "upload_id", None) is not None else 10**9,
        line_index if line_index is not None else 10**9,
        (item.property_code or ""),
        getattr(item, "bookkeeping_expense_item_id", 0) or 0,
    )


def _expense_link_label(item: Any, fallback: str = "-") -> str:
    return RECEIPT_LINK_TEXT if getattr(item, "upload_id", None) else fallback


def _upload_lookup(uploads: Sequence[Any]) -> Dict[int, Any]:
    return {
        int(getattr(upload, "bookkeeping_upload_id")): upload
        for upload in uploads
        if getattr(upload, "bookkeeping_upload_id", None) is not None
    }


def _expense_link_url(item: Any, upload_by_id: Optional[Dict[int, Any]] = None) -> Optional[str]:
    if not upload_by_id:
        return None
    upload = upload_by_id.get(getattr(item, "upload_id", None))
    if not upload:
        return None
    return upload_drive_file_url(upload)


def _reimbursement_link_url(item: Any, upload_by_id: Optional[Dict[int, Any]] = None) -> Optional[str]:
    if not upload_by_id:
        return None
    reimbursement_upload_id = _item_extraction_field(item, "reimbursement_upload_id")
    if reimbursement_upload_id in (None, ""):
        return None
    upload = upload_by_id.get(int(reimbursement_upload_id))
    if not upload:
        return None
    return upload_drive_file_url(upload)


def _set_link_cell(cell: Any, url: Optional[str], label: str = RECEIPT_LINK_TEXT) -> None:
    if not url:
        return
    cell.value = label
    cell.hyperlink = url
    cell.style = "Hyperlink"


def _item_extraction_field(item: Any, key: str, default: Any = None) -> Any:
    extraction_data = getattr(item, "extraction_data", None) or {}
    if key in extraction_data:
        return extraction_data.get(key, default)
    shared = extraction_data.get("shared_fields") or {}
    return shared.get(key, default)


def _sheet_ref(sheet_name: str, cell_ref: str) -> str:
    return f"='{sheet_name}'!{cell_ref}"


def _previous_month_label(report_period_start: date) -> str:
    if report_period_start.month == 1:
        return f"December {report_period_start.year - 1}"
    previous_month_date = report_period_start.replace(month=report_period_start.month - 1)
    return logical_month_label(previous_month_date)


def _style_range(
    sheet: Any,
    cell_range: str,
    *,
    value: Any = None,
    fill_color: Optional[str] = None,
    font: Optional[Font] = None,
    alignment: Optional[Alignment] = None,
    border: Optional[Border] = None,
    number_format: Optional[str] = None,
    merge: bool = True,
) -> Any:
    start_cell = sheet[cell_range.split(":")[0]]
    if merge:
        sheet.merge_cells(cell_range)
    if value is not None:
        start_cell.value = value
    if font is not None:
        start_cell.font = font
    if alignment is not None:
        start_cell.alignment = alignment
    if fill_color:
        fill = PatternFill("solid", fgColor=fill_color)
    else:
        fill = None

    start_row = start_cell.row
    start_col = start_cell.column
    end_cell = sheet[cell_range.split(":")[1]]
    for row in sheet.iter_rows(
        min_row=start_row,
        max_row=end_cell.row,
        min_col=start_col,
        max_col=end_cell.column,
    ):
        for cell in row:
            if fill is not None:
                cell.fill = fill
            if border is not None:
                cell.border = border
            if alignment is not None:
                cell.alignment = alignment
            if number_format:
                cell.number_format = number_format
    return start_cell


def _build_standard_owner_statement_sheet(
    workbook: Any,
    portfolio: Any,
    month_label: str,
    revenue_totals: Dict[str, Dict[str, Any]],
    expense_totals: Dict[str, float],
    direct_refund_total: float,
    management_fee_rate: float,
    software_total_base: float,
) -> Any:
    owner_sheet = workbook.create_sheet(title=f"Owner Statement ({month_label})"[:31])
    for column_letter, width in OWNER_STATEMENT_COLUMN_WIDTHS.items():
        owner_sheet.column_dimensions[column_letter].width = width

    title_font = Font(size=18, bold=True, color="FFFFFFFF")
    header_font = Font(size=11, bold=True, color="FFFFFFFF")
    section_font = Font(size=11, bold=True)
    body_font = Font(size=11)
    body_value_font = Font(size=11, bold=True)
    center = Alignment(horizontal="center", vertical="center")

    _style_range(
        owner_sheet,
        "A1:N2",
        value=f"OWNER STATEMENT ({getattr(portfolio, 'code', None) or getattr(portfolio, 'name', None) or 'Portfolio'})",
        fill_color=OWNER_STATEMENT_HEADER_FILL,
        font=title_font,
        alignment=center,
        border=THIN_BLACK_BORDER,
    )
    _style_range(
        owner_sheet,
        "B3:M3",
        value=f"For {month_label}",
        fill_color=OWNER_STATEMENT_HEADER_FILL,
        font=header_font,
        alignment=center,
    )
    _style_range(
        owner_sheet,
        "B5:E5",
        value="PROPERTY NAME",
        fill_color=OWNER_STATEMENT_HEADER_FILL,
        font=header_font,
        alignment=center,
        border=THIN_BLACK_BORDER,
    )
    _style_range(
        owner_sheet,
        "G5:M5",
        value="PROPERTY ADDRESS",
        fill_color=OWNER_STATEMENT_HEADER_FILL,
        font=header_font,
        alignment=center,
        border=THIN_BLACK_BORDER,
    )
    _style_range(
        owner_sheet,
        "B6:E8",
        value=getattr(portfolio, "property_name", None) or getattr(portfolio, "name", None) or "",
        font=body_font,
        alignment=center,
        border=THIN_BLACK_BORDER,
    )
    _style_range(
        owner_sheet,
        "G6:M8",
        value=getattr(portfolio, "property_address", None) or "",
        font=body_font,
        alignment=center,
        border=THIN_BLACK_BORDER,
    )

    owner_sheet["B10"] = "Revenue Summary"
    owner_sheet["B10"].font = section_font

    revenue_rows = [
        (11, "Airbnb", revenue_totals.get("airbnb", {}).get("gross_total", 0.0)),
        (12, "Booking.com", revenue_totals.get("booking_com", {}).get("gross_total", 0.0)),
        (13, "Vrbo", revenue_totals.get("vrbo", {}).get("gross_total", 0.0)),
        (14, "Hopper Homes", revenue_totals.get("hopper", {}).get("gross_total", 0.0)),
        (15, "Direct Bookings", revenue_totals.get("direct_bookings", {}).get("gross_total", 0.0)),
        (16, "Google", revenue_totals.get("google", {}).get("gross_total", 0.0)),
        (17, "Direct Refunds", -direct_refund_total),
    ]
    for row_index, label, value in revenue_rows:
        _style_range(
            owner_sheet,
            f"B{row_index}:K{row_index}",
            value=label,
            fill_color=OWNER_STATEMENT_BAND_FILL,
            font=body_font,
            alignment=center,
            border=THIN_BLACK_BORDER,
        )
        _style_range(
            owner_sheet,
            f"L{row_index}:M{row_index}",
            value=value,
            fill_color=OWNER_STATEMENT_BAND_FILL,
            font=body_font,
            alignment=center,
            border=THIN_BLACK_BORDER,
            number_format=OWNER_STATEMENT_VALUE_NUMBER_FORMAT,
        )

    _style_range(
        owner_sheet,
        "B19:K19",
        value="TOTAL REVENUE",
        fill_color=OWNER_STATEMENT_BAND_FILL,
        font=section_font,
        alignment=center,
        border=THIN_BLACK_BORDER,
    )
    _style_range(
        owner_sheet,
        "L19:M19",
        value="=SUM(L11:M17)",
        fill_color=OWNER_STATEMENT_BAND_FILL,
        font=section_font,
        alignment=center,
        border=THIN_BLACK_BORDER,
        number_format=OWNER_STATEMENT_VALUE_NUMBER_FORMAT,
    )

    owner_sheet["B21"] = "Expense Summary"
    owner_sheet["B21"].font = section_font

    expense_rows = [
        (22, "Cleaning Fees", expense_totals.get("cleaning", 0.0)),
        (23, "Maintenance", expense_totals.get("maintenance", 0.0)),
        (24, "Supplies", expense_totals.get("supplies", 0.0)),
        (25, "Software Fee", software_total_base),
        (26, "Misc. Expenses", expense_totals.get("misc", 0.0)),
        (27, "Management Fee", f"=L19*{management_fee_rate}"),
        (28, f"Booking.com Commissions - {month_label}", revenue_totals.get("booking_com", {}).get("commission_total", 0.0)),
        (29, "Vrbo Commissions", revenue_totals.get("vrbo", {}).get("commission_total", 0.0)),
        (30, "Hopper Homes Commission", revenue_totals.get("hopper", {}).get("commission_total", 0.0)),
    ]
    for row_index, label, value in expense_rows:
        _style_range(
            owner_sheet,
            f"B{row_index}:K{row_index}",
            value=label,
            fill_color="FFFFFFFF",
            font=body_font,
            alignment=center,
            border=THIN_BLACK_BORDER,
        )
        _style_range(
            owner_sheet,
            f"L{row_index}:M{row_index}",
            value=value,
            fill_color="FFFFFFFF",
            font=body_value_font,
            alignment=center,
            border=THIN_BLACK_BORDER,
            number_format=OWNER_STATEMENT_VALUE_NUMBER_FORMAT,
        )

    _style_range(
        owner_sheet,
        "B32:K32",
        value="TOTAL EXPENSES",
        font=section_font,
        alignment=center,
    )
    _style_range(
        owner_sheet,
        "L32:M32",
        value="=SUM(L22:L30)",
        fill_color=OWNER_STATEMENT_BAND_FILL,
        font=section_font,
        alignment=center,
        border=THIN_BLACK_BORDER,
        number_format=OWNER_STATEMENT_LABEL_NUMBER_FORMAT,
    )
    _style_range(
        owner_sheet,
        "B34:K34",
        value="NET INCOME/LOSS",
        font=section_font,
        alignment=center,
    )
    _style_range(
        owner_sheet,
        "L34:M34",
        value="=L19-L32",
        fill_color=OWNER_STATEMENT_BAND_FILL,
        font=section_font,
        alignment=center,
        border=THIN_BLACK_BORDER,
        number_format=OWNER_STATEMENT_LABEL_NUMBER_FORMAT,
    )

    return owner_sheet


def build_middlefork_bookkeeping_workbook(
    portfolio: Any,
    period: Any,
    uploads: Sequence[Any],
    expense_items: Sequence[Any],
    revenue_items: Optional[Sequence[Any]] = None,
) -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    report_period_start = infer_reporting_period_start(period, uploads, expense_items)
    month_label = logical_month_label(report_period_start)
    previous_month_label = _previous_month_label(report_period_start)
    revenue_rows_by_source = build_revenue_rows_by_source(revenue_items=revenue_items, uploads=uploads)
    expense_totals = aggregate_expense_totals(expense_items)
    management_fee_rate = float(portfolio.management_fee_percentage or 0) / 100.0

    middlefork_titles = {
        "airbnb": f"Airbnb - {month_label}"[:31],
        "airbnb_tax": f"Airbnb Tax Report - {month_label}"[:31],
        "vrbo": f"Vrbo Report - {month_label}"[:31],
        "direct_bookings": f"Direct Bookings Report - {month_label}"[:31],
        "booking_com": f"Booking.com - {month_label}"[:31],
        "hopper": f"Hopper Homes - {month_label}"[:31],
        "refunds": "Refunds",
    }

    upload_by_id = {
        getattr(upload, "bookkeeping_upload_id", None): upload
        for upload in uploads
        if getattr(upload, "bookkeeping_upload_id", None) is not None
    }

    revenue_refs: Dict[str, Dict[str, str]] = {}
    revenue_has_data: Dict[str, bool] = {}

    def merged_source_rows(source: str) -> List[Dict[str, Any]]:
        rows = revenue_rows_by_source.get(source, [])
        return [row for row in rows if _is_data_row_for_summary(source, row)]

    def set_row_values(sheet, row_index: int, headers: Sequence[str], row_values: Dict[str, Any], source_header_map: Optional[Dict[str, str]] = None, formulas: Optional[Dict[str, Any]] = None) -> None:
        header_map = source_header_map or {}
        formula_map = formulas or {}
        for column_index, header in enumerate(headers, start=1):
            if header in formula_map:
                value = formula_map[header](row_index)
            else:
                source_header = header_map.get(header, header)
                value = _coerce_revenue_cell_value(source_header, row_values.get(source_header))
            sheet.cell(row=row_index, column=column_index, value=value)

    airbnb_headers = [
        "Date",
        "Arriving by date",
        "Type",
        "Confirmation code",
        "Booking date",
        "Start date",
        "End date",
        "Nights",
        "Guest",
        "Listing",
        "Details",
        "Reference code",
        "Currency",
        "Amount",
        "Paid out",
        "Service fee",
        "Fast pay fee",
        "Cleaning fee",
        "Pet fee",
        "Gross earnings",
        "Occupancy taxes",
        "Earnings year",
    ]

    airbnb_sheet = workbook.create_sheet(title=middlefork_titles["airbnb"])
    airbnb_rows = merged_source_rows("airbnb")
    _write_sheet_headers(airbnb_sheet, airbnb_headers)
    current_row = 2
    for row in airbnb_rows:
        set_row_values(airbnb_sheet, current_row, airbnb_headers, row)
        current_row += 1
    airbnb_total_row = max(2, current_row)
    airbnb_sheet.cell(row=airbnb_total_row, column=15, value=f"=SUM(O2:O{max(2, airbnb_total_row - 1)})")
    _auto_width(airbnb_sheet, max_width=42)
    revenue_refs["airbnb"] = {"total": f"O{airbnb_total_row}"}
    revenue_has_data["airbnb"] = bool(airbnb_rows)

    airbnb_tax_sheet = workbook.create_sheet(title=middlefork_titles["airbnb_tax"])
    airbnb_tax_rows = [row for row in airbnb_rows if _string_or_none(row.get("Type")) == "Pass Through Tot"]
    _write_sheet_headers(airbnb_tax_sheet, airbnb_headers)
    current_row = 2
    for row in airbnb_tax_rows:
        set_row_values(airbnb_tax_sheet, current_row, airbnb_headers, row)
        current_row += 1
    airbnb_tax_total_row = max(2, current_row)
    airbnb_tax_sheet.cell(row=airbnb_tax_total_row, column=14, value=f"=SUM(N2:N{max(2, airbnb_tax_total_row - 1)})")
    _auto_width(airbnb_tax_sheet, max_width=42)
    revenue_refs["airbnb_tax"] = {"tax": f"N{airbnb_tax_total_row}"}

    def write_channel_sheet(
        source: str,
        headers: Sequence[str],
        table_name: str,
        title: str,
        footer_formulas: Dict[str, str],
        source_header_map: Optional[Dict[str, str]] = None,
        row_formula_map: Optional[Dict[str, Any]] = None,
        minimum_footer_row: int = 2,
    ) -> Any:
        sheet = workbook.create_sheet(title=title)
        rows = merged_source_rows(source)
        _write_sheet_headers(sheet, headers)
        current_row = 2
        for row in rows:
            set_row_values(sheet, current_row, headers, row, source_header_map=source_header_map, formulas=row_formula_map)
            current_row += 1
        while current_row < minimum_footer_row:
            current_row += 1
        footer_row = current_row
        for header, formula in footer_formulas.items():
            column_index = headers.index(header) + 1
            sheet.cell(row=footer_row, column=column_index, value=formula)
        _add_table(sheet, table_name, f"A1:{get_column_letter(len(headers))}{footer_row}")
        _auto_width(sheet, max_width=36)
        revenue_has_data[source] = bool(rows)
        return sheet, rows, footer_row

    vrbo_headers = [
        "Reservation ID",
        "Check-In Date",
        "Check-Out Date",
        "Property",
        "Guest name",
        "Base Rate",
        "Discount",
        "Cleaning Fee",
        "Lodging Tax",
        "Refund",
        "Total Price",
        "Hostaway application fee",
        "Stripe processing fees",
        "Vrbo Commission",
    ]
    vrbo_sheet, vrbo_rows, vrbo_footer_row = write_channel_sheet(
        "vrbo",
        vrbo_headers,
        "Vrbo",
        middlefork_titles["vrbo"],
        footer_formulas={
            "Base Rate": "=SUM(Vrbo[Base Rate])",
            "Discount": "=SUM(Vrbo[Discount])",
            "Cleaning Fee": "=SUM(Vrbo[Cleaning Fee])",
            "Lodging Tax": "=SUM(Vrbo[Lodging Tax])",
            "Refund": "=SUM(Vrbo[Refund])",
            "Total Price": "=SUM(Vrbo[Total Price])",
            "Hostaway application fee": "=SUM(Vrbo[Hostaway application fee])",
            "Stripe processing fees": "=SUM(Vrbo[Stripe processing fees])",
            "Vrbo Commission": "=SUM(Vrbo[Vrbo Commission])",
        },
        row_formula_map={
            "Total Price": lambda row_index: f"=SUM(F{row_index}:J{row_index})",
        },
        minimum_footer_row=3,
    )
    revenue_refs["vrbo"] = {
        "total": f"K{vrbo_footer_row}",
        "tax": f"I{vrbo_footer_row}",
        "commission": f"N{vrbo_footer_row}",
    }

    direct_headers = [
        "Reservation ID",
        "Check-In Date",
        "Check-Out Date",
        "Property",
        "Guest name",
        "Base Rate",
        "Discount",
        "Cleaning Fee",
        "Lodging Tax",
        "Refund",
        "Guest Channel Fee",
        "Total Price",
        "Hostaway application fee",
        "Stripe processing fees",
    ]
    direct_sheet, direct_rows, direct_footer_row = write_channel_sheet(
        "direct_bookings",
        direct_headers,
        "Direct_Bookings",
        middlefork_titles["direct_bookings"],
        footer_formulas={
            "Base Rate": "=SUM(Direct_Bookings[Base Rate])",
            "Discount": "=SUM(Direct_Bookings[Discount])",
            "Cleaning Fee": "=SUM(Direct_Bookings[Cleaning Fee])",
            "Lodging Tax": "=SUM(Direct_Bookings[Lodging Tax])",
            "Refund": "=SUM(Direct_Bookings[Refund])",
            "Guest Channel Fee": "=sum(Direct_Bookings[Guest Channel Fee])",
            "Total Price": "=SUM(Direct_Bookings[Total Price])",
            "Hostaway application fee": "=SUM(Direct_Bookings[Hostaway application fee])",
            "Stripe processing fees": "=SUM(Direct_Bookings[Stripe processing fees])",
        },
        row_formula_map={
            "Total Price": lambda row_index: "=sum(Direct_Bookings[[Base Rate]:[Guest Channel Fee]])",
        },
        minimum_footer_row=3,
    )
    revenue_refs["direct_bookings"] = {
        "total": f"L{direct_footer_row}",
        "tax": f"I{direct_footer_row}",
    }

    booking_headers = [
        "Reservation ID",
        "Check-In Date",
        "Check-Out Date",
        "Property",
        "Guest name",
        "Base Rate",
        "Discount",
        "Cleaning Fee",
        "Total Price",
        "Commission",
    ]
    booking_sheet, booking_rows, booking_footer_row = write_channel_sheet(
        "booking_com",
        booking_headers,
        "Booking.com",
        middlefork_titles["booking_com"],
        footer_formulas={
            "Base Rate": "=SUM(Booking.com[Base Rate])",
            "Discount": "=SUM(Booking.com[Discount])",
            "Cleaning Fee": "=SUM(Booking.com[Cleaning Fee])",
            "Total Price": "=sum(Booking.com[Total Price])",
            "Commission": "=SUM(Booking.com[Commission])",
        },
        source_header_map={
            "Reservation ID": "Reservation ID",
            "Check-In Date": "Check-In Date",
            "Check-Out Date": "Check-Out Date",
            "Property": "Property",
            "Guest name": "Guest name",
            "Base Rate": "Base Rate",
            "Discount": "Discount",
            "Cleaning Fee": "Cleaning Fee",
            "Total Price": "Total Price",
            "Commission": "Commission",
        },
        minimum_footer_row=3,
    )
    revenue_refs["booking_com"] = {
        "total": f"I{booking_footer_row}",
        "commission": f"J{booking_footer_row}",
    }

    hopper_headers = [
        "Reservation ID",
        "Check-In Date",
        "Check-Out Date",
        "Property",
        "Listing Number",
        "Guest name",
        "Base Rate",
        "Cleaning Fee",
        "Lodging Tax",
        "Hopper Homes Commission",
        "Total Price",
    ]
    hopper_sheet, hopper_rows, hopper_footer_row = write_channel_sheet(
        "hopper",
        hopper_headers,
        "Hopper_Homes",
        middlefork_titles["hopper"],
        footer_formulas={
            "Base Rate": "=sum(Hopper_Homes[Base Rate])",
            "Cleaning Fee": "=sum(Hopper_Homes[Cleaning Fee])",
            "Lodging Tax": "=sum(Hopper_Homes[Lodging Tax])",
            "Hopper Homes Commission": "=sum(Hopper_Homes[Hopper Homes Commission])",
            "Total Price": "=SUM(Hopper_Homes[Total Price])",
        },
        source_header_map={
            "Reservation ID": "Reservation ID",
            "Check-In Date": "Check-In Date",
            "Check-Out Date": "Check-Out Date",
            "Property": "Property",
            "Listing Number": "Listing Number",
            "Guest name": "Guest name",
            "Base Rate": "Base Rate",
            "Cleaning Fee": "Cleaning Fee",
            "Lodging Tax": "Lodging Tax",
            "Hopper Homes Commission": "Hopper Homes Commission",
            "Total Price": "Total Price",
        },
        row_formula_map={
            "Total Price": lambda row_index: f"=SUM(G{row_index}:J{row_index})",
        },
        minimum_footer_row=4,
    )
    if not hopper_rows:
        hopper_sheet["G4"] = "=sum(G2:G3)"
        hopper_sheet["H4"] = "=sum(H2:H3)"
        hopper_sheet["I4"] = "=sum(I2:I3)"
        hopper_sheet["J4"] = "=sum(J2:J3)"
        hopper_sheet["K4"] = "=SUM(K2:K3)"
    revenue_refs["hopper"] = {
        "total": f"K{hopper_footer_row}",
        "tax": f"I{hopper_footer_row}",
        "commission": f"J{hopper_footer_row}",
    }

    cleaning_sheet = workbook.create_sheet(title="Cleaning Expenses")
    cleaning_headers = ["Service Date", "Expense", "Vendor", "Amount", "Property", "Payment Method", "Account Holder", "Account Number", "Payment Date", "Receipt"]
    _write_sheet_headers(cleaning_sheet, cleaning_headers)
    cleaning_items = sorted(
        [item for item in expense_items if item.category == "cleaning"],
        key=lambda item: (
            item.service_date or date.max,
            item.property_code or "",
            item.payment_date or date.max,
            getattr(item, "upload_id", None) if getattr(item, "upload_id", None) is not None else 10**9,
            getattr(item, "bookkeeping_expense_item_id", 0) or 0,
        ),
    )
    row_index = 2
    for item in cleaning_items:
        values = [
            item.service_date,
            item.item_name or "Cleaning",
            _normalized_person_label(item.vendor) or item.vendor,
            item.effective_total(),
            item.property_code,
            _display_payment_method_label(item.payment_method),
            _normalized_person_label(item.account_holder) or item.account_holder,
            _venmo_handle_for_item(item) or item.account_number,
            _display_payment_date(item, upload_by_id),
            _expense_link_label(item),
        ]
        for column_index, value in enumerate(values, start=1):
            cleaning_sheet.cell(row=row_index, column=column_index, value=value)
        _set_link_cell(cleaning_sheet.cell(row=row_index, column=10), _expense_link_url(item, upload_by_id))
        if _row_needs_review(item):
            _apply_review_highlight(cleaning_sheet, row_index, len(cleaning_headers), item.review_reason)
        row_index += 1
    cleaning_total_row = max(2, row_index)
    cleaning_sheet.cell(row=cleaning_total_row, column=4, value=f"=SUM(D2:D{max(2, cleaning_total_row - 1)})")
    _auto_width(cleaning_sheet)

    maintenance_sheet = workbook.create_sheet(title="Maintenance Expenses")
    maintenance_headers = ["Service Date", "Expense", "Notes", "Vendor", "Amount", "Property", "Payment Method", "Account/Card Number", "Payment Date", "Receipt"]
    _write_sheet_headers(maintenance_sheet, maintenance_headers)
    maintenance_items = sorted([item for item in expense_items if item.category == "maintenance"], key=_expense_sort_key)
    row_index = 2
    if maintenance_items:
        for item in maintenance_items:
            values = [
                item.service_date,
                item.item_name or "Maintenance",
                item.details,
                _normalized_person_label(item.vendor) or item.vendor,
                item.effective_total(),
                item.property_code,
                _display_payment_method_label(item.payment_method),
                item.account_number,
                _display_payment_date(item, upload_by_id),
                _expense_link_label(item),
            ]
            for column_index, value in enumerate(values, start=1):
                maintenance_sheet.cell(row=row_index, column=column_index, value=value)
            _set_link_cell(maintenance_sheet.cell(row=row_index, column=10), _expense_link_url(item, upload_by_id))
            if _row_needs_review(item):
                _apply_review_highlight(maintenance_sheet, row_index, len(maintenance_headers), item.review_reason)
            row_index += 1
    maintenance_total_row = max(3, row_index)
    maintenance_sheet.cell(row=maintenance_total_row, column=5, value="=sum(E2)" if not maintenance_items else f"=SUM(E2:E{max(2, maintenance_total_row - 1)})")
    _auto_width(maintenance_sheet)

    supplies_sheet = workbook.create_sheet(title="Supplies")
    supplies_headers = [
        "Purchase/Order Date",
        "Purchase/Order Type",
        "Vendor",
        "Store/Site",
        "Item(s)",
        "Total",
        "Order Receipt",
        "Reimbursement Method",
        "Reimbursement Receipt",
    ]
    _write_sheet_headers(supplies_sheet, supplies_headers)
    grouped_supplies: Dict[Any, List[Any]] = defaultdict(list)
    for item in sorted([item for item in expense_items if item.category == "supplies"], key=_expense_sort_key):
        group_key = _item_extraction_field(item, "group_key") or item.upload_id or item.bookkeeping_expense_item_id
        grouped_supplies[group_key].append(item)

    row_index = 2
    for group_items in sorted(
        grouped_supplies.values(),
        key=lambda items: (
            items[0].service_date or items[0].payment_date or date.max,
            items[0].vendor or "",
            items[0].store_name or "",
        ),
    ):
        group_items = sorted(group_items, key=lambda item: (_item_extraction_field(item, "line_index", 0), item.bookkeeping_expense_item_id or 0))
        start_row = row_index
        end_row = row_index + len(group_items) - 1
        first_item = group_items[0]
        shared_total = _float_or_none(_item_extraction_field(first_item, "total"))
        group_total = round(shared_total if shared_total is not None else sum(item.effective_total() for item in group_items), 2)

        for offset, item in enumerate(group_items):
            current_row = start_row + offset
            has_reimbursement = bool(_item_extraction_field(item, "reimbursement_upload_id") or item.reimbursement_method)
            values = [
                item.service_date or item.payment_date if offset == 0 else None,
                _supply_purchase_type_label(item.purchase_type) if offset == 0 else None,
                _supply_vendor_display(item, upload_by_id) if offset == 0 else None,
                item.store_name if offset == 0 else None,
                _compact_supply_item_name(item.item_name),
                group_total if offset == 0 else None,
                _expense_link_label(item) if offset == 0 else None,
                (_display_payment_method_label(item.reimbursement_method) if has_reimbursement else "N/A") if offset == 0 else None,
                "Click/Tap Here" if offset == 0 and has_reimbursement else ("N/A" if offset == 0 else None),
            ]
            for column_index, value in enumerate(values, start=1):
                supplies_sheet.cell(row=current_row, column=column_index, value=value)
            if offset == 0:
                _set_link_cell(supplies_sheet.cell(row=current_row, column=7), _expense_link_url(item, upload_by_id))
                reimbursement_url = _reimbursement_link_url(item, upload_by_id)
                if reimbursement_url:
                    _set_link_cell(supplies_sheet.cell(row=current_row, column=9), reimbursement_url)
            if _row_needs_review(item):
                _apply_review_highlight(supplies_sheet, current_row, len(supplies_headers), item.review_reason)
        row_index = end_row + 1
    supplies_total_row = max(7, row_index)
    supplies_sheet.cell(row=supplies_total_row, column=6, value=f"=sum(F2:F{max(2, supplies_total_row - 1)})")
    _auto_width(supplies_sheet)

    misc_sheet = workbook.create_sheet(title="Misc. Expenses")
    misc_headers = ["Service Date", "Expense", "Vendor", "Amount", "Property", "Payment Method", "Account Holder", "Account Number", "Payment Date", "Receipt"]
    _write_sheet_headers(misc_sheet, misc_headers)
    misc_items = sorted(
        [item for item in expense_items if item.category == "misc"],
        key=lambda item: (
            item.service_date or date.max,
            item.property_code or "",
            item.payment_date or date.max,
            getattr(item, "upload_id", None) if getattr(item, "upload_id", None) is not None else 10**9,
            getattr(item, "bookkeeping_expense_item_id", 0) or 0,
        ),
    )
    row_index = 2
    for item in misc_items:
        values = [
            item.service_date,
            item.item_name or "Misc. Expense",
            _normalized_person_label(item.vendor) or item.vendor,
            item.effective_total(),
            item.property_code,
            _display_payment_method_label(item.payment_method),
            _normalized_person_label(item.account_holder) or item.account_holder,
            _venmo_handle_for_item(item) or item.account_number,
            _display_payment_date(item, upload_by_id),
            _expense_link_label(item),
        ]
        for column_index, value in enumerate(values, start=1):
            misc_sheet.cell(row=row_index, column=column_index, value=value)
        _set_link_cell(misc_sheet.cell(row=row_index, column=10), _expense_link_url(item, upload_by_id))
        if _row_needs_review(item):
            _apply_review_highlight(misc_sheet, row_index, len(misc_headers), item.review_reason)
        row_index += 1
    misc_total_row = max(3, row_index)
    misc_sheet.cell(row=misc_total_row, column=4, value="=SUM(D2)" if len(misc_items) == 1 else f"=SUM(D2:D{max(2, misc_total_row - 1)})")
    _auto_width(misc_sheet)

    software_sheet = workbook.create_sheet(title="Software Fee")
    software_sheet["A1"] = "Software Cost"
    software_sheet["F1"] = "Processing Fee Per Reservation"
    software_sheet["A2"] = "Software"
    software_sheet["B2"] = "Price Per Listing"
    software_sheet["C2"] = "No. of Listings"
    software_sheet["D2"] = "Total"
    software_sheet["F2"] = "Guest"
    software_sheet["G2"] = "Booking Platform"
    software_sheet["H2"] = "Stripe"
    software_sheet["I2"] = "Hostaway"
    for cell in ("A1", "F1", "A2", "B2", "C2", "D2", "F2", "G2", "H2", "I2"):
        software_sheet[cell].font = Font(bold=True)

    listing_count = portfolio.listing_count or 0
    hostaway_price = float(portfolio.hostaway_price_per_listing or 0)
    pricelabs_price = float(portfolio.pricelabs_price_per_listing or 0)
    software_sheet["A3"] = "Hostaway"
    software_sheet["B3"] = hostaway_price
    software_sheet["C3"] = listing_count
    software_sheet["D3"] = "=B3*C3"
    software_sheet["A4"] = "Pricelabs"
    software_sheet["B4"] = pricelabs_price
    software_sheet["C4"] = listing_count
    software_sheet["D4"] = "=B4*C4"
    software_sheet["A5"] = "TOTAL"
    software_sheet["D5"] = "=sum(D3:D4)"
    software_sheet["A5"].font = Font(bold=True)
    software_sheet["D5"].font = Font(bold=True)

    processing_fee_rows = []
    for upload in uploads:
        if getattr(upload, "stage", None) != "revenue":
            continue
        processing_fee_rows.extend((normalized_revenue_upload_payload(upload)["summary"] or {}).get("processing_fee_rows", []))
    processing_fee_rows.sort(
        key=lambda row: (
            0 if row.get("booking_platform") == "Direct Booking" else 1 if row.get("booking_platform") == "Vrbo" else 2,
        )
    )

    fee_row = 3
    for processing_row in processing_fee_rows:
        software_sheet.cell(row=fee_row, column=6, value=processing_row.get("guest"))
        software_sheet.cell(row=fee_row, column=7, value=processing_row.get("booking_platform"))
        software_sheet.cell(row=fee_row, column=8, value=processing_row.get("stripe"))
        software_sheet.cell(row=fee_row, column=9, value=processing_row.get("hostaway"))
        fee_row += 1
    processing_total_row = max(12, fee_row)
    software_sheet.cell(row=processing_total_row, column=6, value="TOTAL")
    software_sheet.cell(row=processing_total_row, column=8, value=f"=sum(H3:H{processing_total_row - 1})")
    software_sheet.cell(row=processing_total_row, column=9, value=f"=sum(I3:I{processing_total_row - 1})")
    software_sheet.cell(row=processing_total_row, column=6).font = Font(bold=True)
    software_sheet.cell(row=processing_total_row, column=8).font = Font(bold=True)
    software_sheet.cell(row=processing_total_row, column=9).font = Font(bold=True)
    software_sheet["A15"] = "TOTAL"
    software_sheet["A16"] = f"=sum(D5,H{processing_total_row},I{processing_total_row})"
    software_sheet["A15"].font = Font(bold=True)
    software_sheet["A16"].font = Font(bold=True)
    _auto_width(software_sheet)

    refunds_sheet = workbook.create_sheet(title=middlefork_titles["refunds"])
    refund_headers = ["Date", "Platform", "Type", "Listing", "Details", "Notes", "Receipt", "Amount"]
    _write_sheet_headers(refunds_sheet, refund_headers)
    refunds_sheet.cell(row=2, column=1, value=f"{portfolio.property_name or portfolio.name} - Refunds For {month_label}")
    refund_items = sorted([item for item in expense_items if item.category == "direct_refund"], key=_expense_sort_key)
    row_index = 3
    for item in refund_items:
        values = [
            item.service_date,
            _display_payment_method_label(item.payment_method) or "Manual",
            item.item_name or "Direct Refund",
            item.property_code,
            item.details or item.description,
            item.review_reason,
            _expense_link_label(item, fallback="-"),
            item.effective_total(),
        ]
        for column_index, value in enumerate(values, start=1):
            refunds_sheet.cell(row=row_index, column=column_index, value=value)
        _set_link_cell(refunds_sheet.cell(row=row_index, column=7), _expense_link_url(item, upload_by_id))
        row_index += 1
    refunds_total_row = max(4, row_index)
    refunds_sheet.cell(row=refunds_total_row, column=1, value="TOTAL")
    refunds_sheet.cell(row=refunds_total_row, column=8, value=f"=SUM(H2:H{max(3, refunds_total_row - 1)})")
    _auto_width(refunds_sheet)

    maintenance_owner_value = _sheet_ref(maintenance_sheet.title, f"E{maintenance_total_row}")
    booking_owner_total = 0.0 if not booking_rows else _sheet_ref(booking_sheet.title, revenue_refs["booking_com"]["total"])
    booking_owner_commission = 0.0 if not booking_rows else _sheet_ref(booking_sheet.title, revenue_refs["booking_com"]["commission"])
    hopper_owner_total = 0.0 if not hopper_rows else _sheet_ref(hopper_sheet.title, revenue_refs["hopper"]["total"])
    hopper_owner_tax = 0.0 if not hopper_rows else _sheet_ref(hopper_sheet.title, revenue_refs["hopper"]["tax"])
    hopper_owner_commission = 0.0 if not hopper_rows else _sheet_ref(hopper_sheet.title, revenue_refs["hopper"]["commission"])

    owner_sheet = workbook.create_sheet(title=f"Owner Statement - {month_label}"[:31], index=0)
    owner_rows = [
        [f"OWNER STATEMENT ({portfolio.property_name or portfolio.name})"],
        [],
        [None, f"For {month_label}"],
        [],
        [None, "PROPERTY NAME", None, None, None, None, "PROPERTY ADDRESS"],
        [None, portfolio.property_name or portfolio.name, None, None, None, None, portfolio.property_address or ""],
        [],
        [],
        [],
        [None, "Revenue Summary"],
        [None, "Airbnb", None, None, None, None, None, None, None, None, None, _sheet_ref(airbnb_sheet.title, revenue_refs["airbnb"]["total"])],
        [None, "Vrbo", None, None, None, None, None, None, None, None, None, _sheet_ref(vrbo_sheet.title, revenue_refs["vrbo"]["total"])],
        [None, "Direct Bookings", None, None, None, None, None, None, None, None, None, _sheet_ref(direct_sheet.title, revenue_refs["direct_bookings"]["total"])],
        [None, "Booking.com", None, None, None, None, None, None, None, None, None, booking_owner_total],
        [None, "Hopper Homes", None, None, None, None, None, None, None, None, None, hopper_owner_total],
        [],
        [None, "TOTAL REVENUE", None, None, None, None, None, None, None, None, None, "=SUM(L11:L15)"],
        [],
        [None, "Tax Summary"],
        [None, "Airbnb Taxes", None, None, None, None, None, None, None, None, None, _sheet_ref(airbnb_tax_sheet.title, revenue_refs["airbnb_tax"]["tax"])],
        [None, "Vrbo Taxes", None, None, None, None, None, None, None, None, None, _sheet_ref(vrbo_sheet.title, revenue_refs["vrbo"]["tax"])],
        [None, "Direct Bookings Taxes", None, None, None, None, None, None, None, None, None, _sheet_ref(direct_sheet.title, revenue_refs["direct_bookings"]["tax"])],
        [None, "Hopper Homes Taxes", None, None, None, None, None, None, None, None, None, hopper_owner_tax],
        [],
        [None, "TOTAL TAX", None, None, None, None, None, None, None, None, None, "=SUM(L20:L23)"],
        [],
        [None, "Expense Summary"],
        [None, "Cleaning Fees", None, None, None, None, None, None, None, None, None, _sheet_ref(cleaning_sheet.title, f"D{cleaning_total_row}")],
        [None, "Maintenance", None, None, None, None, None, None, None, None, None, maintenance_owner_value],
        [None, "Supplies", None, None, None, None, None, None, None, None, None, _sheet_ref(supplies_sheet.title, f"F{supplies_total_row}")],
        [None, "Software Fee", None, None, None, None, None, None, None, None, None, _sheet_ref(software_sheet.title, "A16")],
        [None, "Misc. Expenses", None, None, None, None, None, None, None, None, None, _sheet_ref(misc_sheet.title, f"D{misc_total_row}")],
        [None, "Direct Refunds", None, None, None, None, None, None, None, None, None, 0.0],
        [None, "Management Fee", None, None, None, None, None, None, None, None, None, "='Management Fee Summary'!E5"],
        [None, f"Vrbo Commissions - {previous_month_label}", None, None, None, None, None, None, None, None, None, _sheet_ref(vrbo_sheet.title, revenue_refs["vrbo"]["commission"])],
        [None, "Booking.com Commission", None, None, None, None, None, None, None, None, None, booking_owner_commission],
        [None, "Hopper Homes Commission", None, None, None, None, None, None, None, None, None, hopper_owner_commission],
        [],
        [None, "TOTAL EXPENSES", None, None, None, None, None, None, None, None, None, "=SUM(L28:M37)"],
        [],
        [None, "NET INCOME/LOSS", None, None, None, None, None, None, None, None, None, "=L17-L39"],
    ]
    for row_index, row in enumerate(owner_rows, start=1):
        for column_index, value in enumerate(row, start=1):
            owner_sheet.cell(row=row_index, column=column_index, value=value)
    owner_sheet["A1"].font = Font(size=16, bold=True)
    owner_sheet["B10"].font = Font(bold=True)
    owner_sheet["B19"].font = Font(bold=True)
    owner_sheet["B27"].font = Font(bold=True)
    owner_sheet["L17"].font = Font(bold=True)
    owner_sheet["L39"].font = Font(bold=True)
    owner_sheet["L41"].font = Font(bold=True)
    _auto_width(owner_sheet, max_width=38)

    management_sheet = workbook.create_sheet(title="Management Fee Summary", index=1)
    management_rows = [
        ["Expenses Covered by TRC Ventures", None, None, "Management Fee Summary", None],
        ["Expense", "Cost", None, "Total Revenue", _sheet_ref(owner_sheet.title, "L17")],
        ["Cleaning", _sheet_ref(cleaning_sheet.title, f"D{cleaning_total_row}"), None, "Total Cleaning Expense", _sheet_ref(owner_sheet.title, "L28")],
        ["Maintenance", 0.0 if not maintenance_items else _sheet_ref(maintenance_sheet.title, f"E{maintenance_total_row}"), None, "Revenue After Cleaning Expense", "=E2-E3"],
        ["Supplies", _sheet_ref(supplies_sheet.title, f"F{supplies_total_row}"), None, "Total Management Fee", f"=E4*{management_fee_rate}"],
        ["Misc. Expenses", _sheet_ref(misc_sheet.title, f"D{misc_total_row}"), None, "Expenses Covered by TRC Ventures", "=B15"],
        ["Hostaway Cost", _sheet_ref(software_sheet.title, "D3"), None, "Revenue Collected by TRC Ventures", "=B24"],
        ["Pricelabs Cost", _sheet_ref(software_sheet.title, "D4"), None, "Management Fee", "=E5+E6-E7"],
        ["Stripe Processing Fee", _sheet_ref(software_sheet.title, f"H{processing_total_row}"), None, None, None],
        ["Hostaway Processing Fee", _sheet_ref(software_sheet.title, f"I{processing_total_row}"), None, None, None],
        ["Direct Refunds", _sheet_ref(refunds_sheet.title, f"H{refunds_total_row}"), None, None, None],
        ["Vrbo Commission", _sheet_ref(vrbo_sheet.title, revenue_refs["vrbo"]["commission"]), None, None, None],
        ["Booking.com Commission", _sheet_ref(booking_sheet.title, revenue_refs["booking_com"]["commission"]), None, None, None],
        ["Hopper Homes Commission", _sheet_ref(hopper_sheet.title, revenue_refs["hopper"]["commission"]), None, None, None],
        ["Total", "=SUM(B3:B14)", None, None, None],
        [],
        ["Revenue Collected by TRC Ventures", None, None, None, None],
        ["Revenue Platform", "Amount", None, None, None],
        ["Vrbo", _sheet_ref(vrbo_sheet.title, revenue_refs["vrbo"]["total"]), None, None, None],
        ["Booking.com", 0.0 if not booking_rows else _sheet_ref(booking_sheet.title, revenue_refs["booking_com"]["total"]), None, None, None],
        ["Hopper Homes", _sheet_ref(hopper_sheet.title, revenue_refs["hopper"]["total"]), None, None, None],
        ["Hocking Hills", 0.0, None, None, None],
        ["Direct Bookings", _sheet_ref(direct_sheet.title, revenue_refs["direct_bookings"]["total"]), None, None, None],
        ["Total", "=SUM(B19:B23)", None, None, None],
    ]
    for row_index, row in enumerate(management_rows, start=1):
        for column_index, value in enumerate(row, start=1):
            management_sheet.cell(row=row_index, column=column_index, value=value)
    for cell in ("A1", "D1", "A2", "D2", "D17", "A18"):
        management_sheet[cell].font = Font(bold=True)
    _auto_width(management_sheet, max_width=34)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def build_bookkeeping_workbook(
    portfolio: Any,
    period: Any,
    uploads: Sequence[Any],
    expense_items: Sequence[Any],
    revenue_items: Optional[Sequence[Any]] = None,
) -> bytes:
    if normalize_property_token(getattr(portfolio, "code", None)) == "MIDDLEFORK":
        return build_middlefork_bookkeeping_workbook(portfolio, period, uploads, expense_items, revenue_items=revenue_items)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    report_period_start = infer_reporting_period_start(period, uploads, expense_items)
    month_label = logical_month_label(report_period_start)
    revenue_totals = aggregate_revenue_totals(uploads, revenue_items)
    expense_totals = aggregate_expense_totals(expense_items)
    management_fee_rate = float(portfolio.management_fee_percentage or 0) / 100.0
    configured_hostaway_total = float((portfolio.hostaway_price_per_listing or Decimal("0")) * (portfolio.listing_count or 0))
    configured_pricelabs_total = float((portfolio.pricelabs_price_per_listing or Decimal("0")) * (portfolio.listing_count or 0))
    upload_by_id = _upload_lookup(uploads)

    processing_fee_rows = []
    for source_totals in revenue_totals.values():
        processing_fee_rows.extend(source_totals.get("processing_fee_rows", []))
    processing_fee_rows.sort(
        key=lambda row: (
            0 if row.get("booking_platform") == "Vrbo" else 1 if row.get("booking_platform") == "Direct Booking" else 2,
        )
    )

    manual_software_items = [item for item in expense_items if item.category == "software_fee"]
    software_total_base = (
        configured_hostaway_total
        + configured_pricelabs_total
        + sum(item.effective_total() for item in manual_software_items)
        + sum(row["stripe"] + row["hostaway"] for row in processing_fee_rows)
    )
    direct_refund_total = expense_totals.get("direct_refund", 0.0)

    owner_sheet = _build_standard_owner_statement_sheet(
        workbook,
        portfolio,
        month_label,
        revenue_totals,
        expense_totals,
        direct_refund_total,
        management_fee_rate,
        software_total_base,
    )

    simple_expense_sheets = {
        "cleaning": ("Cleaning Expenses", ["Service Date", "Expense", "Vendor", "Amount", "Property", "Payment Method", "Account Holder", "Account Number", "Payment Date", "Receipt"]),
        "maintenance": ("Maintenance Expenses", ["Service Date", "Expense", "Vendor", "Amount", "Property", "Payment Method", "Account Holder", "Account Number", "Payment Date", "Receipt"]),
        "misc": ("Misc. Expenses", ["Service Date", "Expense", "Vendor", "Amount", "Property", "Payment Method", "Account Holder", "Account Number", "Payment Date", "Receipt"]),
    }
    for category, (title, headers) in simple_expense_sheets.items():
        sheet = workbook.create_sheet(title=title[:31])
        _write_sheet_headers(sheet, headers)
        category_items = sorted([item for item in expense_items if item.category == category], key=_expense_sort_key)
        row_index = 2
        for item in category_items:
            values = [
                item.service_date,
                item.item_name or category.replace("_", " ").title(),
                item.vendor,
                item.effective_total(),
                item.property_code,
                item.payment_method,
                item.account_holder,
                item.account_number,
                item.payment_date,
                _expense_link_label(item),
            ]
            for column_index, value in enumerate(values, start=1):
                sheet.cell(row=row_index, column=column_index, value=value)
            _set_link_cell(sheet.cell(row=row_index, column=10), _expense_link_url(item, upload_by_id))
            if _row_needs_review(item):
                _apply_review_highlight(sheet, row_index, len(headers), item.review_reason)
            row_index += 1
        if category_items:
            sheet.cell(row=row_index, column=4, value=f"=SUM(D2:D{row_index - 1})")
        _auto_width(sheet)

    supplies_sheet = workbook.create_sheet(title="Supplies")
    supplies_headers = [
        "Purchase/Order Date",
        "Purchase/Order Type",
        "Vendor",
        "Store/Site",
        "Item(s)",
        "Quantity",
        "Amount",
        "Sub Total",
        "Discount",
        "Shipping",
        "Tax",
        "Total",
        "Receipts",
        "Payment/Reimbursement Method",
        "Account Holder",
        "Account Number",
        "Reimbursement Receipt",
        "Reimbursement Date",
    ]
    _write_sheet_headers(supplies_sheet, supplies_headers)
    grouped_supplies: Dict[Any, List[Any]] = defaultdict(list)
    for item in sorted([item for item in expense_items if item.category == "supplies"], key=_expense_sort_key):
        group_key = _item_extraction_field(item, "group_key") or item.upload_id or item.bookkeeping_expense_item_id
        grouped_supplies[group_key].append(item)

    row_index = 2
    supply_groups = sorted(
        grouped_supplies.values(),
        key=lambda items: (
            items[0].service_date or items[0].payment_date or date.max,
            items[0].vendor or "",
            items[0].store_name or "",
        ),
    )
    for group_items in supply_groups:
        group_items = sorted(
            group_items,
            key=lambda item: (_item_extraction_field(item, "line_index", 0), item.bookkeeping_expense_item_id or 0),
        )
        start_row = row_index
        end_row = row_index + len(group_items) - 1
        first_item = group_items[0]
        shared_subtotal = _float_or_none(_item_extraction_field(first_item, "subtotal"))
        shared_discount = _float_or_none(_item_extraction_field(first_item, "discount"))
        shared_shipping = _float_or_none(_item_extraction_field(first_item, "shipping"))
        shared_tax = _float_or_none(_item_extraction_field(first_item, "tax"))
        shared_total = _float_or_none(_item_extraction_field(first_item, "total"))

        for offset, item in enumerate(group_items):
            current_row = start_row + offset
            line_amount = item.amount if item.amount is not None else item.effective_total()
            if item.unit_amount is not None and item.quantity is not None:
                line_amount = f"=F{current_row}*{float(item.unit_amount):g}"

            values = [
                item.service_date if offset == 0 else None,
                item.purchase_type if offset == 0 else None,
                item.vendor if offset == 0 else None,
                item.store_name if offset == 0 else None,
                item.item_name,
                float(item.quantity) if item.quantity is not None else None,
                line_amount,
                f"=SUM(G{start_row}:G{end_row})" if offset == 0 else None,
                shared_discount if offset == 0 else None,
                shared_shipping if offset == 0 else None,
                shared_tax if offset == 0 else None,
                (f"=SUM(H{start_row}:K{start_row})" if start_row == end_row else f"=SUM(H{start_row}:K{start_row})") if offset == 0 and shared_total is None else (shared_total if offset == 0 else None),
                _expense_link_label(item) if offset == 0 else None,
                (item.reimbursement_method or item.payment_method) if offset == 0 else None,
                item.account_holder if offset == 0 else None,
                item.account_number if offset == 0 else None,
                "Click/Tap Here" if offset == 0 and _item_extraction_field(item, "reimbursement_upload_id") else None,
                item.reimbursement_date if offset == 0 else None,
            ]
            if offset == 0 and shared_subtotal is not None:
                values[7] = shared_subtotal if len(group_items) == 1 else f"=SUM(G{start_row}:G{end_row})"
            if offset == 0 and shared_total is None:
                values[11] = f"=SUM(H{start_row}:K{start_row})"
            for column_index, value in enumerate(values, start=1):
                supplies_sheet.cell(row=current_row, column=column_index, value=value)
            if offset == 0:
                _set_link_cell(supplies_sheet.cell(row=current_row, column=13), _expense_link_url(item, upload_by_id))
                reimbursement_url = _reimbursement_link_url(item, upload_by_id)
                if reimbursement_url:
                    _set_link_cell(supplies_sheet.cell(row=current_row, column=17), reimbursement_url)
            if _row_needs_review(item):
                _apply_review_highlight(supplies_sheet, current_row, len(supplies_headers), item.review_reason)
        row_index = end_row + 1
    _auto_width(supplies_sheet)

    software_sheet = workbook.create_sheet(title="Software Fee")
    software_sheet["A1"] = "Software Cost"
    software_sheet["F1"] = "Processing Fee Per Reservation"
    software_sheet["A2"] = "Software"
    software_sheet["B2"] = "Price Per Listing"
    software_sheet["C2"] = "No. of Listings"
    software_sheet["D2"] = "Total"
    software_sheet["F2"] = "Guest"
    software_sheet["G2"] = "Booking Platform"
    software_sheet["H2"] = "Listing"
    software_sheet["I2"] = "Stripe"
    software_sheet["J2"] = "Hostaway"
    for cell in ("A1", "F1", "A2", "B2", "C2", "D2", "F2", "G2", "H2", "I2", "J2"):
        software_sheet[cell].font = Font(bold=True)

    left_row = 3
    configured_rows = [
        ("Hostaway", float(portfolio.hostaway_price_per_listing or 0), portfolio.listing_count or 0, configured_hostaway_total),
        ("Pricelabs", float(portfolio.pricelabs_price_per_listing or 0), portfolio.listing_count or 0, configured_pricelabs_total),
    ]
    for software_name, price_per_listing, listing_count, total in configured_rows:
        if price_per_listing == 0 and total == 0 and software_name != "Hostaway":
            continue
        software_sheet.cell(row=left_row, column=1, value=software_name)
        software_sheet.cell(row=left_row, column=2, value=price_per_listing)
        software_sheet.cell(row=left_row, column=3, value=listing_count)
        software_sheet.cell(row=left_row, column=4, value=f"=B{left_row}*C{left_row}")
        left_row += 1

    for item in manual_software_items:
        software_sheet.cell(row=left_row, column=1, value=item.item_name or item.vendor or "Manual Software Fee")
        software_sheet.cell(row=left_row, column=4, value=item.effective_total())
        if _row_needs_review(item):
            _apply_review_highlight(software_sheet, left_row, 4, item.review_reason)
        left_row += 1

    software_total_row = left_row
    software_sheet.cell(row=left_row, column=1, value="TOTAL")
    if left_row > 3:
        software_sheet.cell(row=left_row, column=4, value=f"=SUM(D3:D{left_row - 1})")
    else:
        software_sheet.cell(row=left_row, column=4, value=0)
    software_sheet.cell(row=left_row, column=1).font = Font(bold=True)
    software_sheet.cell(row=left_row, column=4).font = Font(bold=True)

    fee_row = 3
    for processing_row in processing_fee_rows:
        software_sheet.cell(row=fee_row, column=6, value=processing_row.get("guest"))
        software_sheet.cell(row=fee_row, column=7, value=processing_row.get("booking_platform"))
        software_sheet.cell(row=fee_row, column=8, value=processing_row.get("listing"))
        software_sheet.cell(row=fee_row, column=9, value=processing_row.get("stripe"))
        software_sheet.cell(row=fee_row, column=10, value=processing_row.get("hostaway"))
        fee_row += 1
    processing_total_row = max(3, fee_row)
    if processing_fee_rows:
        software_sheet.cell(row=processing_total_row, column=6, value="TOTAL")
        software_sheet.cell(row=processing_total_row, column=9, value=f"=SUM(I3:I{fee_row - 1})")
        software_sheet.cell(row=processing_total_row, column=10, value=f"=SUM(J2:J{fee_row - 1})")
        software_sheet.cell(row=processing_total_row, column=6).font = Font(bold=True)
        software_sheet.cell(row=processing_total_row, column=9).font = Font(bold=True)
        software_sheet.cell(row=processing_total_row, column=10).font = Font(bold=True)

    grand_total_label_row = max(software_total_row, processing_total_row) + 2
    grand_total_formula_row = grand_total_label_row + 1
    software_sheet.cell(row=grand_total_label_row, column=1, value="TOTAL")
    if processing_fee_rows:
        software_sheet.cell(
            row=grand_total_formula_row,
            column=1,
            value=f"=sum(D{software_total_row},I{processing_total_row},J{processing_total_row})",
        )
    else:
        software_sheet.cell(row=grand_total_formula_row, column=1, value=f"=D{software_total_row}")
    _auto_width(software_sheet)

    revenue_rows_by_source = build_revenue_rows_by_source(revenue_items=revenue_items, uploads=uploads)
    refund_sheet = workbook.create_sheet(title="Direct Refund")
    refund_headers = ["Date", "Platform", "Type", "Listing", "Details", "Amount", "Refund Receipt"]
    _write_sheet_headers(refund_sheet, refund_headers)
    refund_sheet.cell(row=2, column=1, value=f"{portfolio.name or portfolio.code} - Refunds for {month_label}")
    row_index = 3
    refund_items = sorted([item for item in expense_items if item.category == "direct_refund"], key=_expense_sort_key)
    for item in refund_items:
        values = [
            item.service_date,
            item.payment_method or "Manual",
            item.item_name or "Direct Refund",
            item.property_code,
            item.details or item.description,
            item.effective_total(),
            _expense_link_label(item, fallback="-"),
        ]
        for column_index, value in enumerate(values, start=1):
            refund_sheet.cell(row=row_index, column=column_index, value=value)
        _set_link_cell(refund_sheet.cell(row=row_index, column=7), _expense_link_url(item, upload_by_id))
        if _row_needs_review(item):
            _apply_review_highlight(refund_sheet, row_index, len(refund_headers), item.review_reason)
        row_index += 1
    refund_total_row = max(4, row_index)
    refund_sheet.cell(row=refund_total_row, column=1, value="TOTAL")
    refund_sheet.cell(row=refund_total_row, column=6, value=f"=SUM(F2:F{max(3, refund_total_row - 1)})")
    refund_sheet.cell(row=refund_total_row, column=7, value="-")
    _auto_width(refund_sheet)

    for source in REVENUE_SOURCES:
        sheet = workbook.create_sheet(title=sheet_title_for_source(source, report_period_start))
        rows = revenue_rows_by_source.get(source, [])
        _write_standardized_revenue_sheet(sheet, source, rows)
        _auto_width(sheet)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
